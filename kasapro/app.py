# -*- coding: utf-8 -*-
"""KasaPro v3 - Uygulama giriÅŸ noktasÄ± (App)

Bu modÃ¼l; kullanÄ±cÄ± giriÅŸi, ÅŸirket seÃ§imi ve ana UI'Ä± baÅŸlatÄ±r.
"""

from __future__ import annotations

import logging
import os
import shutil
import sqlite3
import sys
import threading
from datetime import datetime, date, time, timedelta
from typing import Any, Optional, List, Dict, Tuple

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

from .config import APP_TITLE, HAS_OPENPYXL, APP_BASE_DIR, DB_FILENAME
from .utils import _safe_slug, fmt_amount
from .db.main_db import DB
from .db.users_db import UsersDB
from .services import Services
from .ui.navigation import ScreenRegistry
from .ui.style import apply_modern_style
from .ui.ui_logging import log_ui_event, wrap_callback
from .ui.windows import LoginWindow, SettingsWindow, HelpWindow, ImportWizard
from .ui.frames import (
    KasaFrame,
    TanimlarHubFrame,
    RaporAraclarHubFrame,
    KullanicilarFrame,
    MessagesFrame,
    IntegrationsHubFrame,
    CreateCenterFrame,
    StockWmsFrame,
)
from .ui.plugins.loader import discover_ui_plugins
from .modules.notes_reminders.scheduler import ReminderScheduler
from .modules.integrations.worker import IntegrationWorker

# Import HRContext for typing
try:
    import sys
    import os as _os
    _modules_path = _os.path.abspath(_os.path.join(_os.path.dirname(__file__), "..", "modules"))
    if _modules_path not in sys.path:
        sys.path.insert(0, _modules_path)
    from hr.service import HRContext
except ImportError:
    # Fallback if HR module not available
    from typing import Any as HRContext

class App:
    def __init__(self, root: tk.Tk | None = None):
        """Uygulama baÅŸlat.
        
        Args:
            root: Tkinter root window. None ise otomatik oluÅŸtur.
        """
        if root is None:
            import tkinter as tk
            root = tk.Tk()
            root.title("KasaPro v3")
            root.geometry("1400x900")
            root.minsize(800, 600)
        
        self.root = root
        
        # Attributes ilk olarak tanÄ±mla
        self._test_mode = False
        self.db = None
        self.logger = logging.getLogger(__name__)
        self.user_id = None
        self.company_id = None
        self.current_user = None
        self.current_company = None
        self.status_bar = None  # âœ… Attribute tanÄ±mla
        
        # Initialization Ã§aÄŸÄ±r
        self._init_window()
        self._init_db()
        self._init_ui()
        self._start_services()
    
    def _init_window(self) -> None:
        """Pencere Ã¶zelliklerini ayarla."""
        try:
            self.root.title("KasaPro v3 - Muhasebe & Ä°ÅŸ YÃ¶netim")
            self.root.geometry("1600x1000")
            self.root.minsize(1024, 768)
            
            # âœ… Tema ayarla
            self._setup_theme()
            
            # âœ… Pencere ikonu (varsa)
            self._setup_icon()
            
            # âœ… VarsayÄ±lan font ayarlarÄ±
            self._setup_fonts()
            
        except Exception as e:
            self.logger.exception("Pencere kuruluÅŸu baÅŸarÄ±sÄ±z")
            raise
    
    def _setup_theme(self) -> None:
        """Uygulama temasÄ±nÄ± ayarla."""
        import tkinter.font as tkFont
        
        # Koyu tema (modern, profesyonel)
        style = ttk.Style()
        style.theme_use('clam')
        
        # Renkler
        bg_primary = "#1e1e1e"      # Koyu arka plan
        bg_secondary = "#2d2d2d"    # Ä°kincil arka plan
        fg_primary = "#ffffff"      # Ana metin
        fg_secondary = "#cccccc"    # Ä°kincil metin
        accent = "#007acc"          # Vurgulama rengi
        
        style.configure('TFrame', background=bg_primary, foreground=fg_primary)
        style.configure('TLabel', background=bg_primary, foreground=fg_primary)
        style.configure('TButton', background=bg_secondary, foreground=fg_primary)
        style.map('TButton',
            background=[('active', accent)],
            foreground=[('active', fg_primary)]
        )
        style.configure('TNotebook', background=bg_primary, foreground=fg_primary)
        style.configure('TNotebook.Tab', background=bg_secondary, foreground=fg_primary)
        style.map('TNotebook.Tab',
            background=[('selected', accent)],
            foreground=[('selected', fg_primary)]
        )
        
        self.root.configure(bg=bg_primary)
    
    def _setup_icon(self) -> None:
        """Pencere ikonunu ayarla."""
        import os
        try:
            icon_path = os.path.join(
                os.path.dirname(__file__),
                'assets', 'icon.ico'
            )
            if os.path.exists(icon_path):
                self.root.iconbitmap(icon_path)
        except Exception:
            pass  # Icon yok, devam et
    
    def _setup_fonts(self) -> None:
        """Font ayarlarÄ±nÄ± tanÄ±mla."""
        import tkinter.font as tkFont
        
        self.font_heading = tkFont.Font(family="Segoe UI", size=14, weight="bold")
        self.font_normal = tkFont.Font(family="Segoe UI", size=10)
        self.font_small = tkFont.Font(family="Segoe UI", size=9)
        self.font_mono = tkFont.Font(family="Courier New", size=9)
    
    def _init_db(self) -> None:
        """VeritabanÄ± baÄŸlantÄ±sÄ±nÄ± aÃ§."""
        try:
            self.base_dir = APP_BASE_DIR
            self.usersdb = UsersDB(self.base_dir)

            self.user = self._login()
            if not self.user:
                try:
                    self.usersdb.close()
                except Exception:
                    pass
                self.root.destroy()
                raise SystemExit(0)

            # Aktif veri sahibi (kullanÄ±cÄ±) + aktif ÅŸirket (her ÅŸirketin kendi DB'si vardÄ±r)
            self.is_admin = (self.user["role"] == "admin")
            self.data_owner_username = str(self.user["username"])
            try:
                self.data_owner_user_id = int(self.user["id"])
            except Exception:
                self.data_owner_user_id = None

            # Åirket: son kullanÄ±lan (yoksa ilk)
            self.active_company = None
            try:
                self.active_company = self.usersdb.get_active_company_for_user(self.user)
            except Exception:
                self.active_company = None

            if self.active_company:
                db_path = self.usersdb.get_company_db_path(self.active_company)
                try:
                    self.active_company_id = int(self.active_company["id"])
                except Exception:
                    self.active_company_id = None
                try:
                    self.active_company_name = str(self.active_company["name"])
                except Exception:
                    self.active_company_name = ""
                try:
                    if self.data_owner_user_id and self.active_company_id:
                        self.usersdb.set_last_company_id(int(self.data_owner_user_id), int(self.active_company_id))
                except Exception:
                    pass
            else:
                # GÃ¼venlik: Ã§ok eski kurulumlarda ÅŸirket kaydÄ± yoksa user DB'si ile devam et
                db_path = self.usersdb.get_user_db_path(self.user)
                self.active_company_id = None
                self.active_company_name = "1. Åirket"

            self.db = DB(db_path)
            # Servis katmanÄ± (UI -> services -> DB)
            self.services = Services.build(self.db, self.usersdb, context_provider=self._hr_context)
            self.integrations_worker = IntegrationWorker(self.services.integrations)
            try:
                cname = getattr(self, "active_company_name", "")
                if cname:
                    self.db.log("Login", f"{self.user['username']} ({self.user['role']}) / {cname}")
                else:
                    self.db.log("Login", f"{self.user['username']} ({self.user['role']})")
            except Exception:
                pass
        except Exception as e:
            self.logger.exception("VeritabanÄ± hatasÄ±")
            messagebox.showerror("Hata", f"VeritabanÄ± hatasÄ±: {e}")
            raise
    
    def _init_ui(self) -> None:
        """UI bileÅŸenlerini oluÅŸtur."""
        try:
            # MenÃ¼ Ã§ubuÄŸu
            self._create_menu_bar()
            
            # Ana container
            main_container = ttk.Frame(self.root)
            main_container.pack(fill=tk.BOTH, expand=True)
            
            # Sidebar
            self._create_sidebar(main_container)
            
            # Content area
            self._create_content_area(main_container)
            
            # âœ… Status bar oluÅŸtur
            self.status_bar = StatusBar(self.root)
            
            if not self._test_mode:
                try:
                    self.root.deiconify()
                    self.root.lift()
                    self.root.after(80, self.root.lift)
                except Exception:
                    pass
        except Exception as e:
            self.logger.exception("UI kuruluÅŸu baÅŸarÄ±sÄ±z")
            raise
    
    def _create_menu_bar(self) -> None:
        """MenÃ¼ Ã§ubuÄŸunu oluÅŸtur."""
        try:
            menubar = tk.Menu(self.root)
            self.root.config(menu=menubar)
            
            # Dosya menÃ¼sÃ¼
            file_menu = tk.Menu(menubar, tearoff=0)
            menubar.add_cascade(label="ğŸ“ Dosya", menu=file_menu)
            file_menu.add_command(label="ğŸ¢ Åirket SeÃ§", command=self._select_company)
            file_menu.add_separator()
            file_menu.add_command(label="ğŸšª Ã‡Ä±kÄ±ÅŸ", command=self.root.quit)
            
            # DÃ¼zen menÃ¼sÃ¼
            edit_menu = tk.Menu(menubar, tearoff=0)
            menubar.add_cascade(label="âœï¸ DÃ¼zen", menu=edit_menu)
            edit_menu.add_command(label="âš™ï¸ Ayarlar", command=self._show_settings)
            
            # YardÄ±m menÃ¼sÃ¼
            help_menu = tk.Menu(menubar, tearoff=0)
            menubar.add_cascade(label="â“ YardÄ±m", menu=help_menu)
            help_menu.add_command(label="â„¹ï¸ HakkÄ±nda", command=self._show_about)
        except Exception as e:
            self.logger.exception("MenÃ¼ Ã§ubuÄŸu oluÅŸturma hatasÄ±")
    
    def _create_sidebar(self, parent: ttk.Frame) -> None:
        """Sol sidebar'Ä± oluÅŸtur."""
        try:
            sidebar = ttk.Frame(parent, width=250)
            sidebar.pack(side=tk.LEFT, fill=tk.Y, padx=10, pady=10)
            sidebar.pack_propagate(False)
            
            title = ttk.Label(
                sidebar,
                text="ğŸ¢ KasaPro",
                font=("Segoe UI", 12, "bold"),
            )
            title.pack(pady=10)
            
            self.nav_frame = ttk.Frame(sidebar)
            self.nav_frame.pack(fill=tk.BOTH, expand=True)
        except Exception as e:
            self.logger.exception("Sidebar oluÅŸturma hatasÄ±")
    
    def _create_content_area(self, parent: ttk.Frame) -> None:
        """SaÄŸ iÃ§erik alanÄ±nÄ± oluÅŸtur."""
        try:
            content = ttk.Frame(parent)
            content.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            
            self.notebook = ttk.Notebook(content)
            self.notebook.pack(fill=tk.BOTH, expand=True)
        except Exception as e:
            self.logger.exception("Content area oluÅŸturma hatasÄ±")
    
    def _select_company(self) -> None:
        """Åirket seÃ§me dialog'u."""
        try:
            pass  # Ä°ÅŸlevi uygula
        except Exception as e:
            self.logger.exception("Åirket seÃ§me hatasÄ±")
    
    def _show_settings(self) -> None:
        """Ayarlar penceresini aÃ§."""
        try:
            pass  # Ä°ÅŸlevi uygula
        except Exception as e:
            self.logger.exception("Ayarlar aÃ§ma hatasÄ±")
    
    def _show_about(self) -> None:
        """HakkÄ±nda penceresini aÃ§."""
        try:
            messagebox.showinfo(
                "HakkÄ±nda",
                "KasaPro v3\nMuhasebe & Ä°ÅŸ YÃ¶netim Sistemi"
            )
        except Exception as e:
            self.logger.exception("HakkÄ±nda gÃ¶sterme hatasÄ±")
    
    def _start_services(self) -> None:
        """Arka plan hizmetlerini baÅŸlat."""
        try:
            self._schedule_sales_order_summary()
            self._start_reminder_scheduler()
        except Exception as e:
            self.logger.warning(f"Hizmetler baÅŸlatÄ±lamadÄ±: {e}")
            # Non-critical, don't raise

    def _schedule_sales_order_summary(self) -> None:
        try:
            now = datetime.now()
            next_run = datetime.combine(now.date(), time(23, 59))
            if now >= next_run:
                self._run_sales_order_summary_if_needed()
                next_run = next_run + timedelta(days=1)
            delay_ms = int((next_run - now).total_seconds() * 1000)
            self.root.after(delay_ms, self._schedule_sales_order_summary)
        except Exception:
            pass

    def _run_sales_order_summary_if_needed(self) -> None:
        today = date.today().isoformat()
        try:
            last = self.db.get_setting("satis_siparis_gun_sonu")
            if last == today:
                return
        except Exception:
            pass

        db_path = self.db.path
        company_name = getattr(self, "active_company_name", "") or ""

        def worker():
            try:
                db = DB(db_path)
                summary = db.satis_siparis_acik_ozet(["AÃ§Ä±k", "HazÄ±rlanÄ±yor", "KÄ±smi Sevk"])
                db.set_setting("satis_siparis_gun_sonu", today)
                msg = (
                    f"GÃ¼n Sonu AÃ§Ä±k SipariÅŸ ({company_name or 'Åirket'}): "
                    f"Adet={summary['adet']} â€¢ Toplam={fmt_amount(summary['toplam'])}"
                )
                db.log("GÃ¼n Sonu", msg)
                db.close()
            except Exception:
                return

        threading.Thread(target=worker, daemon=True).start()

    def _install_exception_handlers(self) -> None:
        logger = logging.getLogger(__name__)

        def handle_exception(exc_type, exc, tb):
            logger.exception("Unhandled exception", exc_info=(exc_type, exc, tb))
            try:
                messagebox.showerror(APP_TITLE, f"Beklenmeyen hata:\n{exc}")
            except Exception:
                pass

        def handle_tk_exception(exc, val, tb):
            logger.exception("Tkinter callback exception", exc_info=(exc, val, tb))
            try:
                messagebox.showerror(APP_TITLE, f"ArayÃ¼z hatasÄ±:\n{val}")
            except Exception:
                pass

        sys.excepthook = handle_exception
        self.root.report_callback_exception = handle_tk_exception

    def _load_test_user(self) -> sqlite3.Row:
        user = self.usersdb.get_user_by_username("admin")
        if not user:
            raise RuntimeError("Test modu iÃ§in admin kullanÄ±cÄ± bulunamadÄ±.")
        return user

    def _login(self) -> Optional[sqlite3.Row]:
        # LoginWindow kullanÄ±cÄ± seÃ§imi + ÅŸifre doÄŸrulama yapar
        w = LoginWindow(self.root, self.usersdb)

        # âš ï¸ Ã–nemli:
        # Login penceresi kapanmadan user bilgisi set edilmez.
        # Bu yÃ¼zden burada beklemek gerekiyor; aksi halde uygulama hemen kapanÄ±r
        # ve "pencere hiÃ§ gelmiyor" gibi gÃ¶rÃ¼nÃ¼r.
        try:
            self.root.wait_window(w)
        except Exception:
            # Ã‡ok nadir: wait_window sorun Ã§Ä±karÄ±rsa gÃ¼venli bir manuel dÃ¶ngÃ¼
            try:
                while True:
                    if not w.winfo_exists():
                        break
                    self.root.update()
            except Exception:
                pass

        return w.user

    def _hr_context(self) -> HRContext:
        """HR modÃ¼lÃ¼ iÃ§in context bilgisi saÄŸlar."""
        try:
            company_id = getattr(self, "active_company_id", None) or 1
            username = getattr(self, "data_owner_username", "unknown")
            role = str(getattr(self, "user", {}).get("role", "user"))
            return HRContext(
                company_id=int(company_id),
                actor_username=username,
                actor_role=role
            )
        except Exception:
            # Fallback: minimal context
            return HRContext(company_id=1, actor_username="system", actor_role="user")

    def reload_settings(self):
        """Para birimi / Ã¶deme / kategori gibi ayar listelerini tÃ¼m ekranlarda yeniler."""
        try:
            for fr in (getattr(self, "frames", {}) or {}).values():
                if fr is not None and hasattr(fr, "reload_settings"):
                    try:
                        fr.reload_settings()  # type: ignore
                    except Exception:
                        pass
        except Exception:
            pass

    
    def _update_company_label(self):
        try:
            if hasattr(self, "lbl_company"):
                name = getattr(self, "active_company_name", "") or ""
                self.lbl_company.config(text=f"Åirket: {name}")
        except Exception:
            pass

    def update_messages_badge(self, count: int):
        try:
            btn = (getattr(self, "nav_buttons", {}) or {}).get("mesajlar")
            if not btn:
                return
            base = getattr(self, "_messages_nav_base_text", "ğŸ“¨ Mesajlar")
            text = f"{base} ({count})" if count and count > 0 else base
            btn.config(text=text)
        except Exception:
            pass

    def update_notes_reminders_badge(self, count: int):
        try:
            frame = (getattr(self, "frames", {}) or {}).get("rapor_araclar")
            if frame and hasattr(frame, "refresh_notes_reminders_badge"):
                frame.refresh_notes_reminders_badge(count)
        except Exception:
            pass

    def update_overdue_dashboard(self, count: int):
        try:
            frame = (getattr(self, "frames", {}) or {}).get("kasa")
            if frame and hasattr(frame, "update_overdue_reminders"):
                frame.update_overdue_reminders(count)
        except Exception:
            pass

    def show_toast(self, text: str, duration_ms: int = 3500):
        try:
            toast = tk.Toplevel(self.root)
            toast.overrideredirect(True)
            toast.configure(bg="#333333")
            label = tk.Label(toast, text=text, bg="#333333", fg="#ffffff", padx=12, pady=6)
            label.pack()
            toast.update_idletasks()
            x = self.root.winfo_x() + self.root.winfo_width() - toast.winfo_width() - 24
            y = self.root.winfo_y() + 60
            toast.geometry(f"+{x}+{y}")
            toast.after(duration_ms, toast.destroy)
        except Exception:
            pass

    def _start_message_polling(self):
        self._last_unread_count = None
        self._message_polling_active = True
        try:
            self._poll_messages()
        except Exception:
            pass

    def _start_reminder_scheduler(self):
        if getattr(self, "_test_mode", False):
            return
        try:
            self._reminder_scheduler = ReminderScheduler(self, self.services.notes_reminders)
            self._reminder_scheduler.start()
        except Exception:
            pass

    def _refresh_message_badge(self):
        uid = self.get_active_user_id()
        if not uid:
            return
        try:
            count = self.db.message_unread_count(uid)
            self.update_messages_badge(count)
            self._last_unread_count = count
        except Exception:
            pass

    def _poll_messages(self):
        uid = self.get_active_user_id()
        if uid:
            try:
                count = self.db.message_unread_count(uid)
                self.update_messages_badge(count)
                if self._last_unread_count is not None and count > self._last_unread_count:
                    self.show_toast("Yeni mesajÄ±nÄ±z var.")
                self._last_unread_count = count
            except Exception:
                pass
        try:
            if getattr(self, "_message_polling_active", False):
                self.root.after(15000, self._poll_messages)
        except Exception:
            pass

    def get_active_user_row(self) -> Optional[sqlite3.Row]:
        """Admin iÃ§in seÃ§ili veri sahibi, diÄŸerleri iÃ§in giriÅŸ yapan kullanÄ±cÄ±."""
        try:
            if self.is_admin and getattr(self, "data_owner_username", None):
                u = self.usersdb.get_user_by_username(str(self.data_owner_username))
                if u:
                    return u
        except Exception:
            pass
        return self.user

    def get_active_user_id(self) -> Optional[int]:
        u = self.get_active_user_row()
        try:
            return int(u["id"]) if u else None
        except Exception:
            return None

    def switch_company(self, company_id: int):
        """Aktif ÅŸirketi deÄŸiÅŸtirir (DB dosyasÄ± deÄŸiÅŸir)."""
        u = self.get_active_user_row()
        if not u:
            return
        c = self.usersdb.get_company_by_id(int(company_id))
        if (not c) or (int(c["user_id"]) != int(u["id"])):
            messagebox.showerror(APP_TITLE, "Åirket bulunamadÄ± veya yetkiniz yok.")
            return

        new_path = self.usersdb.get_company_db_path(c)

        try:
            self.db.close()
        except Exception:
            pass
        self.db = DB(new_path)
        self.services = Services.build(self.db, self.usersdb)
        try:
            if hasattr(self, "_reminder_scheduler") and self._reminder_scheduler:
                self._reminder_scheduler.service = self.services.notes_reminders
                self._reminder_scheduler.reset_context()
        except Exception:
            pass

        self.active_company = c
        try:
            self.active_company_id = int(c["id"])
        except Exception:
            self.active_company_id = None
        try:
            self.active_company_name = str(c["name"])
        except Exception:
            self.active_company_name = ""

        try:
            self.usersdb.set_last_company_id(int(u["id"]), int(self.active_company_id or 0))
        except Exception:
            pass

        self.reload_settings()

        try:
            self._update_company_label()
        except Exception:
            pass

        # EkranlarÄ± yenile (varsa)
        try:
            if "tanimlar" in self.frames and hasattr(self.frames["tanimlar"], "refresh"):
                self.frames["tanimlar"].refresh()  # type: ignore
        except Exception:
            pass
        try:
            if "kasa" in self.frames and hasattr(self.frames["kasa"], "refresh"):
                self.frames["kasa"].refresh()  # type: ignore
        except Exception:
            pass
        try:
            for key in ("cari_hareket_ekle", "cari_hareketler"):
                if key in self.frames:
                    fr = self.frames[key]
                    try:
                        if hasattr(fr, "reload_cari"):
                            fr.reload_cari()  # type: ignore
                    except Exception:
                        pass
                    try:
                        if hasattr(fr, "refresh"):
                            fr.refresh()  # type: ignore
                    except Exception:
                        pass
        except Exception:
            pass
        try:
            if "rapor_araclar" in self.frames and hasattr(self.frames["rapor_araclar"], "refresh"):
                self.frames["rapor_araclar"].refresh()  # type: ignore
        except Exception:
            pass
        try:
            if "sirketler" in self.frames and hasattr(self.frames["sirketler"], "refresh"):
                self.frames["sirketler"].refresh()  # type: ignore
        except Exception:
            pass

        try:
            self.db.log("Åirket", f"Aktif ÅŸirket: {self.active_company_name}")
        except Exception:
            pass

        try:
            self._last_unread_count = None
            self._refresh_message_badge()
        except Exception:
            pass

    def get_active_username(self) -> str:
        """Aktif veri sahibi kullanÄ±cÄ± adÄ± (admin iÃ§in seÃ§ili kullanÄ±cÄ±, diÄŸerleri iÃ§in kendi)."""
        try:
            if self.is_admin and getattr(self, "data_owner_username", None):
                return str(self.data_owner_username)
        except Exception:
            pass
        try:
            return str(self.user["username"])
        except Exception:
            return "user"

    def on_users_changed(self):
        """KullanÄ±cÄ± listesi deÄŸiÅŸtiyse (admin combobox) gÃ¼ncelle."""
        if not getattr(self, "is_admin", False):
            return
        if not hasattr(self, "cmb_data_owner"):
            return
        try:
            vals = self.usersdb.list_usernames()
        except Exception:
            vals = [self.get_active_username()]
        try:
            self.cmb_data_owner["values"] = vals
        except Exception:
            pass
        if self.get_active_username() not in vals:
            try:
                self.data_owner_username = str(self.user["username"])
            except Exception:
                self.data_owner_username = "admin"
            try:
                self.cmb_data_owner.set(self.data_owner_username)
            except Exception:
                pass
            try:
                self.switch_data_owner(self.data_owner_username)
            except Exception:
                pass

    def _on_data_owner_change(self):
        if not self.is_admin:
            return
        try:
            new_u = str(self.cmb_data_owner.get()).strip()
        except Exception:
            return
        if not new_u or new_u == self.get_active_username():
            return
        self.switch_data_owner(new_u)

    def switch_data_owner(self, username: str):
        """Admin iÃ§in aktif veri sahibini deÄŸiÅŸtir (baÅŸka kullanÄ±cÄ±nÄ±n DB'sini gÃ¶rÃ¼ntÃ¼le/dÃ¼zenle)."""
        if not self.is_admin:
            return
        try:
            urow = self.usersdb.get_user_by_username(username)
        except Exception:
            urow = None
        if not urow:
            messagebox.showerror(APP_TITLE, f"KullanÄ±cÄ± bulunamadÄ±: {username}")
            return

        # Aktif veri sahibi deÄŸiÅŸince aktif ÅŸirketi de o kullanÄ±cÄ±ya gÃ¶re seÃ§
        try:
            self.data_owner_user_id = int(urow["id"])
        except Exception:
            self.data_owner_user_id = None

        crow = None
        try:
            crow = self.usersdb.get_active_company_for_user(urow)
        except Exception:
            crow = None

        if crow:
            new_path = self.usersdb.get_company_db_path(crow)
            self.active_company = crow
            try:
                self.active_company_id = int(crow["id"])
            except Exception:
                self.active_company_id = None
            try:
                self.active_company_name = str(crow["name"])
            except Exception:
                self.active_company_name = ""
            try:
                if self.data_owner_user_id and self.active_company_id:
                    self.usersdb.set_last_company_id(int(self.data_owner_user_id), int(self.active_company_id))
            except Exception:
                pass

        try:
            self._last_unread_count = None
            self._refresh_message_badge()
        except Exception:
            pass
        else:
            # gÃ¼venlik
            new_path = self.usersdb.get_user_db_path(urow)
            self.active_company = None
            self.active_company_id = None
            self.active_company_name = "1. Åirket"

        try:
            self.db.close()
        except Exception:
            pass
        self.db = DB(new_path)
        self.services = Services.build(self.db, self.usersdb)
        self.data_owner_username = str(username)

        self.reload_settings()
        try:
            self._update_company_label()
        except Exception:
            pass
        try:
            if "sirketler" in self.frames and hasattr(self.frames["sirketler"], "refresh"):
                self.frames["sirketler"].refresh()  # type: ignore
        except Exception:
            pass

        # EkranlarÄ± yenile (varsa)
        try:
            if "tanimlar" in self.frames and hasattr(self.frames["tanimlar"], "refresh"):
                self.frames["tanimlar"].refresh()  # type: ignore
        except Exception:
            pass
        try:
            if "kasa" in self.frames and hasattr(self.frames["kasa"], "refresh"):
                self.frames["kasa"].refresh()  # type: ignore
        except Exception:
            pass
        try:
            for key in ("cari_hareket_ekle", "cari_hareketler"):
                if key in self.frames and hasattr(self.frames[key], "reload_cari"):
                    self.frames[key].reload_cari()  # type: ignore
        except Exception:
            pass
        try:
            if "rapor_araclar" in self.frames and hasattr(self.frames["rapor_araclar"], "refresh"):
                self.frames["rapor_araclar"].refresh()  # type: ignore
        except Exception:
            pass

    def _build_ui(self):
        # Modern tema + okunabilir fontlar
        try:
            self._ui_colors = apply_modern_style(self.root)
        except Exception:
            self._ui_colors = {}

        container = ttk.Frame(self.root, style="TFrame")
        container.pack(fill=tk.BOTH, expand=True)
        log_ui_event("container_created", view="root_container")

        # Sol menÃ¼
        nav = ttk.Frame(container, style="Sidebar.TFrame", width=270)
        nav.pack(side=tk.LEFT, fill=tk.Y)
        try:
            nav.pack_propagate(False)
        except Exception:
            pass

        # SaÄŸ iÃ§erik
        content = ttk.Frame(container, style="TFrame")
        content.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        # Ãœst bar (sayfa baÅŸlÄ±ÄŸÄ± + hÄ±zlÄ± aksiyon)
        topbar = ttk.Frame(content, style="Topbar.TFrame")
        topbar.pack(fill=tk.X, padx=12, pady=(12, 8))

        self.lbl_page_title = ttk.Label(topbar, text="Kasa", style="TopTitle.TLabel")
        self.lbl_page_title.pack(side=tk.LEFT, padx=(10, 10), pady=8)

        self.lbl_page_sub = ttk.Label(topbar, text="", style="TopSub.TLabel")
        self.lbl_page_sub.pack(side=tk.LEFT, pady=10)

        # HÄ±zlÄ± butonlar
        ttk.Button(
            topbar,
            text="â“",
            width=3,
            command=wrap_callback("open_help", self.open_help),
        ).pack(side=tk.RIGHT, padx=(6, 10), pady=8)
        ttk.Button(
            topbar,
            text="âš™ï¸",
            width=3,
            command=wrap_callback("open_settings", self.open_settings),
        ).pack(side=tk.RIGHT, padx=6, pady=8)

        # Ä°Ã§erik gÃ¶vdesi (ekranlar buraya gelecek)
        body = ttk.Frame(content, style="TFrame")
        body.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 12))
        self.screen_registry = ScreenRegistry(body, self)
        self.frames = self.screen_registry.frames

        # Status bar
        self.status_var = tk.StringVar(value="F1: YardÄ±m  â€¢  Ctrl+F: Global Arama  â€¢  Ã‡ift tÄ±k: DÃ¼zenle")
        self.status = ttk.Label(content, textvariable=self.status_var, style="Status.TLabel")
        self.status.pack(fill=tk.X, side=tk.BOTTOM)

        # Sol menÃ¼ Ã¼st bilgi
        hdr = ttk.Frame(nav, style="Sidebar.TFrame")
        hdr.pack(fill=tk.X, padx=14, pady=(14, 10))

        ttk.Label(hdr, text="KasaPro", style="SidebarTitle.TLabel").pack(anchor="w")
        ttk.Label(hdr, text=f"KullanÄ±cÄ±: {self.user['username']} ({self.user['role']})", style="SidebarSub.TLabel").pack(anchor="w", pady=(2, 0))
        self.lbl_company = ttk.Label(hdr, text=f"Åirket: {getattr(self,'active_company_name','')}", style="SidebarSub.TLabel")
        self.lbl_company.pack(anchor="w")

        if self.is_admin:
            try:
                vals = self.usersdb.list_usernames()
            except Exception:
                vals = [str(self.user["username"])]

            rowu = ttk.Frame(hdr, style="Sidebar.TFrame")
            rowu.pack(fill=tk.X, pady=(8, 0))

            ttk.Label(rowu, text="Veri Sahibi:", style="SidebarSub.TLabel").pack(side=tk.LEFT)
            self.cmb_data_owner = ttk.Combobox(rowu, state="readonly", width=16, values=vals)
            self.cmb_data_owner.pack(side=tk.LEFT, padx=(8, 0))

            try:
                self.cmb_data_owner.set(self.data_owner_username)
            except Exception:
                pass
            try:
                self.cmb_data_owner.bind("<<ComboboxSelected>>", lambda _e: self._on_data_owner_change())
            except Exception:
                pass

        ttk.Separator(nav, orient="horizontal").pack(fill=tk.X, padx=12, pady=(6, 10))

        # MenÃ¼ butonlarÄ±
        self.nav_buttons: Dict[str, ttk.Button] = {}

        menu = ttk.Frame(nav, style="Sidebar.TFrame")
        menu.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 8))

        def nav_btn(text, key):
            log_ui_event("menu_added", key=key, text=text)
            b = ttk.Button(
                menu,
                text=text,
                style="Sidebar.TButton",
                command=wrap_callback(f"nav:{key}", lambda k=key: self.show(k)),
            )
            b.pack(fill=tk.X, padx=4, pady=2)
            self.nav_buttons[key] = b
            return b

        def nav_section(title: str):
            """Sol menÃ¼de bÃ¶lÃ¼m baÅŸlÄ±ÄŸÄ± (TanÄ±mlar/Ä°ÅŸlemler/Rapor)."""
            try:
                ttk.Label(menu, text=title, style="SidebarSection.TLabel").pack(fill=tk.X, padx=4, pady=(10, 2))
                ttk.Separator(menu, orient="horizontal").pack(fill=tk.X, padx=10, pady=(0, 6))
                log_ui_event("menu_section_added", title=title)
            except Exception:
                ttk.Label(menu, text=title, style="SidebarSub.TLabel").pack(fill=tk.X, padx=8, pady=(10, 4))

        # UI eklentileri (Ã¶r: cari hareketler + cari hareket ekle)
        # Not: MaaÅŸ Eklentileri artÄ±k MaaÅŸ Takibi iÃ§ine taÅŸÄ±ndÄ±ÄŸÄ± iÃ§in menÃ¼/ekran olarak yÃ¼klenmez.
        self.ui_plugins = [p for p in discover_ui_plugins() if p.key != "maas_eklentileri"]
        self._plugin_titles = {p.key: p.page_title for p in self.ui_plugins}

        # "Ã‡alÄ±ÅŸanlar" gibi bazÄ± ekranlarÄ±, var olan bir plugin ekranÄ±nÄ±n sekmesine yÃ¶nlendirebiliriz.
        if not hasattr(self, "_nav_routes") or not isinstance(getattr(self, "_nav_routes", None), dict):
            self._nav_routes = {}
        # TanÄ±mlar ekranÄ± iÃ§inde alt sekme yÃ¶nlendirmeleri
        self._nav_routes["cariler"] = {"target": "tanimlar", "after": "hub_tab", "tab": "cariler"}
        self._nav_routes["calisanlar"] = {"target": "tanimlar", "after": "hub_tab", "tab": "calisanlar"}
        self._nav_routes["maas_meslekler"] = {"target": "tanimlar", "after": "hub_tab", "tab": "meslekler"}

        # Rapor & AraÃ§lar hub iÃ§ sekme yÃ¶nlendirmeleri (eski kÄ±sayollar uyumlu kalsÄ±n)
        self._nav_routes["raporlar"] = {"target": "rapor_araclar", "after": "hub_tab", "tab": "raporlar"}
        self._nav_routes["satis_raporlari"] = {"target": "rapor_araclar", "after": "hub_tab", "tab": "satis_raporlari"}
        self._nav_routes["search"] = {"target": "rapor_araclar", "after": "hub_tab", "tab": "search"}
        self._nav_routes["loglar"] = {"target": "rapor_araclar", "after": "hub_tab", "tab": "loglar"}
        self._nav_routes["notlar_hatirlatmalar"] = {"target": "rapor_araclar", "after": "hub_tab", "tab": "notlar_hatirlatmalar"}

        # Eski kÄ±sayol uyumluluÄŸu: MaaÅŸ Eklentileri -> MaaÅŸ Takibi / Bankada MaaÅŸ Bul
        self._nav_routes["maas_eklentileri"] = {"target": "maas_takibi", "after": "plugin_tab", "tab": "scan"}

        # ----------------
        # MenÃ¼ bÃ¶lÃ¼mleri
        # ----------------
        nav_section("ğŸ§¾ KAYIT OLUÅTUR")
        nav_btn("ğŸ§¾ KayÄ±t OluÅŸtur (Merkez)", "create_center")

        nav_section("ğŸ“š TANIMLAR & AYARLAR")
        nav_btn("ğŸ“š TanÄ±mlar", "tanimlar")

        # Åirket yÃ¶netimi sol menÃ¼den kaldÄ±rÄ±ldÄ±; âš™ï¸ Ayarlar > Åirketler sekmesinde.

        nav_section("ğŸ’³ Ä°ÅLEMLER")
        nav_btn("ğŸ¦ Kasa", "kasa")
        self._messages_nav_base_text = "ğŸ“¨ Mesajlar"
        nav_btn(self._messages_nav_base_text, "mesajlar")
        for p in self.ui_plugins:
            # Meslekler TanÄ±mlar altÄ±nda gÃ¶steriliyor
            if p.key == "maas_meslekler":
                continue
            # MaaÅŸ Eklentileri artÄ±k MaaÅŸ Takibi iÃ§ine taÅŸÄ±ndÄ±
            if p.key == "maas_eklentileri":
                continue
            nav_btn(p.nav_text, p.key)

        nav_section("ğŸ“¦ STOK & WMS")
        nav_btn("ğŸ“¦ Stok/WMS", "stok_wms")

        nav_section("ğŸ”Œ ENTEGRASYONLAR")
        nav_btn("ğŸ”Œ Entegrasyonlar", "entegrasyonlar")

        nav_section("ğŸ“ˆ RAPORLAR & DASHBOARD")
        nav_btn("ğŸ“ˆ Rapor & AraÃ§lar", "rapor_araclar")
        nav_btn("ğŸ’¹ SatÄ±ÅŸ RaporlarÄ±", "satis_raporlari")

        # KullanÄ±cÄ± yÃ¶netimi Ayarlar (âš™ï¸) iÃ§ine taÅŸÄ±ndÄ±.
        # Sol menÃ¼de ayrÄ± bir "KullanÄ±cÄ±lar" sayfasÄ± gÃ¶stermiyoruz.

        # Alt aksiyonlar
        actions = ttk.Frame(nav, style="Sidebar.TFrame")
        actions.pack(fill=tk.X, padx=8, pady=(0, 12), side=tk.BOTTOM)

        ttk.Separator(actions, orient="horizontal").pack(fill=tk.X, padx=4, pady=(4, 8))

        self.btn_import_excel = ttk.Button(
            actions,
            text="ğŸ“¥ Excel Ä°Ã§e Aktar",
            style="Primary.TButton",
            command=wrap_callback("import_excel", self.import_excel),
        )
        self.btn_import_excel.pack(fill=tk.X, padx=4, pady=2)

        self.btn_export_excel = ttk.Button(
            actions,
            text="ğŸ“¤ Excel Export",
            style="Secondary.TButton",
            command=wrap_callback("export_excel", self.export_excel),
        )
        self.btn_export_excel.pack(fill=tk.X, padx=4, pady=2)

        # openpyxl yoksa Excel import/export devre dÄ±ÅŸÄ±
        if not HAS_OPENPYXL:
            try:
                self.btn_import_excel.config(state="disabled")
                self.btn_export_excel.config(state="disabled")
            except Exception:
                pass

        # DB yedek/geri yÃ¼kle iÅŸlemleri artÄ±k âš™ï¸ Ayarlar > DB sekmelerine taÅŸÄ±ndÄ±.

        ttk.Button(
            actions,
            text="ğŸšª Ã‡Ä±kÄ±ÅŸ",
            style="Danger.TButton",
            command=wrap_callback("on_close", self.on_close),
        ).pack(fill=tk.X, padx=4, pady=(8, 2))

        # Ekranlar
        self.screen_registry.register("kasa", lambda parent, app: KasaFrame(parent, app), title="Kasa")
        self.screen_registry.register("create_center", lambda parent, app: CreateCenterFrame(parent, app), title="KayÄ±t OluÅŸtur (Merkez)")
        self.screen_registry.register("mesajlar", lambda parent, app: MessagesFrame(parent, app), title="Mesajlar")
        self.screen_registry.register("tanimlar", lambda parent, app: TanimlarHubFrame(parent, app), title="TanÄ±mlar")
        self.screen_registry.register("stok_wms", lambda parent, app: StockWmsFrame(parent, app), title="Stok/WMS")
        self.screen_registry.register(
            "rapor_araclar",
            lambda parent, app: RaporAraclarHubFrame(parent, app),
            title="Rapor & AraÃ§lar",
        )

        # Plugin ekranlarÄ±
        for p in getattr(self, "ui_plugins", []) or []:
            self.screen_registry.register(p.key, p.build, title=p.page_title)
            log_ui_event("plugin_ui_registered", key=p.key, title=p.page_title)

        if self.is_admin:
            self.screen_registry.register("kullanicilar", lambda parent, app: KullanicilarFrame(parent, app), title="KullanÄ±cÄ±lar")

        # Ä°lk yÃ¼klemeler
        try:
            if "tanimlar" in self.frames and hasattr(self.frames["tanimlar"], "refresh"):
                self.frames["tanimlar"].refresh()  # type: ignore
        except Exception:
            pass
        try:
            for key in ("cari_hareket_ekle", "cari_hareketler"):
                if key in self.frames and hasattr(self.frames[key], "reload_cari"):
                    self.frames[key].reload_cari()  # type: ignore
        except Exception:
            pass
        try:
            self.frames["kasa"].reload_cari_combo()  # type: ignore
            self.frames["kasa"].refresh()  # type: ignore
        except Exception:
            pass

        for f in self.frames.values():
            f.pack_forget()

        # KÄ±sayollar
        try:
            self.root.bind("<F1>", wrap_callback("shortcut_help", lambda _e: self.open_help()))
            log_ui_event("callback_bound", target="root", tk_event="<F1>", handler="open_help")
        except Exception:
            pass
        try:
            self.root.bind("<Control-f>", wrap_callback("shortcut_search", lambda _e: self.show("search")))
            log_ui_event("callback_bound", target="root", tk_event="<Control-f>", handler="show(search)")
        except Exception:
            pass

        # BaÅŸlangÄ±Ã§ ekranÄ±
        self.show("kasa")

        self.root.protocol("WM_DELETE_WINDOW", wrap_callback("on_close", self.on_close))
        log_ui_event("callback_bound", target="root", tk_event="WM_DELETE_WINDOW", handler="on_close")

    def _ui_on_show(self, key: str, active_nav_key: Optional[str] = None):
        """Ekran deÄŸiÅŸince sol menÃ¼ + baÅŸlÄ±k gibi UI parÃ§alarÄ±nÄ± gÃ¼nceller."""
        title_map = {
            "kasa": "Kasa",
            "create_center": "KayÄ±t OluÅŸtur (Merkez)",
            "tanimlar": "TanÄ±mlar",
            "cariler": "TanÄ±mlar",
            "rapor_araclar": "Rapor & AraÃ§lar",
            "raporlar": "Raporlar",
            "satis_raporlari": "SatÄ±ÅŸ RaporlarÄ±",
            "search": "Global Arama",
            "loglar": "Log",
            "kullanicilar": "KullanÄ±cÄ±lar",
            "mesajlar": "Mesajlar",
            "entegrasyonlar": "Entegrasyonlar",
            "stok_wms": "Stok/WMS",
        }

        # Plugin baÅŸlÄ±klarÄ±
        try:
            title_map.update(getattr(self, "_plugin_titles", {}) or {})
        except Exception:
            pass

        # Aktif menÃ¼ vurgusu
        active_key = active_nav_key or key
        btns = getattr(self, "nav_buttons", {}) or {}
        for k, b in btns.items():
            try:
                b.configure(style=("SidebarActive.TButton" if k == active_key else "Sidebar.TButton"))
            except Exception:
                pass

        # Ãœst baÅŸlÄ±k
        try:
            t = title_map.get(key, key)
            self.lbl_page_title.config(text=t)
        except Exception:
            pass

        # Alt bilgi (ÅŸirket + kullanÄ±cÄ±)
        try:
            cname = getattr(self, "active_company_name", "") or ""
            uname = self.get_active_username() if hasattr(self, "get_active_username") else (str(self.user["username"]) if self.user else "")
            sub = " â€¢ ".join([x for x in (cname, uname) if x])
            self.lbl_page_sub.config(text=sub)
        except Exception:
            pass


    def show(self, key: str):
        self.active_screen_key = key
        # BazÄ± menÃ¼ tuÅŸlarÄ±, var olan bir ekranÄ±n belirli sekmesine yÃ¶nlenebilir (Ã¶rn: Ã‡alÄ±ÅŸanlar -> MaaÅŸ Takibi/Ã‡alÄ±ÅŸanlar sekmesi)
        route = None
        try:
            route = (getattr(self, "_nav_routes", {}) or {}).get(key)
        except Exception:
            route = None

        target_key = key
        if isinstance(route, dict):
            try:
                target_key = str(route.get("target") or key)
            except Exception:
                target_key = key

        if target_key not in self.frames:
            log_ui_event("screen_missing", key=target_key, source=key)
        self.screen_registry.show(target_key)

        # Route sonrasÄ± aksiyonlar
        if isinstance(route, dict):
            after = str(route.get("after") or "")

            # Eski kÄ±sayol: Ã‡alÄ±ÅŸanlar -> MaaÅŸ Takibi/Ã‡alÄ±ÅŸanlar
            if after == "employees":
                try:
                    fr = self.frames.get(target_key)
                    if fr is not None and hasattr(fr, "select_employees_tab"):
                        fr.select_employees_tab()  # type: ignore
                except Exception:
                    pass

            # Yeni: TanÄ±mlar hub iÃ§ sekme yÃ¶nlendirmesi
            if after == "hub_tab":
                try:
                    fr = self.frames.get(target_key)
                    tab = str(route.get("tab") or "")
                    if fr is not None and hasattr(fr, "select_tab"):
                        fr.select_tab(tab)  # type: ignore
                except Exception:
                    pass

            # Plugin iÃ§ sekme yÃ¶nlendirmesi (Ã¶rn: MaaÅŸ Takibi gibi Notebook kullanan pluginler)
            if after == "plugin_tab":
                try:
                    fr = self.frames.get(target_key)
                    tab = str(route.get("tab") or "")
                    if fr is not None and hasattr(fr, "select_tab"):
                        fr.select_tab(tab)  # type: ignore
                except Exception:
                    pass

        try:
            # Sol menÃ¼de tek tuÅŸ altÄ±nda toplanan ekranlar iÃ§in (Ã¶rn: Rapor & AraÃ§lar),
            # baÅŸlÄ±k key'e gÃ¶re, menÃ¼ vurgusu ise target_key'e gÃ¶re olsun.
            self._ui_on_show(key, active_nav_key=(target_key if hasattr(self, 'nav_buttons') and target_key in getattr(self, 'nav_buttons', {}) else None))
        except Exception:
            pass

        if key in ("tanimlar", "cariler", "cari_hareket_ekle", "cari_hareketler", "rapor_araclar", "raporlar", "satis_raporlari", "search", "loglar", "kasa"):
            try:
                for k2 in ("cari_hareket_ekle", "cari_hareketler"):
                    if k2 in self.frames and hasattr(self.frames[k2], "reload_cari"):
                        self.frames[k2].reload_cari()  # type: ignore
            except Exception:
                pass
            try:
                self.frames["kasa"].reload_cari_combo()  # type: ignore
            except Exception:
                pass
            if key in ("rapor_araclar", "raporlar", "satis_raporlari", "search", "loglar"):
                try:
                    if "rapor_araclar" in self.frames and hasattr(self.frames["rapor_araclar"], "refresh"):
                        self.frames["rapor_araclar"].refresh()  # type: ignore
                except Exception:
                    pass
            if key == "mesajlar":
                try:
                    if "mesajlar" in self.frames and hasattr(self.frames["mesajlar"], "refresh"):
                        self.frames["mesajlar"].refresh()  # type: ignore
                except Exception:
                    pass



        if key == "kullanicilar":
            try:
                self.frames["kullanicilar"].refresh()  # type: ignore
            except Exception:
                pass
        if key == "create_center":
            try:
                self.frames["create_center"].refresh()  # type: ignore
            except Exception:
                pass

    def open_create_center(self, form_id: Optional[str] = None, context: Optional[Dict[str, Any]] = None) -> None:
        self.show("create_center")
        try:
            frame = self.frames.get("create_center")
        except Exception:
            frame = None
        if frame is not None and form_id:
            try:
                frame.select_form(form_id, context=context)  # type: ignore
                log_ui_event("create_center_open", form=form_id)
            except Exception:
                pass
    def open_settings(self):
        SettingsWindow(self)


    def open_help(self):
        HelpWindow(self)

    def backup_db(self):
        base = self.base_dir if hasattr(self, 'base_dir') else APP_BASE_DIR
        src = getattr(self.db, 'path', None) or os.path.join(base, DB_FILENAME)
        if not src or not os.path.exists(src):
            raise FileNotFoundError('DB bulunamadÄ±.')
        uname = self.get_active_username() if hasattr(self, 'get_active_username') else (str(self.user['username']) if self.user else 'user')
        cname = _safe_slug(getattr(self, "active_company_name", "") or "sirket")
        dst = os.path.join(base, f"kasa_backup_{uname}_{cname}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db")
        shutil.copy2(src, dst)
        try:
            self.db.log('Yedek', dst)
        except Exception:
            pass
        messagebox.showinfo(APP_TITLE, f"Yedek alÄ±ndÄ±:\n{dst}")

    def restore_db(self):
        p = filedialog.askopenfilename(title='DB Yedek SeÃ§', filetypes=[('DB', '*.db'), ('All', '*.*')])
        if not p:
            return
        uname = self.get_active_username() if hasattr(self, 'get_active_username') else (str(self.user['username']) if self.user else 'user')
        if not messagebox.askyesno(APP_TITLE, f"Geri yÃ¼kleme '{uname}' verisinin DB'sinin Ã¼zerine yazar. Devam?"):
            return
        dst = getattr(self.db, 'path', None)
        if not dst:
            messagebox.showerror(APP_TITLE, 'Hedef DB yolu bulunamadÄ±.')
            return
        try:
            self.db.close()
        except Exception:
            pass
        try:
            shutil.copy2(p, dst)
        except Exception as e:
            messagebox.showerror(APP_TITLE, f'Geri yÃ¼kleme baÅŸarÄ±sÄ±z: {e}')
            # DB'yi tekrar aÃ§mayÄ± deneyelim
            try:
                self.db = DB(dst)
            except Exception:
                pass
            return
        self.db = DB(dst)
        try:
            self.db.log('Restore', os.path.basename(p))
        except Exception:
            pass
        messagebox.showinfo(APP_TITLE, 'Geri yÃ¼klendi. UygulamayÄ± yeniden baÅŸlatman Ã¶nerilir.')

    def import_excel(self):
        p = filedialog.askopenfilename(title="Excel SeÃ§", filetypes=[("Excel", "*.xlsx *.xlsm"), ("All", "*.*")])
        if not p:
            return
        w = ImportWizard(self, p)
        self.root.wait_window(w)
        if w.result_counts:
            try:
                if "tanimlar" in self.frames and hasattr(self.frames["tanimlar"], "refresh"):
                    self.frames["tanimlar"].refresh()  # type: ignore
            except Exception:
                pass
            try:
                self.frames["kasa"].refresh()  # type: ignore
            except Exception:
                pass
            try:
                if "cari_hareketler" in self.frames and hasattr(self.frames["cari_hareketler"], "refresh"):
                    self.frames["cari_hareketler"].refresh()  # type: ignore
            except Exception:
                pass
            try:
                if "banka_hareketleri" in self.frames and hasattr(self.frames["banka_hareketleri"], "refresh"):
                    self.frames["banka_hareketleri"].refresh()  # type: ignore
            except Exception:
                pass

    def export_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except Exception:
            messagebox.showerror(APP_TITLE, "openpyxl kurulu deÄŸil. Kur: pip install openpyxl")
            return

        wb = Workbook()
        wb.remove(wb.active)

        def add_sheet(name: str, headers: List[str], rows: List[Tuple[Any, ...]]):
            ws = wb.create_sheet(title=name)
            ws.append(headers)
            for r in rows:
                ws.append(list(r))
            for i, h in enumerate(headers, start=1):
                ws.column_dimensions[get_column_letter(i)].width = min(45, max(12, len(str(h))+2))
            ws.freeze_panes = "A2"

        caris = self.db.cari_list()
        add_sheet("Cariler", ["id","ad","tur","telefon","notlar","acilis_bakiye","aktif"],
                  [(r["id"], r["ad"], r["tur"], r["telefon"], r["notlar"], r["acilis_bakiye"], r.get("aktif", 1) if hasattr(r, "get") else r["aktif"]) for r in caris])

        ch = self.db.cari_hareket_list()
        add_sheet("Cari_Hareket", ["id","tarih","cari","tip","tutar","para","aciklama","odeme","belge","etiket"],
                  [(r["id"], r["tarih"], r["cari_ad"], r["tip"], r["tutar"], r["para"], r["aciklama"], r["odeme"], r["belge"], r["etiket"]) for r in ch])

        kh = self.db.kasa_list()
        add_sheet("Kasa_Hareket", ["id","tarih","tip","tutar","para","odeme","kategori","cari","aciklama","belge","etiket"],
                  [(r["id"], r["tarih"], r["tip"], r["tutar"], r["para"], r["odeme"], r["kategori"], r["cari_ad"], r["aciklama"], r["belge"], r["etiket"]) for r in kh])

        # Banka
        try:
            bh = self.db.banka_list(limit=200000)
            add_sheet("Banka_Hareket", ["id","tarih","banka","hesap","tip","tutar","para","aciklama","referans","belge","etiket","bakiye"],
                      [(r["id"], r["tarih"], r["banka"], r["hesap"], r["tip"], r["tutar"], r["para"], r["aciklama"], r["referans"], r["belge"], r["etiket"], r["bakiye"]) for r in bh])
        except Exception:
            pass

        # Stok
        try:
            urunler = self.db.stok_urun_list()
            add_sheet(
                "Stok_Urun",
                ["id","kod","ad","kategori","birim","min_stok","kritik_stok","max_stok","raf","tedarikci_id","barkod","aktif","aciklama"],
                [
                    (
                        r["id"],
                        r["kod"],
                        r["ad"],
                        r["kategori"],
                        r["birim"],
                        r["min_stok"],
                        r["kritik_stok"],
                        r["max_stok"],
                        r["raf"],
                        r["tedarikci_id"],
                        r["barkod"],
                        r["aktif"],
                        r["aciklama"],
                    )
                    for r in urunler
                ],
            )
            loks = self.db.stok_lokasyon_list(only_active=False)
            add_sheet("Stok_Lokasyon", ["id","ad","aciklama","aktif"],
                      [(r["id"], r["ad"], r["aciklama"], r["aktif"]) for r in loks])
            partiler = self.db.stok_parti_list()
            add_sheet("Stok_Parti", ["id","urun_id","parti_no","skt","uretim_tarih","aciklama"],
                      [(r["id"], r["urun_id"], r["parti_no"], r["skt"], r["uretim_tarih"], r["aciklama"]) for r in partiler])
            hareketler = self.db.stok_hareket_list(limit=200000)
            add_sheet(
                "Stok_Hareket",
                ["id","tarih","urun_id","tip","miktar","birim","kaynak_lokasyon_id","hedef_lokasyon_id","parti_id","referans_tipi","referans_id","maliyet","aciklama"],
                [
                    (
                        r["id"],
                        r["tarih"],
                        r["urun_id"],
                        r["tip"],
                        r["miktar"],
                        r["birim"],
                        r["kaynak_lokasyon_id"],
                        r["hedef_lokasyon_id"],
                        r["parti_id"],
                        r["referans_tipi"],
                        r["referans_id"],
                        r["maliyet"],
                        r["aciklama"],
                    )
                    for r in hareketler
                ],
            )
        except Exception:
            pass

        ws = wb.create_sheet("Ozet")
        totals = self.db.kasa_toplam()
        ws["A1"] = "Kasa Ã–zet"
        ws["A3"] = "Gelir"
        ws["B3"] = totals["gelir"]
        ws["A4"] = "Gider"
        ws["B4"] = totals["gider"]
        ws["A5"] = "Net"
        ws["B5"] = totals["net"]

        base = APP_BASE_DIR
        out = os.path.join(base, f"kasa_pro_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb.save(out)
        self.db.log("Excel Export", out)
        messagebox.showinfo(APP_TITLE, f"Excel export:\n{out}")

    def on_close(self):
        try:
            self.db.log("Uygulama", "KapandÄ±")
        except Exception:
            pass
        try:
            if hasattr(self, "_reminder_scheduler") and self._reminder_scheduler:
                self._reminder_scheduler.stop()
        except Exception:
            pass
        try:
            self.db.close()
        except Exception:
            pass
        try:
            self.usersdb.close()
        except Exception:
            pass
        self.root.destroy()

    def run(self):
        self.root.mainloop()


def main() -> None:
    """Ana uygulamayÄ± baÅŸlat."""
    import logging
    
    logger = logging.getLogger(__name__)
    
    try:
        app = App()  # âœ… root otomatik oluÅŸturulur
        app.root.mainloop()
        
    except Exception as e:
        logger.exception("Uygulama baÅŸlatma hatasÄ±")
        raise

if __name__ == "__main__":
    main()
