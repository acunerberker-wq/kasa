# -*- coding: utf-8 -*-
"""KasaPro v3 - Cari Ekstre penceresi (Toplevel)."""

from __future__ import annotations

from datetime import date, timedelta
from typing import TYPE_CHECKING

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

from ...config import APP_TITLE, HAS_OPENPYXL, HAS_REPORTLAB
from ...utils import center_window, fmt_tr_date, fmt_amount, ensure_pdf_fonts
from ..widgets import LabeledEntry

if TYPE_CHECKING:
    from ...app import App

class CariEkstreWindow(tk.Toplevel):
    def __init__(self, app: "App", cari_id: int):
        super().__init__(app.root)
        self.app = app
        self.db = app.db
        self.services = getattr(app, "services", None)
        self.cari_service = getattr(self.services, "cari", None) if self.services else None
        self.cari_id = int(cari_id)

        cari = (self.cari_service.get(self.cari_id) if self.cari_service else self.db.cari_get(self.cari_id))
        self.title(f"Cari Ekstre: {cari['ad'] if cari else self.cari_id}")
        self.geometry("1180x720")
        self.resizable(True, True)
        self.transient(app.root)
        self.grab_set()

        self._build()
        self.refresh()
        center_window(self, app.root)

    def _build(self):
        top = ttk.LabelFrame(self, text="Filtre")
        top.pack(fill=tk.X, padx=10, pady=10)

        r1 = ttk.Frame(top)
        r1.pack(fill=tk.X, pady=6)
        self.f_from = LabeledEntry(r1, "Başlangıç:", 12)
        self.f_from.pack(side=tk.LEFT, padx=6)
        self.f_to = LabeledEntry(r1, "Bitiş:", 12)
        self.f_to.pack(side=tk.LEFT, padx=6)
        self.f_q = LabeledEntry(r1, "Ara:", 22)
        self.f_q.pack(side=tk.LEFT, padx=6)

        ttk.Button(r1, text="Son 30 gün", command=self.last30).pack(side=tk.LEFT, padx=6)
        ttk.Button(r1, text="Yenile", command=self.refresh).pack(side=tk.LEFT, padx=6)

        ttk.Separator(top, orient="horizontal").pack(fill=tk.X, padx=6, pady=8)
        r2 = ttk.Frame(top)
        r2.pack(fill=tk.X, pady=6)
        self.btn_pdf = ttk.Button(r2, text="📄 PDF Ekstre", command=self.export_pdf)
        self.btn_pdf.pack(side=tk.LEFT, padx=6)
        if not HAS_REPORTLAB:
            self.btn_pdf.config(state="disabled")

        self.btn_xlsx = ttk.Button(r2, text="📊 Excel Ekstre", command=self.export_excel)
        self.btn_xlsx.pack(side=tk.LEFT, padx=6)
        if not HAS_OPENPYXL:
            self.btn_xlsx.config(state="disabled")

        ttk.Button(r2, text="Kapat", command=self.destroy).pack(side=tk.RIGHT, padx=6)

        self.lbl_sum = ttk.Label(top, text="")
        self.lbl_sum.pack(anchor="w", padx=10, pady=(0,8))

        mid = ttk.LabelFrame(self, text="Hareketler")
        mid.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0,10))

        cols = ("tarih","tip","borc","alacak","para","odeme","belge","etiket","aciklama","bakiye")
        self.tree = ttk.Treeview(mid, columns=cols, show="headings", height=18)
        for c in cols:
            self.tree.heading(c, text=c.upper())

        self.tree.column("tarih", width=90)
        self.tree.column("tip", width=70)
        self.tree.column("borc", width=90, anchor="e")
        self.tree.column("alacak", width=90, anchor="e")
        self.tree.column("para", width=55, anchor="center")
        self.tree.column("odeme", width=110)
        self.tree.column("belge", width=90)
        self.tree.column("etiket", width=90)
        self.tree.column("aciklama", width=360)
        self.tree.column("bakiye", width=110, anchor="e")

        self.tree.pack(fill=tk.BOTH, expand=True, side=tk.LEFT, padx=(6,0), pady=6)

        scr = ttk.Scrollbar(mid, orient="vertical", command=self.tree.yview)
        scr.pack(fill=tk.Y, side=tk.RIGHT, padx=(0,6), pady=6)
        self.tree.configure(yscrollcommand=scr.set)

    def last30(self):
        d_to = date.today()
        d_from = d_to - timedelta(days=30)
        self.f_from.set(d_from.strftime("%d.%m.%Y"))
        self.f_to.set(d_to.strftime("%d.%m.%Y"))
        self.refresh()

    def refresh(self):
        for i in self.tree.get_children():
            self.tree.delete(i)
        data = (
            self.cari_service.ekstre(
                self.cari_id,
                date_from=self.f_from.get(),
                date_to=self.f_to.get(),
                q=self.f_q.get(),
            )
            if self.cari_service
            else self.db.cari_ekstre(
                self.cari_id,
                date_from=self.f_from.get(),
                date_to=self.f_to.get(),
                q=self.f_q.get(),
            )
        )

        self.lbl_sum.config(
            text=(
                f"Açılış: {fmt_amount(data['opening'])} | "
                f"Borç: {fmt_amount(data['total_borc'])} | "
                f"Alacak: {fmt_amount(data['total_alacak'])} | "
                f"Net: {fmt_amount(data['net_degisim'])} | "
                f"Kapanış: {fmt_amount(data['closing'])}"
            )
        )

        for r in data["rows"]:
            self.tree.insert("", tk.END, values=(
                fmt_tr_date(r["tarih"]),
                r["tip"],
                f"{fmt_amount(r['borc'])}" if r["borc"] else "",
                f"{fmt_amount(r['alacak'])}" if r["alacak"] else "",
                r["para"],
                r["odeme"],
                r["belge"],
                r["etiket"],
                r["aciklama"],
                f"{fmt_amount(r['bakiye'])}",
            ))
    def export_excel(self):
        p = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            title="Cari Ekstre Kaydet",
        )
        if not p:
            return

        data = (
            self.cari_service.ekstre(
                self.cari_id,
                date_from=self.f_from.get(),
                date_to=self.f_to.get(),
                q=self.f_q.get(),
            )
            if self.cari_service
            else self.db.cari_ekstre(
                self.cari_id,
                date_from=self.f_from.get(),
                date_to=self.f_to.get(),
                q=self.f_q.get(),
            )
        )

        try:
            if self.cari_service:
                self.cari_service.export_ekstre_excel(data, p)
            else:
                # Legacy fallback (direkt openpyxl kullanımı)
                from openpyxl import Workbook
                from openpyxl.utils import get_column_letter

                wb = Workbook()
                ws = wb.active
                ws.title = "Ekstre"

                ws.append(["Cari", data["cari_ad"]])
                ws.append(["Tarih Aralığı", f"{data['df'] or '-'} → {data['dt'] or '-'}"])
                ws.append(["Arama", data["q"] or "-"])
                ws.append([])
                ws.append(["Açılış", data["opening"]])
                ws.append(["Borç", data["total_borc"]])
                ws.append(["Alacak", data["total_alacak"]])
                ws.append(["Net", data["net_degisim"]])
                ws.append(["Kapanış", data["closing"]])
                ws.append([])

                headers = ["Tarih","Tip","Borç","Alacak","Para","Ödeme","Belge","Etiket","Açıklama","Bakiye"]
                ws.append(headers)

                for r in data["rows"]:
                    ws.append([
                        r["tarih"],
                        r["tip"],
                        r["borc"],
                        r["alacak"],
                        r["para"],
                        r["odeme"],
                        r["belge"],
                        r["etiket"],
                        r["aciklama"],
                        r["bakiye"],
                    ])

                for i, h in enumerate(headers, start=1):
                    ws.column_dimensions[get_column_letter(i)].width = min(45, max(12, len(str(h)) + 2))

                ws.freeze_panes = "A13"
                wb.save(p)

        except Exception as e:
            msg = str(e)
            if "openpyxl" in msg.lower() or "no module named" in msg.lower():
                messagebox.showerror(APP_TITLE, "openpyxl kurulu değil. Kur: pip install openpyxl")
            else:
                messagebox.showerror(APP_TITLE, f"Excel export hatası:\n{e}")
            return

        messagebox.showinfo(APP_TITLE, f"Excel ekstre kaydedildi:\n{p}")
    def export_pdf(self):
        p = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            title="Cari Ekstre PDF Kaydet",
        )
        if not p:
            return

        data = (
            self.cari_service.ekstre(
                self.cari_id,
                date_from=self.f_from.get(),
                date_to=self.f_to.get(),
                q=self.f_q.get(),
            )
            if self.cari_service
            else self.db.cari_ekstre(
                self.cari_id,
                date_from=self.f_from.get(),
                date_to=self.f_to.get(),
                q=self.f_q.get(),
            )
        )

        try:
            if self.cari_service:
                self.cari_service.export_ekstre_pdf(data, p)
            else:
                # Legacy fallback (direkt reportlab kullanımı)
                from reportlab.lib.pagesizes import A4
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.lib import colors

                doc = SimpleDocTemplate(p, pagesize=A4, leftMargin=28, rightMargin=28, topMargin=28, bottomMargin=28)
                font_reg, font_bold = ensure_pdf_fonts()

                styles = getSampleStyleSheet()
                for _k in ("Normal", "Title", "Heading1", "Heading2"):
                    if _k in styles:
                        styles[_k].fontName = (font_bold if _k != "Normal" else font_reg)

                story = []
                story.append(Paragraph(f"<b>Cari Ekstre</b> - {data['cari_ad']}", styles["Title"]))
                story.append(Spacer(1, 8))
                story.append(Paragraph(f"Tarih Aralığı: {data['df'] or '-'} → {data['dt'] or '-'}", styles["Normal"]))
                story.append(Paragraph(f"Arama: {data['q'] or '-'}", styles["Normal"]))
                story.append(Spacer(1, 10))
                story.append(Paragraph(
                    f"Açılış: {fmt_amount(data['opening'])} | Borç: {fmt_amount(data['total_borc'])} | "
                    f"Alacak: {fmt_amount(data['total_alacak'])} | Net: {fmt_amount(data['net_degisim'])} | "
                    f"Kapanış: {fmt_amount(data['closing'])}",
                    styles["Normal"],
                ))
                story.append(Spacer(1, 12))

                table_data = [["Tarih","Tip","Borç","Alacak","Para","Ödeme","Belge","Etiket","Açıklama","Bakiye"]]
                for r in data["rows"]:
                    table_data.append([
                        r["tarih"],
                        r["tip"],
                        f"{fmt_amount(r['borc'])}" if r["borc"] else "",
                        f"{fmt_amount(r['alacak'])}" if r["alacak"] else "",
                        r["para"],
                        (r["odeme"] or "")[:20],
                        (r["belge"] or "")[:16],
                        (r["etiket"] or "")[:16],
                        (r["aciklama"] or "")[:40],
                        f"{fmt_amount(r['bakiye'])}",
                    ])

                tbl = Table(table_data, repeatRows=1)
                tbl.setStyle(TableStyle([
                    ("FONTNAME", (0,0), (-1,0), font_bold),
                    ("FONTNAME", (0,1), (-1,-1), font_reg),
                    ("FONTSIZE", (0,0), (-1,0), 9),
                    ("FONTSIZE", (0,1), (-1,-1), 8),
                    ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
                    ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
                    ("ALIGN", (2,1), (3,-1), "RIGHT"),
                    ("ALIGN", (-1,1), (-1,-1), "RIGHT"),
                ]))

                story.append(tbl)
                doc.build(story)

        except Exception as e:
            msg = str(e)
            if "reportlab" in msg.lower() or "no module named" in msg.lower():
                messagebox.showerror(APP_TITLE, "PDF için reportlab yok. Kur: pip install reportlab")
            else:
                messagebox.showerror(APP_TITLE, f"PDF export hatası:\n{e}")
            return

        messagebox.showinfo(APP_TITLE, f"PDF ekstre kaydedildi:\n{p}")


# FRAMES
# =========================
