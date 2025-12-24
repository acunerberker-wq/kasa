# -*- coding: utf-8 -*-
"""Banka Excel Çalışma Alanı (Treeview tabanlı tablo editörü)

Amaç:
- Banka hareketlerini Excel benzeri bir tabloda görüntülemek
- Hücre bazında düzenlemek (çift tık ile)
- Değişiklikleri DB'ye kaydetmek
- Excel'e aktarmak (openpyxl)

Not:
- Bu ekran admin yetkisi olan kullanıcılar için düzenleme/kaydetme işlemlerini açar.
"""

from __future__ import annotations

from typing import Any, Dict, List, Optional, Tuple, TYPE_CHECKING

import json
import threading
import queue

import os
import time
import subprocess
import sys

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog

from ...config import APP_TITLE, HAS_OPENPYXL
from ...utils import safe_float, fmt_amount, parse_date_smart, center_window
from ...core.banka_macros import build_tag_suggestions, DEFAULT_TAG_RULES
from ...core.fuzzy import best_substring_similarity, amount_score, combine_scores
from .banka_analysis import BankaAnalizWindow

if TYPE_CHECKING:
    from ...app import App

_COLS: Tuple[str, ...] = (
    "id",
    "tarih",
    "banka",
    "hesap",
    "tip",
    "tutar",
    "para",
    "aciklama",
    "referans",
    "belge",
    "etiket",
    "bakiye",
)

_EDITABLE: set[str] = set(_COLS) - {"id"}


class BankaWorkspaceWindow(tk.Toplevel):
    def __init__(
        self,
        app: "App",
        *,
        ids: Optional[List[int]] = None,
        initial_filters: Optional[Dict[str, str]] = None,
        title_suffix: str = "",
    ):
        super().__init__(app.root)
        self.app = app
        self.db = app.db
        self.ids = [int(x) for x in (ids or []) if str(x).strip().lstrip("-").isdigit()]
        self.initial_filters = initial_filters or {}

        self.dirty: set[int] = set()
        self._editor: Optional[tk.Widget] = None
        self._editor_info: Optional[Tuple[str, str, int]] = None  # (iid, col, row_id)
        self._active_cell: Optional[Tuple[str, str]] = None  # (iid, col)
        self._last_excel_path: Optional[str] = None
        self._last_excel_mtime: float = 0.0

        self.title(f"{APP_TITLE} - Banka Çalışma Alanı{(' - ' + title_suffix) if title_suffix else ''}")
        self.geometry("1180x760")

        self._build()
        self._apply_permissions()
        self._sync_excel_buttons()

        # Başlangıç verisi
        self.refresh()

        center_window(self, self.app.root)

    # -----------------
    # UI
    # -----------------
    def _build(self):
        top = ttk.Frame(self)
        top.pack(fill=tk.X, padx=10, pady=(10, 6))

        ttk.Button(top, text="Yenile", command=self.refresh).pack(side=tk.LEFT)
        self.btn_save = ttk.Button(top, text="💾 Değişiklikleri Kaydet", command=self.save_changes)
        self.btn_save.pack(side=tk.LEFT, padx=6)

        self.btn_add = ttk.Button(top, text="➕ Yeni Satır", command=self.add_row)
        self.btn_add.pack(side=tk.LEFT, padx=6)

        self.btn_del = ttk.Button(top, text="🗑 Seçileni Sil", command=self.delete_selected)
        self.btn_del.pack(side=tk.LEFT, padx=6)

        ttk.Separator(top, orient="vertical").pack(side=tk.LEFT, fill=tk.Y, padx=10)

        ttk.Button(top, text="📤 Excel'e Aktar", command=self.export_excel).pack(side=tk.LEFT)
        self.btn_excel_open = ttk.Button(top, text="🟩 Excel'de Aç", command=self.open_in_excel)
        self.btn_excel_open.pack(side=tk.LEFT, padx=6)
        self.btn_excel_import = ttk.Button(top, text="🔄 Excel'den Güncelle", command=self.import_from_excel)
        self.btn_excel_import.pack(side=tk.LEFT)
        self.btn_excel_import.config(state="disabled")

        ttk.Separator(top, orient="vertical").pack(side=tk.LEFT, fill=tk.Y, padx=10)

        # Makrolar (Toblo eklentisi)
        self.btn_macro_tag = ttk.Button(top, text="🧠 Açıklamadan Etiketle", command=self.macro_auto_tag)
        self.btn_macro_tag.pack(side=tk.LEFT)
        self.btn_macro_salary = ttk.Button(top, text="💼 Maaşları Bul", command=self.macro_find_salary)
        self.btn_macro_salary.pack(side=tk.LEFT, padx=6)
        self.btn_macro_rules = ttk.Button(top, text="⚙️ Etiket Kuralları", command=self.open_tag_rules_editor)
        self.btn_macro_rules.pack(side=tk.LEFT, padx=6)
        self.btn_macro_report = ttk.Button(top, text="📊 Analiz Raporu", command=self.open_analysis_report)
        self.btn_macro_report.pack(side=tk.LEFT, padx=6)

        # (Makrolar) Excel'deki v8/v12 makrolarının uygulama içi karşılığı

        self.lbl_info = ttk.Label(top, text="")
        self.lbl_info.pack(side=tk.RIGHT)

        self.lbl_calc = ttk.Label(top, text="")
        self.lbl_calc.pack(side=tk.RIGHT, padx=(0, 12))

        # Filtre alanı (ids verilmişse kilitli gibi çalışır)
        fbox = ttk.LabelFrame(self, text="Filtre")
        fbox.pack(fill=tk.X, padx=10, pady=(0, 8))

        row = ttk.Frame(fbox)
        row.pack(fill=tk.X, pady=6)

        self.var_q = tk.StringVar(value=self.initial_filters.get("q", ""))
        self.var_tip = tk.StringVar(value=self.initial_filters.get("tip", "(Tümü)") or "(Tümü)")
        self.var_banka = tk.StringVar(value=self.initial_filters.get("banka", ""))
        self.var_hesap = tk.StringVar(value=self.initial_filters.get("hesap", ""))
        self.var_import = tk.StringVar(value=self.initial_filters.get("import_grup", "(Tümü)") or "(Tümü)")
        self.var_from = tk.StringVar(value=self.initial_filters.get("date_from", ""))
        self.var_to = tk.StringVar(value=self.initial_filters.get("date_to", ""))

        ttk.Label(row, text="Ara:").pack(side=tk.LEFT, padx=(8, 4))
        ttk.Entry(row, textvariable=self.var_q, width=26).pack(side=tk.LEFT)

        ttk.Label(row, text="Tip:").pack(side=tk.LEFT, padx=(10, 4))
        ttk.Combobox(row, textvariable=self.var_tip, values=["(Tümü)", "Giriş", "Çıkış"], state="readonly", width=10).pack(side=tk.LEFT)

        ttk.Label(row, text="Banka:").pack(side=tk.LEFT, padx=(10, 4))
        ttk.Entry(row, textvariable=self.var_banka, width=14).pack(side=tk.LEFT)

        ttk.Label(row, text="Hesap:").pack(side=tk.LEFT, padx=(10, 4))
        ttk.Entry(row, textvariable=self.var_hesap, width=18).pack(side=tk.LEFT)

        ttk.Label(row, text="Import:").pack(side=tk.LEFT, padx=(10, 4))
        try:
            groups = ["(Tümü)"] + (self.app.db.banka_import_groups(limit=80) or [])
        except Exception:
            groups = ["(Tümü)"]
        self.cmb_import = ttk.Combobox(row, textvariable=self.var_import, values=groups, state="readonly", width=28)
        self.cmb_import.pack(side=tk.LEFT)

        ttk.Label(row, text="Başlangıç:").pack(side=tk.LEFT, padx=(10, 4))
        ttk.Entry(row, textvariable=self.var_from, width=12).pack(side=tk.LEFT)

        ttk.Label(row, text="Bitiş:").pack(side=tk.LEFT, padx=(10, 4))
        ttk.Entry(row, textvariable=self.var_to, width=12).pack(side=tk.LEFT)

        ttk.Button(row, text="Uygula", command=self.refresh).pack(side=tk.LEFT, padx=10)

        # Tablo
        mid = ttk.LabelFrame(self, text="Excel Tablosu")
        mid.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        self.tree = ttk.Treeview(mid, columns=_COLS, show="headings", selectmode="extended")
        vsb = ttk.Scrollbar(mid, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(mid, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        mid.columnconfigure(0, weight=1)
        mid.rowconfigure(0, weight=1)

        # Başlıklar + kolon genişlikleri
        widths = {
            "id": 55,
            "tarih": 95,
            "banka": 140,
            "hesap": 160,
            "tip": 70,
            "tutar": 110,
            "para": 60,
            "aciklama": 320,
            "referans": 120,
            "belge": 100,
            "etiket": 110,
            "bakiye": 110,
        }
        for c in _COLS:
            self.tree.heading(c, text=c.upper(), command=lambda cc=c: self._sort_by(cc))
            self.tree.column(c, width=widths.get(c, 120), anchor=("e" if c in ("tutar", "bakiye") else "w"))

        # Renklendirme (maaş tespit vb.)
        try:
            self.tree.tag_configure("salary_hit", background="#DBEAFE")
            self.tree.tag_configure("salary_hit_weak", background="#FEF3C7")
        except Exception:
            pass

        # Hücre edit
        self.tree.bind("<Double-1>", self._on_double_click)
        self.tree.bind("<Button-1>", self._on_click)
        self.bind("<Escape>", lambda _e: self._close_editor(commit=False))

        # Kopyala / Yapıştır (çoklu hücre)
        self.bind_all("<Control-c>", self._on_copy)
        self.bind_all("<Control-C>", self._on_copy)
        self.bind_all("<Control-v>", self._on_paste)
        self.bind_all("<Control-V>", self._on_paste)
        self.tree.bind("<<TreeviewSelect>>", lambda _e: self._update_calc())

    def _apply_permissions(self):
        editable = bool(getattr(self.app, "is_admin", False))
        state = ("normal" if editable else "disabled")
        try:
            self.btn_save.config(state=state)
            self.btn_add.config(state=state)
            self.btn_del.config(state=state)
        except Exception:
            pass


    def _sync_excel_buttons(self):
        try:
            if not HAS_OPENPYXL:
                self.btn_excel_open.config(state="disabled")
                self.btn_excel_import.config(state="disabled")
                return
            self.btn_excel_open.config(state="normal")
            self.btn_excel_import.config(state=("normal" if self._last_excel_path else "disabled"))
        except Exception:
            pass

    # -----------------
    # Data load
    # -----------------
    def refresh(self):
        self._close_editor(commit=True)

        for i in self.tree.get_children():
            self.tree.delete(i)

        rows: List[Any] = []
        if self.ids:
            try:
                rows = self.db.banka_get_many(self.ids)
            except Exception:
                # geri dönüş: tek tek
                rows = []
                for hid in self.ids:
                    r = self.db.banka_get(int(hid))
                    if r:
                        rows.append(r)
        else:
            q = (self.var_q.get() or "").strip()
            tip = (self.var_tip.get() or "").strip()
            if tip == "(Tümü)":
                tip = ""
            banka = (self.var_banka.get() or "").strip()
            hesap = (self.var_hesap.get() or "").strip()
            imp = (self.var_import.get() or "").strip()
            import_grup = "" if imp in ("(Tümü)", "", None) else imp
            date_from = (self.var_from.get() or "").strip()
            date_to = (self.var_to.get() or "").strip()
            rows = self.db.banka_list(q=q, date_from=date_from, date_to=date_to, tip=tip, banka=banka, hesap=hesap, import_grup=import_grup, limit=8000)

        for r in rows:
            ig = ""
            try:
                ig = str(r["import_grup"] or "")
            except Exception:
                ig = ""
            self.tree.insert(
                "",
                tk.END,
                iid=str(r["id"]),
                tags=((f"IG:{ig}",) if ig else ()),
                values=(
                    r["id"],
                    r["tarih"],
                    r["banka"],
                    r["hesap"],
                    r["tip"],
                    fmt_amount(r["tutar"]),
                    r["para"],
                    (r["aciklama"] or ""),
                    (r["referans"] or ""),
                    (r["belge"] or ""),
                    (r["etiket"] or ""),
                    ("" if r["bakiye"] is None else fmt_amount(r["bakiye"])),
                ),
            )

        self.dirty.clear()
        self._update_info()

    def _update_info(self):
        try:
            total_rows = len(self.tree.get_children())
        except Exception:
            total_rows = 0
        self.lbl_info.config(text=f"Satır: {total_rows}  •  Değişen: {len(self.dirty)}")
        self._update_calc()

    def _update_calc(self):
        """Görünüm ve seçili satırlar için hızlı toplam hesapları."""
        def _calc(iids: List[str]) -> Tuple[float, float, float]:
            giris = 0.0
            cikis = 0.0
            for iid in iids:
                try:
                    tip = str(self.tree.set(iid, "tip") or "")
                    tutar = abs(safe_float(self.tree.set(iid, "tutar")))
                    if tip == "Giriş":
                        giris += tutar
                    else:
                        cikis += tutar
                except Exception:
                    pass
            return giris, cikis, (giris - cikis)

        try:
            all_iids = list(self.tree.get_children())
        except Exception:
            all_iids = []
        sel_iids = list(self.tree.selection() or [])

        vg, vc, vn = _calc(all_iids)
        sg, sc, sn = _calc(sel_iids)
        self.lbl_calc.config(
            text=(
                f"Görünüm: +{fmt_amount(vg)}  -{fmt_amount(vc)}  = {fmt_amount(vn)}"
                + (f"   |   Seçili: {fmt_amount(sn)}" if sel_iids else "")
            )
        )

    # -----------------
    # Sorting
    # -----------------
    def _sort_by(self, col: str):
        try:
            data = [(self.tree.set(k, col), k) for k in self.tree.get_children("")]
        except Exception:
            return

        def key_fn(v: str):
            if col in ("tutar", "bakiye"):
                return safe_float(v)
            if col == "id":
                try:
                    return int(v)
                except Exception:
                    return 0
            return (v or "").lower()

        try:
            data.sort(key=lambda t: key_fn(t[0]))
        except Exception:
            data.sort(key=lambda t: (t[0] or ""))

        # ters sırala (aynı kolona ikinci tıklama davranışı basit)
        # Treeview'de state saklamıyoruz; excel gibi hızlı olsun diye her tıklamada ters yapıyoruz.
        data.reverse()

        for index, (_val, k) in enumerate(data):
            self.tree.move(k, "", index)

    # -----------------
    # Active cell + Clipboard
    # -----------------
    def _on_click(self, event):
        """Tek tık... set active cell for Ctrl+V start."""
        # Editörü kapat (Excel gibi davran)
        self._close_editor(commit=True)

        iid = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)
        if not iid or not col_id:
            self._active_cell = None
            return
        try:
            col_index = int(col_id.replace("#", "")) - 1
        except Exception:
            return
        if 0 <= col_index < len(_COLS):
            self._active_cell = (iid, _COLS[col_index])

    def _normalize_cell(self, col: str, v: str) -> str:
        v = "" if v is None else str(v)
        s = v.strip()
        if col == "tarih":
            try:
                return parse_date_smart(s)
            except Exception:
                return s
        if col == "tip":
            n = s.lower()
            if any(k in n for k in ("cik", "çık", "debit", "borc", "borç", "gider")):
                return "Çıkış"
            if any(k in n for k in ("gir", "giriş", "credit", "alacak", "gelir")):
                return "Giriş"
            # boş ise dokunma
            return "Giriş" if not s else s
        if col == "para":
            return (s or "TL").upper()
        if col in ("tutar", "bakiye"):
            if col == "bakiye" and s == "":
                return ""
            return fmt_amount(safe_float(s))
        return s

    def _window_is_active(self) -> bool:
        try:
            w = self.focus_get()
            if w is None:
                return False
            return w.winfo_toplevel() == self
        except Exception:
            return False

    def _on_copy(self, _event=None):
        if not self._window_is_active():
            return
        cols = [c for c in _COLS if c in _EDITABLE and c != "id"]
        if not cols:
            return

        sel = list(self.tree.selection() or [])
        if not sel and self._active_cell:
            sel = [self._active_cell[0]]
        if not sel:
            return

        start_col = cols[0]
        if self._active_cell and self._active_cell[1] in cols:
            start_col = self._active_cell[1]
        try:
            start_i = cols.index(start_col)
        except Exception:
            start_i = 0
        use_cols = cols[start_i:]

        # Görünüm sırasına göre selection'ı sırala
        try:
            order = list(self.tree.get_children())
            idx = {iid: i for i, iid in enumerate(order)}
            sel.sort(key=lambda x: idx.get(x, 10**9))
        except Exception:
            pass

        lines: List[str] = []
        for iid in sel:
            row = [str(self.tree.set(iid, c) or "") for c in use_cols]
            lines.append("\t".join(row))
        text = "\n".join(lines)

        try:
            self.clipboard_clear()
            self.clipboard_append(text)
        except Exception:
            pass

    def _target_rows_for_paste(self, n_rows: int) -> List[str]:
        if n_rows <= 0:
            return []
        try:
            children = list(self.tree.get_children())
        except Exception:
            children = []
        if not children:
            return []

        active_iid = None
        if self._active_cell:
            active_iid = self._active_cell[0]
        if not active_iid:
            active_iid = self.tree.focus() or (self.tree.selection()[0] if self.tree.selection() else None)
        if not active_iid:
            return []

        # Eğer çoklu satır seçildiyse ve aktif satır seçimin içindeyse, selection sırasıyla doldur.
        sel = list(self.tree.selection() or [])
        try:
            idx = {iid: i for i, iid in enumerate(children)}
            sel_sorted = sorted(sel, key=lambda x: idx.get(x, 10**9))
        except Exception:
            sel_sorted = sel
        if sel_sorted and active_iid in sel_sorted and len(sel_sorted) >= n_rows:
            start = sel_sorted.index(active_iid)
            return sel_sorted[start : start + n_rows]

        # Aksi halde görüntü sırasına göre aşağı doğru doldur.
        try:
            start_i = children.index(active_iid)
        except Exception:
            start_i = 0
        return children[start_i : start_i + n_rows]

    def _on_paste(self, _event=None):
        if not self._window_is_active():
            return
        if not getattr(self.app, "is_admin", False):
            return

        try:
            raw = self.clipboard_get()
        except Exception:
            raw = ""
        if not raw:
            return

        text = raw.replace("\r\n", "\n").replace("\r", "\n")
        grid = [r.split("\t") for r in text.split("\n")]
        # sonda boş satırlar
        while grid and all((c.strip() == "" for c in grid[-1])):
            grid.pop()
        if not grid:
            return

        cols = [c for c in _COLS if c in _EDITABLE and c != "id"]
        if not cols:
            return

        start_col = cols[0]
        if self._active_cell and self._active_cell[1] in cols:
            start_col = self._active_cell[1]
        try:
            start_j = cols.index(start_col)
        except Exception:
            start_j = 0

        targets = self._target_rows_for_paste(len(grid))
        if not targets:
            return

        for i, row_vals in enumerate(grid):
            if i >= len(targets):
                break
            iid = targets[i]
            # satır bazında yapıştır
            for j, v in enumerate(row_vals):
                cj = start_j + j
                if cj >= len(cols):
                    break
                col = cols[cj]
                nv = self._normalize_cell(col, v)
                self.tree.set(iid, col, nv)
            try:
                self.dirty.add(int(iid))
            except Exception:
                pass

        self._update_info()

    # -----------------
    # Editing
    # -----------------
    def _on_double_click(self, event):
        if not getattr(self.app, "is_admin", False):
            return

        iid = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)  # '#1'
        if not iid or not col_id:
            return

        try:
            col_index = int(col_id.replace("#", "")) - 1
        except Exception:
            return
        if col_index < 0 or col_index >= len(_COLS):
            return

        col = _COLS[col_index]
        if col not in _EDITABLE:
            return

        bbox = self.tree.bbox(iid, col)
        if not bbox:
            return
        x, y, w, h = bbox

        # Önce var olan editoru kapat
        self._close_editor(commit=True)

        value = self.tree.set(iid, col)

        if col == "tip":
            ed = ttk.Combobox(self.tree, values=["Giriş", "Çıkış"], state="readonly")
            ed.set(value or "Giriş")
        elif col == "para":
            ed = ttk.Combobox(self.tree, values=["TL", "USD", "EUR", "GBP"], state="readonly")
            ed.set(value or "TL")
        else:
            ed = ttk.Entry(self.tree)
            ed.insert(0, value or "")

        ed.place(x=x, y=y, width=w, height=h)
        ed.focus_set()

        def commit_and_close(_e=None):
            self._commit_editor()

        ed.bind("<Return>", commit_and_close)
        ed.bind("<FocusOut>", commit_and_close)
        self._editor = ed
        try:
            row_id = int(iid)
        except Exception:
            row_id = 0
        self._editor_info = (iid, col, row_id)

    def _commit_editor(self):
        if not self._editor or not self._editor_info:
            return
        iid, col, row_id = self._editor_info
        try:
            new_val = self._editor.get()
        except Exception:
            new_val = ""

        # Normalize
        if col == "tarih":
            try:
                new_val = parse_date_smart(new_val)
            except Exception:
                pass
        if col in ("tutar", "bakiye"):
            # format numeric
            if (new_val or "").strip() == "" and col == "bakiye":
                # bakiye boş bırakılabilir
                pass
            else:
                new_val = fmt_amount(safe_float(new_val))

        self.tree.set(iid, col, new_val)

        if row_id > 0:
            self.dirty.add(row_id)
        self._close_editor(commit=False)
        self._update_info()

    def _close_editor(self, commit: bool):
        if self._editor:
            if commit:
                try:
                    self._commit_editor()
                except Exception:
                    pass
            try:
                self._editor.destroy()
            except Exception:
                pass
        self._editor = None
        self._editor_info = None

    # -----------------
    # CRUD
    # -----------------
    def _selected_ids(self) -> List[int]:
        out: List[int] = []
        for s in (self.tree.selection() or []):
            try:
                out.append(int(s))
            except Exception:
                pass
        return out

    def add_row(self):
        if not getattr(self.app, "is_admin", False):
            return

        # Basit yeni kayıt: bugün, giriş, 0 tutar ile başlat
        try:
            today = parse_date_smart("today")
        except Exception:
            today = ""

        imp = (self.var_import.get() or "").strip()
        import_grup = "" if imp in ("(Tümü)", "", None) else imp

        try:
            hid = self.db.banka_add(
                today,
                banka=self.var_banka.get().strip(),
                hesap=self.var_hesap.get().strip(),
                tip="Giriş",
                tutar=0.0,
                para="TL",
                aciklama="",
                referans="",
                belge=self.db.next_belge_no("B"),
                etiket="",
                import_grup=import_grup,
                bakiye=None,
            )
            messagebox.showinfo(APP_TITLE, "Yeni satır eklendi. (Tutarı düzenleyip Kaydet'e basabilirsin)")
        except Exception as e:
            messagebox.showerror(APP_TITLE, str(e))
            return

        # ids kilitliyse yeni id'yi de listeye ekle ki görünmeye devam etsin
        if self.ids:
            self.ids.append(int(hid))
        self.refresh()

        # yeni satırı seç
        try:
            self.tree.selection_set(str(hid))
            self.tree.see(str(hid))
        except Exception:
            pass

    def delete_selected(self):
        if not getattr(self.app, "is_admin", False):
            return
        ids = self._selected_ids()
        if not ids:
            return
        if not messagebox.askyesno(APP_TITLE, f"Seçili {len(ids)} satır silinsin mi?"):
            return

        try:
            for hid in ids:
                self.db.banka_delete(int(hid))
                if self.ids and int(hid) in self.ids:
                    self.ids.remove(int(hid))
        except Exception as e:
            messagebox.showerror(APP_TITLE, str(e))
            return

        self.refresh()
    def save_changes(self):
        if not getattr(self.app, "is_admin", False):
            messagebox.showerror(APP_TITLE, "Bu işlem için admin yetkisi gerekiyor.")
            return

        if not self.dirty:
            messagebox.showinfo(APP_TITLE, "Kaydedilecek bir değişiklik yok.")
            return

        # Değişen satırları DB'ye yaz (tek transaction)
        items: List[Dict[str, Any]] = []
        ok_ids: List[int] = []
        fail = 0

        for hid in list(self.dirty):
            try:
                vals = self.tree.item(str(hid), "values")
                if not vals:
                    continue

                tarih = vals[1]
                banka = vals[2]
                hesap = vals[3]
                tip = vals[4]
                tutar = safe_float(vals[5])
                para = vals[6]
                aciklama = vals[7]
                referans = vals[8]
                belge = vals[9]
                etiket = vals[10]
                bakiye = vals[11]

                bakiye_val: Optional[float] = None
                if str(bakiye).strip() != "":
                    bakiye_val = safe_float(bakiye)

                # import_grup: tabloda görünmez; satır tag'inden koruyoruz
                import_grup = ""
                try:
                    tags = list(self.tree.item(str(hid), "tags") or [])
                    for t in tags:
                        if isinstance(t, str) and t.startswith("IG:"):
                            import_grup = t[3:]
                            break
                except Exception:
                    import_grup = ""

                items.append(
                    {
                        "id": int(hid),
                        "tarih": tarih,
                        "banka": str(banka or ""),
                        "hesap": str(hesap or ""),
                        "tip": str(tip or "Giriş"),
                        "tutar": float(abs(tutar)),
                        "para": str(para or "TL"),
                        "aciklama": str(aciklama or ""),
                        "referans": str(referans or ""),
                        "belge": str(belge or ""),
                        "etiket": str(etiket or ""),
                        "import_grup": import_grup,
                        "bakiye": bakiye_val,
                    }
                )
                ok_ids.append(int(hid))
            except Exception:
                fail += 1

        if not items:
            messagebox.showinfo(APP_TITLE, "Kaydedilecek bir değişiklik yok.")
            return

        try:
            self.db.banka_update_many(items)
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Kaydetme başarısız: {e}")
            return

        for hid in ok_ids:
            self.dirty.discard(int(hid))

        self._update_info()
        ok = len(ok_ids)
        if fail == 0:
            messagebox.showinfo(APP_TITLE, f"Kaydedildi: {ok} satır")
        else:
            messagebox.showwarning(
                APP_TITLE,
                f"Kaydedildi: {ok} satır  •  Hata: {fail} satır\n(Bazı satırlar kaydedilemedi.)",
            )

    # -----------------
    # Toblo Makroları (Eklenti)
    # -----------------
    def _gather_rows_from_tree(self, ids: Optional[List[int]] = None) -> List[Dict[str, object]]:
        """Treeview'den satırları dict olarak toplar (analiz/makro için)."""
        out: List[Dict[str, object]] = []
        iids: List[str]
        if ids:
            iids = [str(int(x)) for x in ids]
        else:
            iids = list(self.tree.get_children())

        for iid in iids:
            vals = self.tree.item(str(iid), "values")
            if not vals:
                continue
            d: Dict[str, object] = {}
            for k, v in zip(_COLS, vals):
                d[k] = v
            # tutar/bakiye numeric olsun
            try:
                d["tutar"] = float(abs(safe_float(d.get("tutar"))))
            except Exception:
                d["tutar"] = 0.0
            out.append(d)
        return out

    def _load_tag_rules(self) -> List[Dict[str, object]]:
        """Etiket kural listesini (settings) okur. Yoksa varsayılanı döndürür."""
        try:
            raw = self.db.settings.get("banka_tag_rules")  # type: ignore[attr-defined]
            if raw:
                data = json.loads(str(raw))
                if isinstance(data, list):
                    out: List[Dict[str, object]] = []
                    for r in data:
                        if isinstance(r, dict):
                            out.append(dict(r))
                    if out:
                        return out
        except Exception:
            pass
        # kopya döndür
        return [dict(x) for x in (DEFAULT_TAG_RULES or [])]

    def _save_tag_rules(self, rules: List[Dict[str, object]]) -> None:
        try:
            self.db.settings.set("banka_tag_rules", json.dumps(rules, ensure_ascii=False))  # type: ignore[attr-defined]
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Kurallar kaydedilemedi: {e}")

    def open_tag_rules_editor(self):
        """Kural tabanlı etiketleme için basit kural editörü."""
        win = tk.Toplevel(self)
        win.title(f"{APP_TITLE} - Etiket Kuralları")
        win.geometry("860x520")
        center_window(win, self)

        rules: List[Dict[str, object]] = self._load_tag_rules()

        frm = ttk.Frame(win)
        frm.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        cols = ("priority", "tag", "pattern")
        tv = ttk.Treeview(frm, columns=cols, show="headings", height=16)
        tv.heading("priority", text="Öncelik")
        tv.heading("tag", text="Etiket")
        tv.heading("pattern", text="Pattern (Regex)")
        tv.column("priority", width=80, anchor=tk.CENTER)
        tv.column("tag", width=160)
        tv.column("pattern", width=520)

        ysb = ttk.Scrollbar(frm, orient="vertical", command=tv.yview)
        tv.configure(yscrollcommand=ysb.set)
        tv.grid(row=0, column=0, sticky="nsew")
        ysb.grid(row=0, column=1, sticky="ns")

        frm.rowconfigure(0, weight=1)
        frm.columnconfigure(0, weight=1)

        def refresh():
            tv.delete(*tv.get_children())
            for idx, r in enumerate(rules):
                pr = int(r.get("priority") or (10 * (idx + 1)))
                tag = str(r.get("tag") or "")
                pat = str(r.get("pattern") or "")
                tv.insert("", "end", iid=str(idx), values=(pr, tag, pat))

        def get_sel_index() -> Optional[int]:
            sel = tv.selection()
            if not sel:
                return None
            try:
                return int(sel[0])
            except Exception:
                return None

        def add_rule():
            tag = simpledialog.askstring(APP_TITLE, "Etiket adı:", parent=win)
            if tag is None:
                return
            tag = tag.strip()
            if not tag:
                return
            pat = simpledialog.askstring(APP_TITLE, "Regex pattern (örn: \\b(BIM|A101)\\b ):", parent=win)
            if pat is None:
                return
            pat = pat.strip()
            if not pat:
                return
            pr = simpledialog.askinteger(APP_TITLE, "Öncelik (küçük olan önce):", parent=win, initialvalue=50, minvalue=1, maxvalue=9999)
            if pr is None:
                pr = 100
            rules.append({"priority": int(pr), "tag": tag, "pattern": pat})
            refresh()

        def edit_rule():
            idx = get_sel_index()
            if idx is None or idx < 0 or idx >= len(rules):
                return
            r = rules[idx]
            tag0 = str(r.get("tag") or "")
            pat0 = str(r.get("pattern") or "")
            pr0 = int(r.get("priority") or 100)
            tag = simpledialog.askstring(APP_TITLE, "Etiket adı:", initialvalue=tag0, parent=win)
            if tag is None:
                return
            pat = simpledialog.askstring(APP_TITLE, "Regex pattern:", initialvalue=pat0, parent=win)
            if pat is None:
                return
            pr = simpledialog.askinteger(APP_TITLE, "Öncelik:", initialvalue=pr0, parent=win, minvalue=1, maxvalue=9999)
            if pr is None:
                pr = pr0
            rules[idx] = {"priority": int(pr), "tag": tag.strip(), "pattern": pat.strip()}
            refresh()

        def del_rule():
            idx = get_sel_index()
            if idx is None or idx < 0 or idx >= len(rules):
                return
            if not messagebox.askyesno(APP_TITLE, "Seçili kural silinsin mi?", parent=win):
                return
            rules.pop(idx)
            refresh()

        def move(delta: int):
            idx = get_sel_index()
            if idx is None:
                return
            j = idx + delta
            if j < 0 or j >= len(rules):
                return
            rules[idx], rules[j] = rules[j], rules[idx]
            refresh()
            tv.selection_set(str(j))

        def load_defaults():
            if not messagebox.askyesno(APP_TITLE, "Varsayılan kurallar yüklensin mi? (Mevcutlar değişir)", parent=win):
                return
            rules.clear()
            rules.extend([dict(x) for x in (DEFAULT_TAG_RULES or [])])
            refresh()

        btns = ttk.Frame(win)
        btns.pack(fill=tk.X, padx=10, pady=(0, 10))

        ttk.Button(btns, text="➕ Ekle", command=add_rule).pack(side=tk.LEFT)
        ttk.Button(btns, text="✏️ Düzenle", command=edit_rule).pack(side=tk.LEFT, padx=6)
        ttk.Button(btns, text="🗑 Sil", command=del_rule).pack(side=tk.LEFT, padx=6)

        ttk.Separator(btns, orient="vertical").pack(side=tk.LEFT, fill=tk.Y, padx=10)

        ttk.Button(btns, text="↑", width=3, command=lambda: move(-1)).pack(side=tk.LEFT)
        ttk.Button(btns, text="↓", width=3, command=lambda: move(1)).pack(side=tk.LEFT, padx=(6, 0))

        ttk.Separator(btns, orient="vertical").pack(side=tk.LEFT, fill=tk.Y, padx=10)

        ttk.Button(btns, text="↺ Varsayılanlar", command=load_defaults).pack(side=tk.LEFT)

        def save_and_close():
            # priority sırasına göre düzenleyip kaydedelim
            try:
                rules_sorted = sorted(rules, key=lambda r: int(r.get("priority") or 100))
            except Exception:
                rules_sorted = rules
            self._save_tag_rules(rules_sorted)
            win.destroy()
            messagebox.showinfo(APP_TITLE, "Kurallar kaydedildi.", parent=self)

        ttk.Button(btns, text="💾 Kaydet", command=save_and_close).pack(side=tk.RIGHT)
        ttk.Button(btns, text="Kapat", command=win.destroy).pack(side=tk.RIGHT, padx=6)

        refresh()

    def macro_find_salary(self):
        """Tablodaki satırlarda çalışan isimlerini (fuzzy) bularak renklendirir.

        - Büyük/küçük harf farkını ve ufak yazım hatalarını görmezden gelir.
        - İsteğe bağlı: Etiket kolonu "Maaş: <Ad>" olarak doldurulabilir.
        """

        try:
            emps = self.db.maas_calisan_list(only_active=True)
        except Exception:
            emps = []

        if not emps:
            messagebox.showinfo(APP_TITLE, "Aktif çalışan bulunamadı. (Maaş modülünden çalışan ekleyin)", parent=self)
            return

        # Etiket yazalım mı?
        write_tag = messagebox.askyesno(
            APP_TITLE,
            "Eşleşen satırların 'ETİKET' alanına 'Maaş: <Ad>' yazılsın mı?",
            parent=self,
        )
        overwrite = False
        if write_tag:
            overwrite = messagebox.askyesno(
                APP_TITLE,
                "Etiket doluysa üzerine yazılsın mı?",
                parent=self,
            )

        # Hangi satırlarda çalışalım?
        ids = list(self.tree.selection() or [])
        if not ids:
            ids = list(self.tree.get_children() or [])

        total = len(ids)
        if total == 0:
            messagebox.showinfo(APP_TITLE, "Tabloda satır yok.", parent=self)
            return

        # Küçük bir ilerleme penceresi
        dlg = tk.Toplevel(self)
        dlg.title("Maaşları Bul")
        dlg.transient(self)
        dlg.grab_set()
        ttk.Label(dlg, text="Maaş eşleşmeleri aranıyor...").pack(padx=12, pady=(12, 6))
        pb = ttk.Progressbar(dlg, length=420, mode="determinate", maximum=total)
        pb.pack(padx=12, pady=6)
        lbl = ttk.Label(dlg, text="0/0")
        lbl.pack(padx=12, pady=(0, 12))

        updated = 0
        colored = 0

        # Çalışanları hazırlayalım
        emp_list: list[tuple[str, float]] = []
        for e in emps:
            try:
                emp_list.append((str(e["ad"]), float(e["aylik_tutar"] or 0.0)))
            except Exception:
                continue

        def step(i: int = 0):
            nonlocal updated, colored
            end = min(i + 250, total)
            for k in range(i, end):
                iid = str(ids[k])
                try:
                    if not self.tree.exists(iid):
                        continue
                    vals = list(self.tree.item(iid, "values") or [])
                    if not vals or len(vals) < len(_COLS):
                        continue
                    row = dict(zip(_COLS, vals))

                    tip = str(row.get("tip") or "")
                    if "CIK" not in tip.upper() and "ÇIK" not in tip.upper():
                        continue
                    desc = str(row.get("aciklama") or "")
                    if not desc.strip():
                        continue

                    tutar = abs(float(safe_float(row.get("tutar") or 0)))

                    best_score = 0.0
                    best_name = ""
                    for ename, expected in emp_list:
                        name_sc = best_substring_similarity(ename, desc)
                        if name_sc < 0.55:
                            continue
                        if expected > 0:
                            a_sc = amount_score(tutar, expected, abs_tol=5.0, pct_tol=0.15)
                            if a_sc > 0:
                                score = combine_scores(name_sc, a_sc, w_name=0.80, w_amt=0.20)
                            else:
                                score = name_sc * 0.90
                        else:
                            score = name_sc
                        if score > best_score:
                            best_score = score
                            best_name = ename

                    if best_score >= 0.78:
                        # tag ekle
                        old_tags = set(self.tree.item(iid, "tags") or ())
                        old_tags.discard("salary_hit")
                        old_tags.discard("salary_hit_weak")
                        if best_score >= 0.87:
                            old_tags.add("salary_hit")
                        else:
                            old_tags.add("salary_hit_weak")
                        self.tree.item(iid, tags=tuple(old_tags))
                        colored += 1

                        if write_tag:
                            cur = str(row.get("etiket") or "").strip()
                            if (not overwrite) and cur:
                                pass
                            else:
                                new_tag = f"Maaş: {best_name}"[:60]
                                # etiket kolon indexi 10
                                vals[10] = new_tag
                                self.tree.item(iid, values=tuple(vals))
                                try:
                                    rid = int(row.get("id") or int(iid))
                                    self.dirty.add(int(rid))
                                except Exception:
                                    pass
                                updated += 1
                    else:
                        # eşleşme yoksa maaş taglerini kaldır
                        old_tags = set(self.tree.item(iid, "tags") or ())
                        if "salary_hit" in old_tags or "salary_hit_weak" in old_tags:
                            old_tags.discard("salary_hit")
                            old_tags.discard("salary_hit_weak")
                            self.tree.item(iid, tags=tuple(old_tags))
                except Exception:
                    continue

            pb["value"] = end
            lbl.config(text=f"{end}/{total}")
            if end >= total:
                try:
                    dlg.destroy()
                except Exception:
                    pass
                self._update_info()
                messagebox.showinfo(
                    APP_TITLE,
                    f"Tamamlandı.\n\nRenklendirilen satır: {colored}\nEtiketi güncellenen satır: {updated}\n\nNot: Etiket değişikliklerini kalıcı yapmak için 'Değişiklikleri Kaydet' kullanın.",
                    parent=self,
                )
                return
            self.after(1, lambda: step(end))

        self.after(10, step)

    def macro_auto_tag(self):
        """Açıklamaya göre otomatik etiketleme (kural + öğrenen + fuzzy)."""
        if not getattr(self.app, "is_admin", False):
            messagebox.showerror(APP_TITLE, "Bu işlem için admin yetkisi gerekiyor.")
            return

        ids = self._selected_ids()
        if not ids:
            ids = [int(x) for x in self.tree.get_children()]
        if not ids:
            return

        if not messagebox.askyesno(
            APP_TITLE,
            "Açıklamaya göre otomatik etiketleme çalışsın mı?\\n\\n"
            "Not: Bu işlem tabloyu günceller; kaydetmek için ayrıca 'Değişiklikleri Kaydet' demen gerekir.",
            parent=self,
        ):
            return

        overwrite = messagebox.askyesno(
            APP_TITLE,
            "Mevcut ETİKET dolu olan satırların üstüne yazılsın mı?\\n\\n"
            "Evet: hepsini yeniden etiketler\\nHayır: sadece boş etiketleri doldurur",
            parent=self,
        )

        split_pm = messagebox.askyesno(
            APP_TITLE,
            "(+)/(−) ayrımı yapılsın mı?\\n\\n"
            "Evet: aynı açıklama için Giriş/Çıkış ayrı etiket alır (A (+), A (-)).",
            parent=self,
        )

        rows = self._gather_rows_from_tree(ids)
        if not rows:
            return

        rules = self._load_tag_rules()
        target_only_empty = (not overwrite)

        # Progress / İptal penceresi
        dlg = tk.Toplevel(self)
        dlg.title(f"{APP_TITLE} - Etiketleme")
        dlg.geometry("520x160")
        dlg.resizable(False, False)
        center_window(dlg, self)
        dlg.transient(self)
        dlg.grab_set()

        lbl = ttk.Label(dlg, text="Hazırlanıyor...", anchor="w", justify="left")
        lbl.pack(fill=tk.X, padx=12, pady=(12, 6))

        pb = ttk.Progressbar(dlg, mode="determinate", length=480)
        pb.pack(padx=12, pady=(0, 10))

        btns = ttk.Frame(dlg)
        btns.pack(fill=tk.X, padx=12)

        cancel_event = threading.Event()

        def on_cancel():
            cancel_event.set()
            lbl.config(text="İptal ediliyor...")

        ttk.Button(btns, text="İptal", command=on_cancel).pack(side=tk.RIGHT)

        q: "queue.Queue[tuple]" = queue.Queue()

        def worker():
            try:
                tag_map, group_count = build_tag_suggestions(
                    rows,
                    rules=rules,
                    target_only_empty=target_only_empty,
                    split_plus_minus=split_pm,
                    progress_cb=lambda phase, cur, tot: q.put(("progress", phase, cur, tot)),
                    should_cancel=cancel_event.is_set,
                )
                q.put(("done", tag_map, group_count))
            except Exception as e:
                q.put(("error", str(e)))

        threading.Thread(target=worker, daemon=True).start()

        updated = 0
        def apply_tags(tag_map: Dict[int, str], group_count: int):
            nonlocal updated
            items = list(tag_map.items())
            total = len(items)
            if total == 0:
                try:
                    dlg.destroy()
                except Exception:
                    pass
                messagebox.showinfo(APP_TITLE, "Uygulanacak etiket bulunamadı.", parent=self)
                return

            pb["maximum"] = total
            pb["value"] = 0

            def step(i: int = 0):
                nonlocal updated
                if cancel_event.is_set():
                    try:
                        dlg.destroy()
                    except Exception:
                        pass
                    messagebox.showinfo(APP_TITLE, "Etiketleme iptal edildi.", parent=self)
                    return

                end = min(i + 300, total)
                for k in range(i, end):
                    rid, tag = items[k]
                    iid = str(int(rid))
                    try:
                        if not self.tree.exists(iid):
                            continue
                        vals = list(self.tree.item(iid, "values") or [])
                        if not vals or len(vals) < 11:
                            continue
                        cur = str(vals[10] or "").strip()
                        if (not overwrite) and cur:
                            continue
                        vals[10] = str(tag)
                        self.tree.item(iid, values=tuple(vals))
                        self.dirty.add(int(rid))
                        updated += 1
                    except Exception:
                        continue

                pb["value"] = end
                lbl.config(text=f"Uygulanıyor... {end}/{total}")

                if end >= total:
                    try:
                        dlg.destroy()
                    except Exception:
                        pass
                    self._update_info()
                    messagebox.showinfo(
                        APP_TITLE,
                        f"Etiketleme tamamlandı.\n\nGüncellenen satır: {updated}\nFuzzy grup sayısı: {group_count}\nKural sayısı: {len(rules)}\n\nNot: Değişiklikleri kalıcı yapmak için 'Değişiklikleri Kaydet' butonunu kullan.",
                        parent=self,
                    )
                    return

                self.after(1, lambda: step(end))

            step(0)

        def poll():
            if cancel_event.is_set():
                # worker bitince zaten error/done döner; şimdilik bekle
                pass
            try:
                while True:
                    msg = q.get_nowait()
                    if not msg:
                        continue
                    kind = msg[0]
                    if kind == "progress":
                        _, phase, cur, tot = msg
                        if cancel_event.is_set():
                            continue
                        pb["maximum"] = max(int(tot), 1)
                        pb["value"] = int(cur)
                        lbl.config(text=f"{phase}: {cur}/{tot}")
                    elif kind == "error":
                        _, err = msg
                        try:
                            dlg.destroy()
                        except Exception:
                            pass
                        messagebox.showerror(APP_TITLE, f"Etiketleme başarısız: {err}", parent=self)
                        return
                    elif kind == "done":
                        _, tag_map, group_count = msg
                        apply_tags(tag_map, int(group_count))
                        return
            except queue.Empty:
                pass

            if dlg.winfo_exists():
                dlg.after(60, poll)

        poll()

    def open_analysis_report(self):
        """Excel'deki v12 benzeri: tablo verisiyle analiz penceresi açar."""
        ids = self._selected_ids()
        title_suffix = "Seçili" if ids else "Görüntülenen"
        rows = self._gather_rows_from_tree(ids if ids else None)
        if not rows:
            return
        try:
            BankaAnalizWindow(self.app, rows, title_suffix=title_suffix)
        except Exception as e:
            messagebox.showerror(APP_TITLE, str(e))

    # -----------------
    # Export
    # -----------------
    # -----------------
    # Excel Export / Import (Toblo <-> Excel)
    # -----------------
    def _export_excel_to_path(self, path: str) -> None:
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl kurulu değil. Kur: pip install openpyxl")
        try:
            import openpyxl  # type: ignore
        except Exception as e:
            raise RuntimeError("openpyxl import edilemedi. Kur: pip install openpyxl") from e

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BankaHareket"

        ws.append(list(_COLS))
        for iid in self.tree.get_children():
            vals = list(self.tree.item(iid, "values"))
            ws.append(vals)

        wb.save(path)

    def export_excel(self):
        if not HAS_OPENPYXL:
            messagebox.showerror(APP_TITLE, "openpyxl kurulu değil. Kur: pip install openpyxl")
            return

        p = filedialog.asksaveasfilename(
            title="Excel Kaydet",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("All", "*.*")],
        )
        if not p:
            return

        try:
            self._export_excel_to_path(p)
        except Exception as e:
            messagebox.showerror(APP_TITLE, str(e))
            return

        # Son excel yolu
        try:
            self._last_excel_path = str(p)
            self._last_excel_mtime = os.path.getmtime(self._last_excel_path)
        except Exception:
            pass
        self._sync_excel_buttons()
        messagebox.showinfo(APP_TITLE, "Excel dosyası oluşturuldu.")

    def _open_file(self, path: str) -> None:
        """Sistemde dosyayı varsayılan uygulama ile aç."""
        if not path:
            return
        try:
            if sys.platform.startswith("win"):
                os.startfile(path)  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                subprocess.Popen(["open", path])
            else:
                subprocess.Popen(["xdg-open", path])
        except Exception as e:
            raise RuntimeError(f"Dosya açılamadı: {e}")

    def open_in_excel(self):
        """Görüntülenen tabloyu Excel'e çıkar ve Excel'de aç.

        Not: Excel'de yapılan değişiklikleri tekrar uygulamaya almak için
        '🔄 Excel'den Güncelle' butonunu kullan.
        """
        if not HAS_OPENPYXL:
            messagebox.showerror(APP_TITLE, "openpyxl kurulu değil. Kur: pip install openpyxl")
            return

        # Kullanıcı dosya yeri seçsin (Excel açıkken kaydetme/izleme daha sağlıklı)
        default_name = f"BankaHareket_{int(time.time())}.xlsx"
        p = filedialog.asksaveasfilename(
            title="Excel'de Aç (Kaydet)",
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("All", "*.*")],
        )
        if not p:
            return

        try:
            self._export_excel_to_path(p)
            self._open_file(p)
            self._last_excel_path = str(p)
            self._last_excel_mtime = os.path.getmtime(self._last_excel_path)
        except Exception as e:
            messagebox.showerror(APP_TITLE, str(e))
            return

        self._sync_excel_buttons()

    def import_from_excel(self):
        """Excel dosyasındaki (ID bazlı) değişiklikleri Toblo tablosuna uygular.

        - Excel'de ilk satır header olmalı
        - 'id' kolonu zorunlu
        - Eşleşen id'lerde, bulunan kolonlar güncellenir
        - Güncellenen satırlar 'dirty' işaretlenir (Kaydet ile DB'ye yazılır)
        """
        if not getattr(self.app, "is_admin", False):
            messagebox.showerror(APP_TITLE, "Bu işlem için admin yetkisi gerekiyor.")
            return
        if not HAS_OPENPYXL:
            messagebox.showerror(APP_TITLE, "openpyxl kurulu değil. Kur: pip install openpyxl")
            return

        use_last = False
        if self._last_excel_path and os.path.exists(self._last_excel_path):
            use_last = messagebox.askyesno(
                APP_TITLE,
                f"Son Excel dosyası kullanılsın mı?\n\n{self._last_excel_path}",
                parent=self,
            )

        if use_last:
            p = self._last_excel_path
        else:
            p = filedialog.askopenfilename(
                title="Excel'den Güncelle",
                filetypes=[("Excel", "*.xlsx"), ("All", "*.*")],
            )
        if not p:
            return

        try:
            import openpyxl  # type: ignore
            wb = openpyxl.load_workbook(p)
            ws = wb.active

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                raise RuntimeError("Excel boş.")

            header = [str(x).strip() if x is not None else "" for x in rows[0]]
            idx = {h.lower(): i for i, h in enumerate(header) if h}
            if 'id' not in idx:
                raise RuntimeError("Excel'de 'id' kolonu bulunamadı.")

            # Güncellenebilir kolonlar
            editable_cols = [c for c in _COLS if c in _EDITABLE]

            updated = 0
            for r in rows[1:]:
                if r is None:
                    continue
                rid = r[idx['id']]
                if rid is None or str(rid).strip() == '':
                    continue
                try:
                    iid = str(int(rid))
                except Exception:
                    continue
                if not self.tree.exists(iid):
                    continue

                cur_vals = list(self.tree.item(iid, 'values') or [])
                if not cur_vals or len(cur_vals) < len(_COLS):
                    continue

                changed = False
                for col in editable_cols:
                    if col.lower() not in idx:
                        continue
                    j = idx[col.lower()]
                    if j >= len(r):
                        continue
                    v = r[j]
                    nv = self._normalize_cell(col, '' if v is None else str(v))
                    # tree values index
                    k = list(_COLS).index(col)
                    if str(cur_vals[k]) != str(nv):
                        cur_vals[k] = nv
                        changed = True

                if changed:
                    self.tree.item(iid, values=tuple(cur_vals))
                    try:
                        self.dirty.add(int(iid))
                    except Exception:
                        pass
                    updated += 1

            self._update_info()
            self._last_excel_path = str(p)
            try:
                self._last_excel_mtime = os.path.getmtime(self._last_excel_path)
            except Exception:
                pass
            self._sync_excel_buttons()

            messagebox.showinfo(APP_TITLE, f"Excel'den güncellendi.\n\nGüncellenen satır: {updated}\n\nNot: Kalıcı olması için 'Değişiklikleri Kaydet'")

        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Excel içe aktarma başarısız: {e}")
