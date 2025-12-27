# -*- coding: utf-8 -*-
"""KasaPro v3 - Excel içe aktarma eşleştirme sihirbazı."""

from __future__ import annotations

import logging
import os
import re
import threading
from datetime import datetime, date
from typing import Any, Optional, List, Dict, Tuple, TYPE_CHECKING, Callable

import tkinter as tk
from tkinter import ttk, messagebox

from ...config import APP_TITLE
from ...utils import (
    center_window,
    now_iso,
    parse_date_smart,
    safe_float,
    norm_header,
)
from ...db.main_db import DB

# Maaş/isim eşleştirme için fuzzy yardımcılar
from ...core.fuzzy import best_substring_similarity, normalize_text, similarity

if TYPE_CHECKING:
    from ...app import App

# IMPORT WIZARD (Excel mapping)
# =========================

def _auto_header_row(ws, search_rows: int = 30) -> int:
    best_row = 1
    best_score = -1
    max_col = min(ws.max_column or 1, 60)
    for r in range(1, min(ws.max_row or 1, search_rows) + 1):
        texts = 0
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip():
                texts += 1
        if texts > best_score:
            best_score = texts
            best_row = r
    return best_row

def _headers(ws, header_row: int) -> List[str]:
    max_col = min(ws.max_column or 1, 60)
    out = []
    for c in range(1, max_col + 1):
        v = ws.cell(header_row, c).value
        out.append("" if v is None else str(v).strip())
    return out

def _suggest_col(headers: List[str], keys: List[str]) -> int:
    norm = [norm_header(h) for h in headers]
    for k in keys:
        kk = norm_header(k)
        for i, h in enumerate(norm):
            if h == kk:
                return i
    for k in keys:
        kk = norm_header(k)
        for i, h in enumerate(norm):
            if kk and kk in h:
                return i
    return -1

class ImportWizard(tk.Toplevel):
    def __init__(self, app: "App", xlsx_path: str, mode: str = "full", context: Optional[Dict[str, Any]] = None):
        super().__init__(app.root)
        self.app = app
        self.db = app.db
        self.xlsx_path = xlsx_path
        self.mode = (mode or "full").strip().lower()
        self.context: Dict[str, Any] = dict(context or {})
        self.title("Excel İçe Aktar - Eşleştirme Sihirbazı")
        self.geometry("1080x720")
        self.resizable(True, True)
        self.grab_set()

        try:
            import openpyxl  # type: ignore
        except Exception:
            messagebox.showerror(APP_TITLE, "openpyxl kurulu değil. Kur: pip install openpyxl")
            self.destroy()
            return

        self.openpyxl = openpyxl
        try:
            self.wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        except Exception as exc:
            logging.getLogger(__name__).exception("Excel yükleme başarısız: %s", xlsx_path)
            messagebox.showerror(APP_TITLE, f"Excel dosyası açılamadı:\n{exc}")
            self.destroy()
            return
        self.sheetnames = self.wb.sheetnames[:]
        self.result_counts: Optional[Dict[str, int]] = None
        self.mappings: Dict[str, Dict[str, Any]] = {}
        self._import_in_progress = False

        self._build()
        center_window(self, app.root)
        try:
            self.protocol("WM_DELETE_WINDOW", self._on_close)
        except Exception:
            pass

    def _on_close(self) -> None:
        self._cleanup()
        self.destroy()

    def _cleanup(self) -> None:
        if self._cleanup_done:
            return
        self._cleanup_done = True
        try:
            wb = getattr(self, "wb", None)
            if wb is not None and hasattr(wb, "close"):
                wb.close()
        except Exception:
            pass
        try:
            self.wb = None  # type: ignore[assignment]
        except Exception:
            pass

    def _build(self):
        ttk.Label(self, text=f"Dosya: {os.path.basename(self.xlsx_path)}", font=("Calibri", 12, "bold")).pack(anchor="w", padx=12, pady=(10, 4))
        ttk.Label(self, text="Sayfa seç → Kolon eşleştir → İçe Aktar. (Başlıklar farklı olabilir, sorun değil.)", foreground="#666").pack(anchor="w", padx=12, pady=(0, 10))

        self.nb = ttk.Notebook(self)
        self.nb.pack(fill=tk.BOTH, expand=True, padx=12, pady=8)

        # Modlar:
        # - full: Cariler + CariHareket + KasaHareket + BankaHareket
        # - bank: sadece BankaHareket
        # - maas: Maaş Ödeme (MaasOdeme)
        if self.mode in ("full", "cariler"):
            self._add_tab("Cariler", fields=[
            ("ad", "Cari Adı (Zorunlu)"),
            ("tur", "Tür"),
            ("telefon", "Telefon"),
            ("notlar", "Notlar"),
            ("acilis", "Açılış Bakiyesi"),
        ], suggest={
            "ad": ["cari", "cari adı", "ad", "unvan", "isim"],
            "tur": ["tur", "tip"],
            "telefon": ["telefon", "tel", "gsm"],
            "notlar": ["not", "aciklama", "açıklama"],
            "acilis": ["acilis", "açılış", "acilis bakiyesi", "bakiye"],
        })

        if self.mode in ("full", "cari", "carihareket"):
            self._add_tab("CariHareket", fields=[
            ("tarih", "Tarih"),
            ("cari", "Cari (Zorunlu)"),
            ("tip", "Tip (Borç/Alacak)"),
            ("tutar", "Tutar (Zorunlu)"),
            ("para", "Para"),
            ("odeme", "Ödeme"),
            ("belge", "Belge"),
            ("etiket", "Etiket"),
            ("aciklama", "Açıklama"),
        ], suggest={
            "tarih": ["tarih", "date"],
            "cari": ["cari", "cari adı", "unvan", "isim"],
            "tip": ["tip", "borç", "borc", "alacak"],
            "tutar": ["tutar", "miktar", "amount", "tahsilat", "odeme", "ödeme", "gelir", "gider", "gelen"],
            "para": ["para", "pb", "doviz", "döviz"],
            "odeme": ["odeme", "ödeme", "odeme tipi", "ödeme tipi"],
            "belge": ["belge", "fis", "fiş", "dekont"],
            "etiket": ["etiket", "tag"],
            "aciklama": ["aciklama", "açıklama", "not"],
        })

        if self.mode in ("full", "kasa", "kasahareket"):
            self._add_tab("KasaHareket", fields=[
            ("tarih", "Tarih"),
            ("tip", "Tip (Gelir/Gider)"),
            ("tutar", "Tutar (Zorunlu)"),
            ("para", "Para"),
            ("odeme", "Ödeme"),
            ("kategori", "Kategori/Grup"),
            ("cari", "Cari (opsiyonel)"),
            ("belge", "Belge"),
            ("etiket", "Etiket"),
            ("aciklama", "Açıklama"),
        ], suggest={
            "tarih": ["tarih", "date"],
            "tip": ["tip", "gelir", "gider"],
            "tutar": ["tutar", "miktar", "amount", "tahsilat", "odeme", "ödeme", "gelir", "gider", "gelen"],
            "para": ["para", "pb", "doviz", "döviz"],
            "odeme": ["odeme", "ödeme", "odeme tipi", "ödeme tipi"],
            "kategori": ["kategori", "grup", "group"],
            "cari": ["cari", "unvan", "isim"],
            "belge": ["belge", "fis", "fiş", "dekont"],
            "etiket": ["etiket", "tag"],
            "aciklama": ["aciklama", "açıklama", "not"],
        })

        if self.mode in ("full", "bank", "banka"):
            self._add_tab("BankaHareket", fields=[
                ("tarih", "Tarih"),
                ("banka", "Banka"),
                ("hesap", "Hesap"),
                ("aciklama", "Açıklama"),
                ("tutar", "Tutar (+/-)"),
                ("borc", "Borç/Çıkış"),
                ("alacak", "Alacak/Giriş"),
                ("para", "Para"),
                ("bakiye", "Bakiye"),
                ("referans", "Referans"),
                ("belge", "Belge"),
                ("etiket", "Etiket"),
            ], suggest={
                "tarih": ["tarih", "date", "işlem tarihi", "islem tarihi"],
                "banka": ["banka", "bank"],
                "hesap": ["hesap", "iban", "account", "şube", "sube"],
                "aciklama": ["açıklama", "aciklama", "işlem açıklaması", "islem aciklamasi", "detay", "description"],
                "tutar": ["tutar", "miktar", "amount", "tahsilat", "odeme", "ödeme", "gelir", "gider", "gelen", "giden"],
                "borc": ["borç", "borc", "debit", "çıkış", "cikis"],
                "alacak": ["alacak", "credit", "giriş", "giris"],
                "para": ["para", "pb", "döviz", "doviz", "currency"],
                "bakiye": ["bakiye", "balance"],
                "referans": ["referans", "ref", "referans no", "işlem no", "islem no"],
                "belge": ["belge", "dekont", "fiş", "fis"],
                "etiket": ["etiket", "tag"],
            })

        if self.mode in ("maas", "salary"):
            self._add_tab(
                "MaasOdeme",
                fields=[
                    ("calisan", "Çalışan Adı (Zorunlu)"),
                    ("tutar", "Tutar (Zorunlu)"),
                    ("para", "Para"),
                    ("odendi", "Ödendi (0/1 / Evet/Hayır)"),
                    ("odeme_tarihi", "Ödeme Tarihi"),
                    ("aciklama", "Açıklama"),
                ],
                suggest={
                    "calisan": ["calisan", "çalışan", "personel", "ad", "ad soyad", "isim", "unvan", "alici", "alıcı"],
                    "tutar": ["tutar", "miktar", "net", "maas", "maaş", "ucret", "ücret", "amount", "odeme"],
                    "para": ["para", "pb", "doviz", "döviz", "currency"],
                    "odendi": ["odendi", "ödendi", "odeme durumu", "durum", "paid"],
                    "odeme_tarihi": ["odeme tarihi", "ödeme tarihi", "tarih", "date"],
                    "aciklama": ["aciklama", "açıklama", "not"],
                },
            )

        bottom = ttk.Frame(self)
        bottom.pack(fill=tk.X, padx=12, pady=10)

        self.var_create_missing_cari = tk.BooleanVar(value=True)
        # Cari oluşturma seçeneği sadece Cari/Kasa importlarında anlamlı
        if self.mode in ("full", "cariler", "cari", "carihareket", "kasa", "kasahareket"):
            ttk.Checkbutton(bottom, text="İçe aktarma sırasında cari yoksa otomatik oluştur", variable=self.var_create_missing_cari).pack(side=tk.LEFT)

        self.btn_cancel = ttk.Button(bottom, text="İptal", command=self._cancel)
        self.btn_cancel.pack(side=tk.RIGHT, padx=6)
        self.btn_import = ttk.Button(bottom, text="İçe Aktar", command=self._do_import)
        self.btn_import.pack(side=tk.RIGHT, padx=6)

    def _guess_sheet(self, tab_name: str) -> str:
        t = tab_name.lower()
        for s in self.sheetnames:
            sl = s.lower()
            if t == "cariler" and ("cari" in sl and "hareket" not in sl):
                return s
            if t == "carihareket" and ("cari" in sl and "hareket" in sl):
                return s
            if t == "kasahareket" and ("kasa" in sl):
                return s
            if t == "bankahareket" and ("banka" in sl or "ekstre" in sl or "hesap" in sl or "hesap hareket" in sl):
                return s
            if t == "maasodeme" and ("maas" in sl or "maaş" in sl or "ucret" in sl or "ücret" in sl or "personel" in sl):
                return s
        return self.sheetnames[0] if self.sheetnames else ""

    def _add_tab(self, name: str, fields: List[Tuple[str, str]], suggest: Dict[str, List[str]]):
        tab = ttk.Frame(self.nb)
        self.nb.add(tab, text=name)

        top = ttk.Frame(tab)
        top.pack(fill=tk.X, padx=10, pady=10)
        ttk.Label(top, text="Sayfa:", width=8).pack(side=tk.LEFT)

        guess = self._guess_sheet(name)
        sheet_var = tk.StringVar(value=(guess if guess else "(Atla)"))
        sheet_cmb = ttk.Combobox(top, textvariable=sheet_var, values=["(Atla)"] + self.sheetnames, state="readonly", width=28)
        sheet_cmb.pack(side=tk.LEFT, padx=(0,10))
        ttk.Button(top, text="Önizleme/Yenile", command=lambda: self._refresh_preview(name)).pack(side=tk.LEFT)

        mapbox = ttk.LabelFrame(tab, text="Kolon Eşleştirme")
        mapbox.pack(fill=tk.X, padx=10, pady=(0,10))

        # Maaş import için çalışan listesi (sistemden)
        emp_names: List[str] = []
        if name == "MaasOdeme":
            try:
                emp_rows = self.db.maas_calisan_list(q="", only_active=False)  # type: ignore
                # sqlite3.Row nesnesi dict gibi .get() desteklemez; bu yüzden güvenli şekilde çekiyoruz.
                emp_names = []
                for rr in (emp_rows or []):
                    ad_val = None
                    try:
                        ad_val = rr["ad"]
                    except Exception:
                        try:
                            # bazı eski sorgularda tek kolon dönebiliyor
                            ad_val = rr[0]
                        except Exception:
                            ad_val = None
                    if ad_val is None:
                        continue
                    s = str(ad_val).strip()
                    if s:
                        emp_names.append(s)
            except Exception:
                emp_names = []

        # Maaş import: Kullanıcı "Çalışan" alanında sistemdeki çalışandan seçsin; Excel'de ad/soyad ile satırları bulsun.
        maas_emp_var: Optional[tk.StringVar] = None
        maas_fixed_emp_var: Optional[tk.StringVar] = None
        maas_fixed_mode_var: Optional[tk.BooleanVar] = None
        maas_only_selected_rows_var: Optional[tk.BooleanVar] = None
        maas_rowmatch_rows: List[int] = []
        maas_calisan_emp_cmb: Optional[ttk.Combobox] = None
        maas_calisan_find_btn: Optional[ttk.Button] = None

        mapping_vars: Dict[str, tk.StringVar] = {}
        mapping_widgets: Dict[str, ttk.Combobox] = {}

        for row, (key, label) in enumerate(fields):
            ttk.Label(mapbox, text=label, width=22).grid(row=row, column=0, sticky="w", padx=10, pady=4)
            # Maaş tabında 'Çalışan' alanı: Excel kolonundan seçilmez.
            # Kullanıcı sistemde kayıtlı çalışanı seçer ve Excel içinde adı/soyadı aranır.
            if name == "MaasOdeme" and key == "calisan":
                # mapping_vars yine olsun (plan oluştururken alan eksik kalmasın) ama kolon seçimi değil.
                mapping_vars[key] = tk.StringVar(value="(Sistem)")

                # Sistem çalışanı ile arama modu daima açık
                maas_fixed_mode_var = tk.BooleanVar(value=True)
                maas_fixed_emp_var = tk.StringVar(value=(emp_names[0] if emp_names else ""))

                cont = ttk.Frame(mapbox)
                cont.grid(row=row, column=1, sticky="w", padx=6, pady=4)

                ttk.Label(cont, text="Sistem Çalışanı:").grid(row=0, column=0, sticky="w")
                maas_calisan_emp_cmb = ttk.Combobox(
                    cont,
                    textvariable=maas_fixed_emp_var,
                    values=(emp_names or ["(Önce Çalışan ekleyin)"]),
                    state=("readonly" if emp_names else "disabled"),
                    width=30,
                )
                maas_calisan_emp_cmb.grid(row=0, column=1, sticky="w", padx=(8, 0))

                maas_calisan_find_btn = ttk.Button(
                    cont,
                    text="Excel'de Ara (Satırı/Satırları Getir)",
                    command=self._make_find_employee_rows_command(name),
                )
                maas_calisan_find_btn.grid(row=0, column=2, sticky="w", padx=8)

                ttk.Label(
                    cont,
                    text="Not: Burada Excel kolon seçimi yok. Seçtiğin çalışanın adı/soyadı Excel içinde aranır; bulunan satırlar önizlemeye gelir.",
                    foreground="#666",
                ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(4, 0))

                # Çalışan için Excel kolon combobox'ı yok.
                continue

            # Normal alanlar (Excel kolon eşleştirme)
            var = tk.StringVar(value="(Yok)")
            mapping_vars[key] = var
            cmb = ttk.Combobox(mapbox, textvariable=var, values=["(Yok)"], state="readonly", width=40)
            cmb.grid(row=row, column=1, sticky="w", padx=6, pady=4)
            mapping_widgets[key] = cmb

        # Maaş tabı: sadece seçili satırları içe aktar (bulunan satırları seçip getirme mantığı)
        if name == "MaasOdeme" and emp_names:
            base = len(fields)
            ttk.Separator(mapbox, orient="horizontal").grid(row=base, column=0, columnspan=2, sticky="ew", padx=10, pady=(8, 6))
            maas_only_selected_rows_var = tk.BooleanVar(value=True)
            ttk.Checkbutton(
                mapbox,
                text="Sadece tabloda seçili satırları içe aktar (önerilir)",
                variable=maas_only_selected_rows_var,
            ).grid(row=base + 1, column=0, columnspan=2, sticky="w", padx=10, pady=(0, 6))

        prev = ttk.LabelFrame(tab, text="Önizleme")
        prev.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0,10))

        tree = ttk.Treeview(prev, show="headings", height=14)
        try:
            tree.configure(selectmode="extended")
        except Exception:
            pass
        tree.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)
        scr = ttk.Scrollbar(prev, orient="vertical", command=tree.yview)
        scr.pack(fill=tk.Y, side=tk.RIGHT)
        tree.configure(yscrollcommand=scr.set)

        self.mappings[name] = {
            "tab": tab,
            "sheet_var": sheet_var,
            "sheet_cmb": sheet_cmb,
            "mapping_vars": mapping_vars,
            "mapping_widgets": mapping_widgets,
            "tree": tree,
            "fields": fields,
            "suggest": suggest,
            "header_row": 1,
            "headers": [],
            "maas_emp_var": maas_emp_var,
            "maas_fixed_mode_var": maas_fixed_mode_var,
            "maas_fixed_emp_var": maas_fixed_emp_var,
            "maas_only_selected_rows_var": maas_only_selected_rows_var,
            "maas_rowmatch_rows": maas_rowmatch_rows,
            # Maaş tabı: 'Çalışan' alanındaki özel widget'lar
            "maas_calisan_emp_cmb": maas_calisan_emp_cmb,
            "maas_calisan_excel_cmb": None,
            "maas_calisan_find_btn": maas_calisan_find_btn,
            "preview_row_map": [],
        }

        sheet_cmb.bind("<<ComboboxSelected>>", lambda _e: self._refresh_preview(name))
        self._refresh_preview(name)

    def _refresh_preview(self, tab_name: str):
        info = self.mappings[tab_name]
        sheet_name = info["sheet_var"].get()
        tree: ttk.Treeview = info["tree"]

        if not sheet_name or sheet_name == "(Atla)":
            for i in tree.get_children():
                tree.delete(i)
            tree["columns"] = ()
            return

        ws = self.wb[sheet_name]
        header_row = _auto_header_row(ws, search_rows=30)
        headers = _headers(ws, header_row)
        info["header_row"] = header_row
        info["headers"] = headers

        col_options = ["(Yok)"]
        for idx, h in enumerate(headers, start=1):
            col_letter = self.openpyxl.utils.get_column_letter(idx)
            col_options.append(f"{col_letter}: {h}" if h else f"{col_letter}: (boş)")

        # Sadece kolon eşleştirme combobox'larını güncelle (maaş çalışan bulucu combobox'ını bozma)
        for cmb in (info.get("mapping_widgets") or {}).values():
            try:
                cmb["values"] = col_options
            except Exception:
                pass

        suggest: Dict[str, List[str]] = info["suggest"]
        for key, var in info["mapping_vars"].items():
            # Maaş "sabit çalışan" modunda çalışan kolonunu otomatik önermeye çalışma
            if tab_name == "MaasOdeme" and key == "calisan":
                try:
                    if self._get_info_bool(info, "maas_fixed_mode_var"):
                        continue
                except Exception:
                    pass
            if var.get() not in ("", "(Yok)"):
                continue
            idx0 = _suggest_col(headers, suggest.get(key, []))
            if idx0 >= 0:
                col_letter = self.openpyxl.utils.get_column_letter(idx0 + 1)
                h = headers[idx0] if idx0 < len(headers) else ""
                var.set(f"{col_letter}: {h}" if h else f"{col_letter}: (boş)")

        for item_id in tree.get_children():
            tree.delete(item_id)

        max_col = min(ws.max_column or 1, max(1, len(headers)))
        cols = [f"C{c}" for c in range(1, max_col + 1)]
        tree["columns"] = cols
        for col_index, c in enumerate(cols, start=1):
            h = headers[col_index - 1] if col_index - 1 < len(headers) else ""
            col_letter = self.openpyxl.utils.get_column_letter(col_index)
            tree.heading(c, text=f"{col_letter} {h}".strip())
            tree.column(c, width=140, anchor="w")

        # Maaş tabında "sabit çalışan" modu ile satır filtreleme varsa, sadece o satırları göster.
        preview_rows: List[int] = []
        if tab_name == "MaasOdeme" and self._get_info_bool(info, "maas_fixed_mode_var"):
            rlist = list(info.get("maas_rowmatch_rows") or [])
            if rlist:
                preview_rows = [int(x) for x in rlist if int(x) > header_row]

        if not preview_rows:
            start = header_row + 1
            end = min(ws.max_row or start, start + 49)
            preview_rows = list(range(start, end + 1))

        # row map: Treeview item -> Excel row index
        info["preview_row_map"] = []
        for r in preview_rows[:200]:
            vals = []
            for col_idx in range(1, max_col + 1):
                v = ws.cell(r, col_idx).value
                if isinstance(v, (datetime, date)):
                    vals.append(v.strftime("%Y-%m-%d"))
                else:
                    s = "" if v is None else str(v)
                    vals.append(s[:60] + ("…" if len(s) > 60 else ""))
            try:
                iid = str(r)
                tree.insert("", tk.END, iid=iid, values=vals)
            except Exception:
                tree.insert("", tk.END, values=vals)
            info["preview_row_map"].append(int(r))

        # Önizleme başlığına bilgi ekle
        try:
            if tab_name == "MaasOdeme" and self._get_info_bool(info, "maas_fixed_mode_var"):
                cnt = len(list(info.get("maas_rowmatch_rows") or []))
                prev_title = f"Önizleme (Bulunan satırlar: {cnt})" if cnt else "Önizleme"
            else:
                prev_title = "Önizleme"
            lf = tree.master
            if isinstance(lf, ttk.LabelFrame):
                lf.configure(text=prev_title)
        except Exception:
            pass

    def _toggle_maas_fixed_mode(self, tab_name: str):
        """Geriye dönük uyumluluk için bırakıldı.

        Bu sürümde Maaş importında çalışan seçimi Excel kolonu üzerinden değil,
        sistemde kayıtlı çalışan üzerinden yapılır; bu yüzden toggle fiilen kullanılmaz.
        """
        info = self.mappings.get(tab_name) or {}
        if tab_name != "MaasOdeme":
            return
        try:
            info["maas_rowmatch_rows"] = []
        except Exception:
            pass
        self._refresh_preview(tab_name)


    def _find_employee_rows(self, tab_name: str):
        """Seçili 'sabit çalışan' adını Excel sayfasında arar ve eşleşen satırları önizlemeye getirir."""
        info = self.mappings.get(tab_name) or {}
        if tab_name != "MaasOdeme":
            return

        emp_var = info.get("maas_fixed_emp_var")
        emp_name = str(emp_var.get() if emp_var else "").strip()
        if (not emp_name) or emp_name.startswith("("):
            messagebox.showinfo(APP_TITLE, "Önce Maaş > Çalışanlar bölümüne çalışan ekleyin ve burada seçin.")
            return

        sheet_name = self._get_info_str(info, "sheet_var")
        if not sheet_name or sheet_name == "(Atla)":
            messagebox.showinfo(APP_TITLE, "Önce bir Excel sayfası seç.")
            return

        try:
            ws = self.wb[sheet_name]
        except Exception:
            messagebox.showerror(APP_TITLE, "Sayfa okunamadı.")
            return

        hr = int(info.get("header_row") or 1)
        headers: List[str] = list(info.get("headers") or [])
        max_col = min(ws.max_column or 1, max(1, len(headers)), 120)
        max_row = min(ws.max_row or (hr + 1), hr + 8000)

        emp_full = str(emp_name)
        emp_norm = normalize_text(emp_full)
        parts = [p for p in re.split(r"\s+", emp_full.strip()) if p]
        first = parts[0] if parts else emp_full
        last = parts[-1] if len(parts) >= 2 else ""
        first_n = normalize_text(first)
        last_n = normalize_text(last)

        # Eşikler
        thr_full = 0.82 if len(parts) >= 2 else 0.92
        thr_tok = 0.92

        matches: List[Tuple[int, float]] = []

        for r in range(hr + 1, max_row + 1):
            best_sc = 0.0
            row_tokens_hit = 0
            row_text_norm_parts: List[str] = []
            for c in range(1, max_col + 1):
                v = ws.cell(r, c).value
                if v is None:
                    continue
                s = str(v).strip()
                if not s:
                    continue
                ns = normalize_text(s)
                row_text_norm_parts.append(ns)
                if emp_norm and emp_norm in ns:
                    best_sc = 1.0
                    break
                if first_n and first_n in ns:
                    row_tokens_hit = max(row_tokens_hit, 1)
                    best_sc = max(best_sc, 0.96)
                if last_n and last_n in ns:
                    row_tokens_hit = max(row_tokens_hit, 2)
                    best_sc = max(best_sc, 0.96)

                # fuzzy sadece gerektiğinde
                if best_sc < 0.90:
                    try:
                        best_sc = max(best_sc, float(best_substring_similarity(emp_full, s)))
                        if first and best_sc < 0.90:
                            best_sc = max(best_sc, float(best_substring_similarity(first, s)))
                        if last and best_sc < 0.90:
                            best_sc = max(best_sc, float(best_substring_similarity(last, s)))
                    except Exception:
                        pass

            if best_sc >= 0.995:
                matches.append((r, best_sc))
                continue

            # Satırın genel metninde (kolon bazlı değil) hızlı içerme
            if best_sc < 0.90 and row_text_norm_parts:
                whole = " ".join(row_text_norm_parts)
                if emp_norm and emp_norm in whole:
                    best_sc = max(best_sc, 0.99)
                elif first_n and last_n and (first_n in whole and last_n in whole):
                    best_sc = max(best_sc, 0.97)

            if len(parts) >= 2:
                ok = (best_sc >= thr_full) or (first_n and last_n and row_tokens_hit >= 2)
            else:
                ok = (best_sc >= thr_tok) or (first_n and row_tokens_hit >= 1)

            if ok:
                matches.append((r, best_sc))

        # Skora göre sırala (yüksek skor üstte), eşitlikte excel satır sırası
        matches.sort(key=lambda x: (-x[1], x[0]))
        rows = [int(r) for r, _sc in matches]
        info["maas_rowmatch_rows"] = rows
        self._refresh_preview(tab_name)

        # Varsayılan seçim: en iyi eşleşen 1 satır.
        try:
            tree = info.get("tree")
            if isinstance(tree, ttk.Treeview) and rows:
                iid = str(rows[0])
                tree.selection_set(iid)
                tree.see(iid)
        except Exception:
            pass

        if len(rows) <= 1:
            messagebox.showinfo(APP_TITLE, f"'{emp_name}' için Excel'de {len(rows)} satır bulundu.")
        else:
            messagebox.showinfo(APP_TITLE, f"'{emp_name}' için Excel'de {len(rows)} satır bulundu. Varsayılan olarak en iyi 1 satır seçildi; istersen birden fazla satır seçip içe aktarabilirsin (toplayarak kaydeder).")

    def _find_and_set_employee_column(self, tab_name: str):
        """Maaş importında, sistemdeki bir çalışan adını Excel içinde arayıp
        'Çalışan' kolon seçimini (mapping) otomatik doldurur."""
        info = self.mappings.get(tab_name) or {}
        emp_var = info.get("maas_emp_var")
        if not emp_var:
            messagebox.showinfo(APP_TITLE, "Sistem çalışan listesi bulunamadı.")
            return

        emp_name = str(emp_var.get() or "").strip()
        if not emp_name:
            messagebox.showinfo(APP_TITLE, "Bir çalışan seç.")
            return

        sheet_name = self._get_info_str(info, "sheet_var")
        if not sheet_name or sheet_name == "(Atla)":
            messagebox.showinfo(APP_TITLE, "Önce bir Excel sayfası seç.")
            return

        try:
            ws = self.wb[sheet_name]
        except Exception:
            messagebox.showerror(APP_TITLE, "Sayfa okunamadı.")
            return

        hr = int(info.get("header_row") or 1)
        headers: List[str] = list(info.get("headers") or [])
        max_col = min(ws.max_column or 1, max(1, len(headers)), 80)
        max_row = min(ws.max_row or (hr + 1), hr + 250)

        # Eşik: tek kelime adlarda daha katı
        nemp = normalize_text(emp_name)
        thr = 0.92 if len(nemp.split()) < 2 else 0.82

        best_sc = 0.0
        best_col: Optional[int] = None

        for r in range(hr + 1, max_row + 1):
            for c in range(1, max_col + 1):
                v = ws.cell(r, c).value
                if v is None:
                    continue
                s = str(v).strip()
                if not s:
                    continue
                # Hız: önce basit içerme, sonra fuzzy
                if nemp and nemp in normalize_text(s):
                    sc = 1.0
                else:
                    sc = float(best_substring_similarity(emp_name, s))
                if sc > best_sc:
                    best_sc = sc
                    best_col = c
                if best_sc >= 0.995:
                    break
            if best_sc >= 0.995:
                break

        if not best_col or best_sc < thr:
            messagebox.showinfo(APP_TITLE, f"Excel içinde '{emp_name}' bulunamadı. (skor={best_sc:.2f})")
            return

        try:
            col_letter = self.openpyxl.utils.get_column_letter(best_col)
        except Exception:
            col_letter = str(best_col)
        header = headers[best_col - 1] if (best_col - 1) < len(headers) else ""
        choice = f"{col_letter}: {header}" if header else f"{col_letter}: (boş)"

        mv = info.get("mapping_vars") or {}
        if "calisan" not in mv:
            messagebox.showinfo(APP_TITLE, "Bu sekmede 'Çalışan' alanı yok.")
            return
        try:
            mv["calisan"].set(choice)
        except Exception:
            pass

        messagebox.showinfo(APP_TITLE, f"Çalışan kolonu bulundu: {choice}  (skor={best_sc:.2f})")

    def _resolve_employee_name(self, excel_name: str) -> str:
        """Excel'den gelen çalışan adını, sistemdeki çalışanlar ile fuzzy eşleştirerek
        mümkünse tekilleştirir. Böylece 1 harf farkı / büyük-küçük gibi farklar nedeniyle
        yeni (mükerrer) çalışan kaydı açılmasının önüne geçer."""

        name = str(excel_name or "").strip()
        if not name:
            return name

        # Cache (import boyunca tek kez çek)
        if not hasattr(self, "_maas_emp_cache"):
            try:
                rows = self.db.maas_calisan_list(q="", only_active=False)  # type: ignore
                cache: List[str] = []
                for rr in (rows or []):
                    ad_val = None
                    try:
                        ad_val = rr["ad"]
                    except Exception:
                        try:
                            ad_val = rr[0]
                        except Exception:
                            ad_val = None
                    if ad_val is None:
                        continue
                    s = str(ad_val).strip()
                    if s:
                        cache.append(s)
                self._maas_emp_cache = cache
            except Exception:
                self._maas_emp_cache = []

        emp_names: List[str] = list(getattr(self, "_maas_emp_cache", []) or [])
        if not emp_names:
            return name

        nn = normalize_text(name)
        thr = 0.92 if len(nn.split()) < 2 else 0.82

        best_name = ""
        best_sc = 0.0
        for e in emp_names:
            try:
                # İsimler için SequenceMatcher bazlı benzerlik + substring desteği
                sc = float(similarity(nn, normalize_text(e)))
                if sc < 0.60:
                    sc = float(best_substring_similarity(name, e))
            except Exception:
                continue
            if sc > best_sc:
                best_sc = sc
                best_name = str(e)

        if best_name and best_sc >= thr:
            return best_name
        return name

    def _parse_col_choice(self, choice: str) -> Optional[int]:
        if not choice or choice == "(Yok)":
            return None
        m = re.match(r"^([A-Z]{1,3}):", choice.strip())
        if not m:
            return None
        return self.openpyxl.utils.column_index_from_string(m.group(1))

    def _make_find_employee_rows_command(self, tab_name: str) -> Callable[[], None]:
        return lambda: self._find_employee_rows(tab_name)

    @staticmethod
    def _get_info_bool(info: Dict[str, Any], key: str) -> bool:
        var = info.get(key)
        return bool(var.get()) if isinstance(var, tk.BooleanVar) else False

    @staticmethod
    def _get_info_str(info: Dict[str, Any], key: str) -> str:
        var = info.get(key)
        return str(var.get() or "") if isinstance(var, tk.StringVar) else ""

    def _cancel(self):
        if self._import_in_progress:
            return
        self.result_counts = None
        self.destroy()

    def _do_import(self):
        if self._import_in_progress:
            return
        plan = self._build_import_plan()
        if plan is None:
            return
        self._start_background_import(plan)

    def _build_import_plan(self) -> Optional[Dict[str, Dict[str, Any]]]:
        plan: Dict[str, Dict[str, Any]] = {}
        for tab_name, info in self.mappings.items():
            cols = {}
            for key, _label in info["fields"]:
                cols[key] = self._parse_col_choice(info["mapping_vars"][key].get())
            tab_plan: Dict[str, Any] = {"sheet": info["sheet_var"].get(), "header_row": info["header_row"], "cols": cols}

            # Maaş: sabit çalışan + satır seçimi modundan gelen bilgileri plana ekle
            if tab_name == "MaasOdeme":
                fixed_on = self._get_info_bool(info, "maas_fixed_mode_var")
                tab_plan["fixed_mode"] = fixed_on
                tab_plan["fixed_employee"] = str(info.get("maas_fixed_emp_var").get() if info.get("maas_fixed_emp_var") else "").strip()
                tab_plan["rowmatch_rows"] = list(info.get("maas_rowmatch_rows") or [])
                only_sel = bool(info.get("maas_only_selected_rows_var").get() if info.get("maas_only_selected_rows_var") else False)
                tab_plan["only_selected"] = only_sel
                if fixed_on and only_sel:
                    try:
                        tree: ttk.Treeview = info.get("tree")
                        sel = [int(iid) for iid in (tree.selection() or []) if str(iid).isdigit()]
                    except Exception:
                        sel = []
                    tab_plan["selected_rows"] = sel

            plan[tab_name] = tab_plan

        ch_plan = plan.get("CariHareket")
        if ch_plan and ch_plan.get("sheet") and ch_plan.get("sheet") != "(Atla)":
            if plan["CariHareket"]["cols"].get("cari") is None or plan["CariHareket"]["cols"].get("tutar") is None:
                messagebox.showerror(APP_TITLE, "CariHareket için en az 'Cari' ve 'Tutar' kolonlarını seçmelisin.")
                return None

        kh_plan = plan.get("KasaHareket")
        if kh_plan and kh_plan.get("sheet") and kh_plan.get("sheet") != "(Atla)":
            if plan["KasaHareket"]["cols"].get("tutar") is None:
                messagebox.showerror(APP_TITLE, "KasaHareket için en az 'Tutar' kolonunu seçmelisin.")
                return None

        bh_plan = plan.get("BankaHareket")
        if bh_plan and bh_plan.get("sheet") and bh_plan.get("sheet") != "(Atla)":
            cols = plan["BankaHareket"]["cols"]
            if cols.get("tutar") is None and cols.get("borc") is None and cols.get("alacak") is None:
                messagebox.showerror(APP_TITLE, "BankaHareket için en az 'Tutar' veya 'Borç/Alacak' kolonlarını seçmelisin.")
                return None

        mp_plan = plan.get("MaasOdeme")
        if mp_plan and mp_plan.get("sheet") and mp_plan.get("sheet") != "(Atla)":
            cols = plan["MaasOdeme"]["cols"]
            fixed_on = bool(mp_plan.get("fixed_mode"))
            if fixed_on:
                if cols.get("tutar") is None:
                    messagebox.showerror(APP_TITLE, "MaasOdeme için en az 'Tutar' kolonunu seçmelisin.")
                    return None
                if not str(mp_plan.get("fixed_employee") or "").strip():
                    messagebox.showerror(APP_TITLE, "Bir çalışan seçmelisin.")
                    return None

                only_sel = bool(mp_plan.get("only_selected"))
                rowmatch = list(mp_plan.get("rowmatch_rows") or [])
                selected = list(mp_plan.get("selected_rows") or [])

                # Kullanıcı satırları getirmeden içe aktarmasın
                if only_sel:
                    if not selected:
                        messagebox.showerror(APP_TITLE, "Sadece seçili satırlar içe aktar seçili ama tabloda satır seçmedin.")
                        return None
                else:
                    if not rowmatch:
                        messagebox.showerror(APP_TITLE, "Önce çalışanı seçip 'Excel'de Ara' ile satırı/satırları getir.")
                        return None

            else:
                if cols.get("calisan") is None or cols.get("tutar") is None:
                    messagebox.showerror(APP_TITLE, "MaasOdeme için en az 'Çalışan' ve 'Tutar' kolonlarını seçmelisin.")
                    return None
            donem = str(self.context.get("donem") or "").strip()
            if not donem:
                messagebox.showerror(APP_TITLE, "Maaş import için dönem (YYYY-MM) seçilmemiş.")
                return None

        return plan

    def _start_background_import(self, plan: Dict[str, Dict[str, Any]]) -> None:
        self._import_in_progress = True
        self._set_busy(True)
        create_missing = bool(self.var_create_missing_cari.get())
        db_path = getattr(self.app.db, "path", "")
        logger = logging.getLogger(__name__)

        def worker() -> None:
            try:
                db = DB(db_path)
                counts = self._run_import(plan, create_missing_cari=create_missing, db=db)
                db.close()
                self.after(0, lambda: self._finish_import_success(counts))
            except Exception as exc:
                logger.exception("Excel import failed")
                self.after(0, lambda: self._finish_import_error(exc))

        threading.Thread(target=worker, daemon=True).start()

    def _set_busy(self, is_busy: bool) -> None:
        try:
            if hasattr(self, "btn_import"):
                self.btn_import.config(state=tk.DISABLED if is_busy else tk.NORMAL)
            if hasattr(self, "btn_cancel"):
                self.btn_cancel.config(state=tk.DISABLED if is_busy else tk.NORMAL)
            self.config(cursor="watch" if is_busy else "")
        except Exception:
            pass

    def _finish_import_success(self, counts: Dict[str, int]) -> None:
        self._import_in_progress = False
        self._set_busy(False)
        self.result_counts = counts
        msg_lines = []
        if "cariler" in counts:
            msg_lines.append(f"Cariler: {counts['cariler']}")
        if "cari_hareket" in counts:
            msg_lines.append(f"Cari Hareket: {counts['cari_hareket']}")
        if "kasa" in counts:
            msg_lines.append(f"Kasa: {counts['kasa']}")
        if "banka" in counts:
            msg_lines.append(f"Banka: {counts['banka']}")
        if "maas" in counts:
            msg_lines.append(f"Maaş: {counts['maas']}")
        messagebox.showinfo(APP_TITLE, "İçe aktarıldı:\n" + "\n".join(msg_lines))
        self.destroy()

    def _finish_import_error(self, exc: Exception) -> None:
        self._import_in_progress = False
        self._set_busy(False)
        messagebox.showerror(APP_TITLE, f"İçe aktarma sırasında hata oluştu:\n{exc}")

    def _run_import(self, plan: Dict[str, Dict[str, Any]], create_missing_cari: bool, db: Optional[DB] = None) -> Dict[str, int]:
        db = db or self.db
        counts = {"cariler": 0, "cari_hareket": 0, "kasa": 0, "banka": 0, "maas": 0}
        bank_ids: List[int] = []

        def cell(ws, r, col):
            if col is None:
                return None
            return ws.cell(r, col).value

        car = plan.get("Cariler")
        if car and car.get("sheet") and car.get("sheet") != "(Atla)":
            ws = self.wb[car["sheet"]]
            hr = int(car["header_row"])
            c = car["cols"]
            if c.get("ad"):
                for r in range(hr + 1, (ws.max_row or hr) + 1):
                    ad = cell(ws, r, c.get("ad"))
                    if not ad or not str(ad).strip():
                        continue
                    tur = str(cell(ws, r, c.get("tur")) or "")
                    tel = str(cell(ws, r, c.get("telefon")) or "")
                    notlar = str(cell(ws, r, c.get("notlar")) or "")
                    acilis = safe_float(cell(ws, r, c.get("acilis")))
                    db.cari_upsert(str(ad), tur, tel, notlar, acilis)
                    counts["cariler"] += 1

        ch = plan.get("CariHareket")
        if ch and ch.get("sheet") and ch.get("sheet") != "(Atla)":
            ws = self.wb[ch["sheet"]]
            hr = int(ch["header_row"])
            c = ch["cols"]
            for r in range(hr + 1, (ws.max_row or hr) + 1):
                cari = cell(ws, r, c.get("cari"))
                tutar = cell(ws, r, c.get("tutar"))
                if not cari or not str(cari).strip():
                    continue
                if safe_float(tutar) == 0:
                    continue

                cari_name = str(cari).strip()
                c_row = db.cari_get_by_name(cari_name)
                if not c_row:
                    if not create_missing_cari:
                        continue
                    cari_id = db.cari_upsert(cari_name)
                else:
                    cari_id = int(c_row["id"])

                tarih = cell(ws, r, c.get("tarih"))
                tip_raw = str(cell(ws, r, c.get("tip")) or "Borç")
                tipn = "Alacak" if "alacak" in norm_header(tip_raw) else "Borç"
                para = str(cell(ws, r, c.get("para")) or "TL")
                odeme = str(cell(ws, r, c.get("odeme")) or "")
                belge = str(cell(ws, r, c.get("belge")) or "")
                if not belge.strip():
                    belge = db.next_belge_no("C")
                etiket = str(cell(ws, r, c.get("etiket")) or "")
                aciklama = str(cell(ws, r, c.get("aciklama")) or "")

                db.cari_hareket_add(tarih, cari_id, tipn, safe_float(tutar), para, aciklama, odeme, belge, etiket)
                counts["cari_hareket"] += 1

        kh = plan.get("KasaHareket")
        if kh and kh.get("sheet") and kh.get("sheet") != "(Atla)":
            ws = self.wb[kh["sheet"]]
            hr = int(kh["header_row"])
            c = kh["cols"]
            for r in range(hr + 1, (ws.max_row or hr) + 1):
                tutar = cell(ws, r, c.get("tutar"))
                if safe_float(tutar) == 0:
                    continue

                tarih = cell(ws, r, c.get("tarih"))
                tip_raw = str(cell(ws, r, c.get("tip")) or "Gider")
                tipn = "Gelir" if "gelir" in norm_header(tip_raw) else "Gider"
                para = str(cell(ws, r, c.get("para")) or "TL")
                odeme = str(cell(ws, r, c.get("odeme")) or "")
                kategori = str(cell(ws, r, c.get("kategori")) or "")
                belge = str(cell(ws, r, c.get("belge")) or "")
                if not belge.strip():
                    belge = db.next_belge_no("K")
                etiket = str(cell(ws, r, c.get("etiket")) or "")
                aciklama = str(cell(ws, r, c.get("aciklama")) or "")

                cari_id = None
                cari = cell(ws, r, c.get("cari"))
                if cari and str(cari).strip():
                    cari_name = str(cari).strip()
                c_row = db.cari_get_by_name(cari_name)
                if not c_row:
                    if create_missing_cari:
                        cari_id = db.cari_upsert(cari_name)
                else:
                    cari_id = int(c_row["id"])

                db.kasa_add(tarih, tipn, safe_float(tutar), para, odeme, kategori, cari_id, aciklama, belge, etiket)
                counts["kasa"] += 1

        # Banka hareketleri
        bh = plan.get("BankaHareket")
        if bh and bh.get("sheet") and bh.get("sheet") != "(Atla)":
            ws = self.wb[bh["sheet"]]
            hr = int(bh["header_row"])
            c = bh["cols"]
            import_grup = f"{now_iso()} | {os.path.basename(self.xlsx_path)}"
            for r in range(hr + 1, (ws.max_row or hr) + 1):
                tarih = cell(ws, r, c.get("tarih"))
                banka = str(cell(ws, r, c.get("banka")) or "")
                hesap = str(cell(ws, r, c.get("hesap")) or "")
                aciklama = str(cell(ws, r, c.get("aciklama")) or "")
                para = str(cell(ws, r, c.get("para")) or "TL")
                referans = str(cell(ws, r, c.get("referans")) or "")
                belge = str(cell(ws, r, c.get("belge")) or "")
                if not belge.strip():
                    belge = db.next_belge_no("B")
                etiket = str(cell(ws, r, c.get("etiket")) or "")
                bakiye = cell(ws, r, c.get("bakiye"))

                # tutar/borç/alacak toleranslı
                alacak = safe_float(cell(ws, r, c.get("alacak")))
                borc = safe_float(cell(ws, r, c.get("borc")))
                tutar_val = safe_float(cell(ws, r, c.get("tutar")))

                tipn = ""
                amount = 0.0
                if alacak != 0:
                    tipn = "Giriş"
                    amount = abs(alacak)
                elif borc != 0:
                    tipn = "Çıkış"
                    amount = abs(borc)
                elif tutar_val != 0:
                    tipn = "Çıkış" if tutar_val < 0 else "Giriş"
                    amount = abs(tutar_val)
                else:
                    continue

                hid = db.banka_add(

                    tarih,
                    banka=banka,
                    hesap=hesap,
                    tip=tipn,
                    tutar=amount,
                    para=para,
                    aciklama=aciklama,
                    referans=referans,
                    belge=belge,
                    etiket=etiket,
                    import_grup=import_grup,
                    bakiye=(None if bakiye is None else safe_float(bakiye)),
                )
                bank_ids.append(int(hid))
                counts["banka"] += 1

            # UI'nin sonradan tek tıkla çağırabilmesi için sakla
            self.last_import_bank_group = import_grup

        # Maaş ödemeleri
        mp = plan.get("MaasOdeme")
        if mp and mp.get("sheet") and mp.get("sheet") != "(Atla)":
            donem = str(self.context.get("donem") or "").strip()
            ws = self.wb[mp["sheet"]]
            hr = int(mp["header_row"])
            c = mp["cols"]
            fixed_on = bool(mp.get("fixed_mode"))
            fixed_emp = str(mp.get("fixed_employee") or "").strip()

            # Hangi satırlar içe aktarılacak?
            if fixed_on and bool(mp.get("only_selected")):
                rows_to_import = [int(x) for x in (mp.get("selected_rows") or []) if int(x) > hr]
            elif fixed_on and list(mp.get("rowmatch_rows") or []):
                # Otomatik bulunan satırlarda varsayılan: en iyi 1 satır
                rr = [int(x) for x in (mp.get("rowmatch_rows") or []) if int(x) > hr]
                rows_to_import = rr[:1]
            else:
                rows_to_import = []

            def _paid(v: Any) -> int:
                if v is None:
                    return 0
                if isinstance(v, (int, float)):
                    return 1 if float(v) != 0 else 0
                s = str(v).strip().lower()
                if s in ("1", "true", "yes", "y", "evet", "e", "odendi", "ödendi", "paid"):
                    return 1
                if s in ("0", "false", "no", "n", "hayir", "hayır"):
                    return 0
                return 0

            # Sabit çalışan modunda (sistem çalışanı seçili):
            # - Tek satır seçiliyse normal import
            # - Birden fazla satır seçiliyse tutarı toplayıp tek kayıt olarak kaydeder (maas_odeme UNIQUE olduğu için)
            if fixed_on and bool(mp.get("only_selected")) and len(rows_to_import) > 1:
                total = 0.0
                para = "TL"
                odendi = 0
                best_date = ""
                notes = []
                for r in rows_to_import:
                    total += safe_float(cell(ws, r, c.get("tutar")))
                    p = str(cell(ws, r, c.get("para")) or "").strip()
                    if p:
                        para = p
                    odendi = max(odendi, _paid(cell(ws, r, c.get("odendi"))))
                    dt_raw = cell(ws, r, c.get("odeme_tarihi"))
                    dt_iso = parse_date_smart(dt_raw) if dt_raw else ""
                    if dt_iso and dt_iso > best_date:
                        best_date = dt_iso
                    ac = str(cell(ws, r, c.get("aciklama")) or "").strip()
                    if ac:
                        notes.append(ac)

                if fixed_emp and total != 0:
                    ad_norm = self._resolve_employee_name(str(fixed_emp).strip())
                    aciklama = " | ".join(list(dict.fromkeys(notes)))
                    db.maas_odeme_upsert_from_excel(
                        donem,
                        str(ad_norm).strip(),
                        float(total),
                        para=para or "TL",
                        odendi=int(odendi or 0),
                        odeme_tarihi=best_date,
                        aciklama=aciklama,
                    )
                    counts["maas"] += 1
            else:
                for r in rows_to_import:
                    ad = fixed_emp if fixed_on else cell(ws, r, c.get("calisan"))
                    tutar = safe_float(cell(ws, r, c.get("tutar")))
                    if not ad or not str(ad).strip():
                        continue
                    if tutar == 0:
                        continue
                    ad_norm = self._resolve_employee_name(str(ad).strip())
                    para = str(cell(ws, r, c.get("para")) or "TL")
                    odendi = _paid(cell(ws, r, c.get("odendi")))
                    odeme_tarihi = cell(ws, r, c.get("odeme_tarihi"))
                    aciklama = str(cell(ws, r, c.get("aciklama")) or "")
                    db.maas_odeme_upsert_from_excel(
                        donem,
                        str(ad_norm).strip(),
                        float(tutar),
                        para=para,
                        odendi=odendi,
                        odeme_tarihi=odeme_tarihi,
                        aciklama=aciklama,
                    )
                    counts["maas"] += 1

        db.log("Excel Import Wizard", f"{os.path.basename(self.xlsx_path)} | {counts}")
        self.last_import_bank_ids = list(bank_ids)
        return counts


# =========================
# CARİ EKSTRE WINDOW
# =========================
