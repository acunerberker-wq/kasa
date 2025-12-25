# -*- coding: utf-8 -*-
"""Satın Alma Sipariş Raporları."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, timedelta
import csv
import queue
import threading
from typing import Any, Dict, List, Optional, Tuple, TYPE_CHECKING

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

from ...config import APP_TITLE, HAS_OPENPYXL, HAS_REPORTLAB
from ...db.main_db import DB
from ...utils import fmt_amount, fmt_tr_date, parse_date_smart, ensure_pdf_fonts
from ..base import BaseView
from ..widgets import LabeledCombo, LabeledEntry

if TYPE_CHECKING:
    from ...app import App


STATUS_LIST = ["oluşturuldu", "onaylandı", "kısmi teslim alındı", "teslim alındı", "iptal"]


@dataclass
class ReportSpec:
    key: str
    title: str
    columns: List[Tuple[str, str]]


REPORTS = {
    "acik": ReportSpec(
        key="acik",
        title="Açık Satın Alma Siparişleri (Tedarikçi Bazlı)",
        columns=[
            ("tedarikci", "Tedarikçi"),
            ("para", "Para"),
            ("siparis_adet", "Sipariş"),
            ("acik_tutar", "Açık Tutar"),
            ("tl_tutar", "TL Karşılığı"),
        ],
    ),
    "performans": ReportSpec(
        key="performans",
        title="Teslim Alma Performansı",
        columns=[
            ("siparis_no", "Sipariş No"),
            ("tarih", "Tarih"),
            ("teslim_tarihi", "Planlanan"),
            ("son_teslim", "Son Teslim"),
            ("tedarikci", "Tedarikçi"),
            ("durum", "Durum"),
            ("miktar", "Sipariş Miktar"),
            ("teslim", "Teslim Miktar"),
            ("gecikme", "Gecikme (gün)"),
            ("kismi", "Kısmi"),
        ],
    ),
    "maliyet": ReportSpec(
        key="maliyet",
        title="Beklenen Maliyet (Açık PO Toplamı)",
        columns=[
            ("siparis_no", "Sipariş No"),
            ("tarih", "Tarih"),
            ("tedarikci", "Tedarikçi"),
            ("para", "Para"),
            ("acik_tutar", "Açık Tutar"),
            ("kur", "Kur"),
            ("tl_tutar", "TL Karşılığı"),
            ("iskonto", "İskonto"),
            ("durum", "Durum"),
        ],
    ),
    "eslesme": ReportSpec(
        key="eslesme",
        title="Mal Kabul → Fatura Eşleşme Kontrolü",
        columns=[
            ("teslim_id", "Teslim ID"),
            ("tarih", "Tarih"),
            ("siparis_no", "Sipariş No"),
            ("tedarikci", "Tedarikçi"),
            ("depo", "Depo"),
            ("fatura_no", "Fatura No"),
            ("fatura_durum", "Fatura Durum"),
            ("eslesme", "Eşleşme"),
        ],
    ),
}


class SatinAlmaRaporlarFrame(BaseView):
    def __init__(self, master, app: "App"):
        self.app = app
        super().__init__(master, app)
        self.report_trees: Dict[str, ttk.Treeview] = {}
        self.report_labels: Dict[str, ttk.Label] = {}
        self.report_rows: Dict[str, List[Dict[str, Any]]] = {}
        self.report_headers: Dict[str, List[str]] = {}
        self.report_keys: Dict[str, List[str]] = {}
        self._queue: "queue.Queue[Tuple[str, Any]]" = queue.Queue()
        self._worker_active = False
        self.build_ui()

    def build_ui(self) -> None:
        self._build()

    def _build(self):
        top = ttk.LabelFrame(self, text="Filtreler")
        top.pack(fill=tk.X, padx=10, pady=(10, 6))

        row1 = ttk.Frame(top)
        row1.pack(fill=tk.X, pady=4)
        self.f_tedarikci = LabeledCombo(row1, "Tedarikçi:", ["(Tümü)"], 28)
        self.f_tedarikci.pack(side=tk.LEFT, padx=6)
        self.f_urun = LabeledCombo(row1, "Ürün:", ["(Tümü)"], 28)
        self.f_urun.pack(side=tk.LEFT, padx=6)
        self.f_durum = LabeledCombo(row1, "Durum:", ["(Tümü)"] + STATUS_LIST, 18)
        self.f_durum.pack(side=tk.LEFT, padx=6)
        self.f_depo = LabeledCombo(row1, "Depo:", ["(Tümü)"], 20)
        self.f_depo.pack(side=tk.LEFT, padx=6)

        row2 = ttk.Frame(top)
        row2.pack(fill=tk.X, pady=4)
        self.f_from = LabeledEntry(row2, "Başlangıç:", 12)
        self.f_from.pack(side=tk.LEFT, padx=6)
        self.f_to = LabeledEntry(row2, "Bitiş:", 12)
        self.f_to.pack(side=tk.LEFT, padx=6)

        self.lbl_company = ttk.Label(row2, text="")
        self.lbl_company.pack(side=tk.LEFT, padx=(12, 6))

        ttk.Button(row2, text="🔄 Filtreleri Yenile", command=self._reload_filters).pack(side=tk.LEFT, padx=6)
        ttk.Button(row2, text="📊 Raporu Çalıştır", command=self.run_report).pack(side=tk.LEFT, padx=6)

        self.progress = ttk.Progressbar(top, mode="indeterminate")
        self.progress.pack(fill=tk.X, padx=8, pady=(6, 2))

        summary = ttk.LabelFrame(self, text="Otomatik Özet (Son 7 Gün)")
        summary.pack(fill=tk.X, padx=10, pady=(0, 6))
        self.lbl_summary = ttk.Label(summary, text="")
        self.lbl_summary.pack(anchor="w", padx=8, pady=6)

        self.nb = ttk.Notebook(self)
        self.nb.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        for spec in REPORTS.values():
            tab = ttk.Frame(self.nb)
            self.nb.add(tab, text=spec.title)
            self._build_report_tab(tab, spec)

        self._reload_filters()
        self._refresh_summary()

    def refresh(self, data=None):
        self._reload_filters()
        self._refresh_summary()

    def _reload_filters(self):
        tedarikciler = ["(Tümü)"]
        for r in self.app.db.cari_list(only_active=True):
            tedarikciler.append(f"{r['id']} - {r['ad']}")
        self.f_tedarikci.cmb.config(values=tedarikciler)
        if not self.f_tedarikci.get():
            self.f_tedarikci.set("(Tümü)")

        urunler = ["(Tümü)"]
        for r in self.app.db.stok_urun_list(only_active=True):
            urunler.append(f"{r['id']} - {r['ad']}")
        self.f_urun.cmb.config(values=urunler)
        if not self.f_urun.get():
            self.f_urun.set("(Tümü)")

        depolar = ["(Tümü)"]
        for r in self.app.db.stok_lokasyon_list(only_active=True):
            depolar.append(f"{r['id']} - {r['ad']}")
        self.f_depo.cmb.config(values=depolar)
        if not self.f_depo.get():
            self.f_depo.set("(Tümü)")

        cname = getattr(self.app, "active_company_name", "") or ""
        self.lbl_company.config(text=f"Şirket: {cname}")

    def _refresh_summary(self):
        try:
            dt_to = date.today()
            dt_from = dt_to - timedelta(days=7)
            df = dt_from.strftime("%d.%m.%Y")
            dt = dt_to.strftime("%d.%m.%Y")

            siparisler = self.app.db.satin_alma_siparis_list(date_from=df, date_to=dt)
            teslimler = self.app.db.satin_alma_teslim_list(date_from=df, date_to=dt)
            ids = [int(r["id"]) for r in siparisler]
            kalem_totals = self.app.db.satin_alma_siparis_kalem_totals(ids)
            teslim_totals = self.app.db.satin_alma_teslim_summary_by_siparis(ids)

            open_total_tl = 0.0
            for r in siparisler:
                if str(r["durum"] or "") == "iptal":
                    continue
                sid = int(r["id"])
                kalem = kalem_totals.get(sid, {"toplam": 0, "iskonto": 0})
                teslim = teslim_totals.get(sid, {"toplam": 0})
                net = float(kalem.get("toplam", 0)) - float(kalem.get("iskonto", 0))
                kalan = net - float(teslim.get("toplam", 0))
                if kalan > 0:
                    open_total_tl += kalan * float(r["kur"] or 1)

            self.lbl_summary.config(
                text=(
                    f"Oluşturulan sipariş: {len(siparisler)}  •  "
                    f"Teslim alma: {len(teslimler)}  •  "
                    f"Açık PO (TL): {fmt_amount(open_total_tl)}"
                )
            )
        except Exception:
            self.lbl_summary.config(text="Özet hazırlanamadı.")

    def _build_report_tab(self, parent: ttk.Frame, spec: ReportSpec):
        toolbar = ttk.Frame(parent)
        toolbar.pack(fill=tk.X, padx=6, pady=(6, 4))

        ttk.Button(toolbar, text="📥 CSV", command=lambda k=spec.key: self.export_csv(k)).pack(side=tk.LEFT)
        btn_excel = ttk.Button(toolbar, text="📊 Excel", command=lambda k=spec.key: self.export_excel(k))
        btn_excel.pack(side=tk.LEFT, padx=6)
        btn_pdf = ttk.Button(toolbar, text="🖨️ Yazdır/PDF", command=lambda k=spec.key: self.export_pdf(k))
        btn_pdf.pack(side=tk.LEFT)

        if not HAS_OPENPYXL:
            btn_excel.config(state="disabled")
        if not HAS_REPORTLAB:
            btn_pdf.config(state="disabled")

        cols = [c for c, _ in spec.columns]
        headers = [h for _, h in spec.columns]
        tree = ttk.Treeview(parent, columns=cols, show="headings", height=14, selectmode="browse")
        for cid, header in spec.columns:
            tree.heading(cid, text=header)
            tree.column(cid, width=120, anchor="w")
        tree.pack(fill=tk.BOTH, expand=True, padx=6, pady=(0, 6))

        scr = ttk.Scrollbar(parent, orient="vertical", command=tree.yview)
        scr.place(in_=tree, relx=1.0, rely=0, relheight=1.0, x=0)
        tree.configure(yscrollcommand=scr.set)

        lbl = ttk.Label(parent, text="")
        lbl.pack(anchor="w", padx=6, pady=(0, 6))

        self.report_trees[spec.key] = tree
        self.report_labels[spec.key] = lbl
        self.report_rows[spec.key] = []
        self.report_headers[spec.key] = headers
        self.report_keys[spec.key] = cols

    def _parse_id(self, val: str) -> Optional[int]:
        if not val or val == "(Tümü)":
            return None
        try:
            return int(str(val).split("-", 1)[0].strip())
        except Exception:
            return None

    def _filters(self) -> Dict[str, Any]:
        return {
            "tedarikci_id": self._parse_id(self.f_tedarikci.get()),
            "urun_id": self._parse_id(self.f_urun.get()),
            "durum": self.f_durum.get(),
            "date_from": self.f_from.get(),
            "date_to": self.f_to.get(),
            "depo_id": self._parse_id(self.f_depo.get()),
        }

    def run_report(self):
        if self._worker_active:
            return
        try:
            tab_id = self.nb.select()
        except Exception:
            tab_id = ""
        report_key = "acik"
        for key, tab in zip(REPORTS.keys(), self.nb.tabs()):
            if tab == tab_id:
                report_key = key
                break

        self._worker_active = True
        self.progress.start(10)
        self._clear_tree(report_key)
        self.report_labels.get(report_key, ttk.Label()).config(text="Rapor hazırlanıyor...")

        filters = self._filters()

        def worker():
            db = DB(self.app.db.path)
            try:
                data = self._generate_report(db, report_key, filters)
                self._queue.put(("done", report_key, data))
            except Exception as e:
                self._queue.put(("error", report_key, str(e)))
            finally:
                db.close()

        threading.Thread(target=worker, daemon=True).start()
        self.after(100, self._poll_queue)

    def _poll_queue(self):
        if not self._worker_active:
            return
        try:
            msg = self._queue.get_nowait()
        except queue.Empty:
            self.after(120, self._poll_queue)
            return

        kind = msg[0]
        if kind == "done":
            _, key, payload = msg
            rows, summary = payload
            self._populate_tree(key, rows, summary)
        elif kind == "error":
            _, key, err = msg
            self.report_labels.get(key, ttk.Label()).config(text="Rapor hazırlanamadı.")
            messagebox.showerror(APP_TITLE, f"Rapor hatası: {err}")

        self.progress.stop()
        self._worker_active = False

    def _clear_tree(self, key: str):
        tree = self.report_trees.get(key)
        if tree:
            tree.delete(*tree.get_children())

    def _populate_tree(self, key: str, rows: List[Dict[str, Any]], summary: str):
        tree = self.report_trees.get(key)
        if not tree:
            return
        tree.delete(*tree.get_children())
        keys = self.report_keys.get(key, [])
        for r in rows:
            tree.insert("", tk.END, values=[r.get(k, "") for k in keys])
        self.report_labels.get(key, ttk.Label()).config(text=summary)
        self.report_rows[key] = rows

    def _generate_report(self, db: DB, key: str, filters: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], str]:
        if key == "acik":
            return self._report_open_orders(db, filters)
        if key == "performans":
            return self._report_performance(db, filters)
        if key == "maliyet":
            return self._report_expected_cost(db, filters)
        if key == "eslesme":
            return self._report_matching(db, filters)
        return [], ""

    def _report_open_orders(self, db: DB, filters: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], str]:
        siparisler = db.satin_alma_siparis_list(**filters)
        ids = [int(r["id"]) for r in siparisler]
        kalem_totals = db.satin_alma_siparis_kalem_totals(ids)
        teslim_totals = db.satin_alma_teslim_summary_by_siparis(ids)

        grouped: Dict[Tuple[str, str], Dict[str, float]] = {}
        for r in siparisler:
            if str(r["durum"] or "") in ("teslim alındı", "iptal"):
                continue
            sid = int(r["id"])
            kalem = kalem_totals.get(sid, {"toplam": 0, "iskonto": 0})
            teslim = teslim_totals.get(sid, {"toplam": 0})
            net = float(kalem.get("toplam", 0)) - float(kalem.get("iskonto", 0))
            kalan = net - float(teslim.get("toplam", 0))
            if kalan <= 0:
                continue
            key_tuple = (str(r["tedarikci_ad"] or ""), str(r["para"] or "TL"))
            if key_tuple not in grouped:
                grouped[key_tuple] = {"siparis": 0, "acik": 0.0, "tl": 0.0}
            grouped[key_tuple]["siparis"] += 1
            grouped[key_tuple]["acik"] += kalan
            grouped[key_tuple]["tl"] += kalan * float(r["kur"] or 1)

        rows = []
        for (tedarikci, para), data in sorted(grouped.items(), key=lambda x: x[0][0]):
            rows.append(
                {
                    "tedarikci": tedarikci,
                    "para": para,
                    "siparis_adet": int(data["siparis"]),
                    "acik_tutar": fmt_amount(data["acik"]),
                    "tl_tutar": fmt_amount(data["tl"]),
                }
            )

        summary = f"Açık sipariş (tedarikçi bazlı): {len(rows)} kayıt"
        return rows, summary

    def _report_performance(self, db: DB, filters: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], str]:
        siparisler = db.satin_alma_siparis_list(**filters)
        ids = [int(r["id"]) for r in siparisler]
        kalem_totals = db.satin_alma_siparis_kalem_totals(ids)
        teslim_totals = db.satin_alma_teslim_summary_by_siparis(ids)

        rows = []
        today = date.today()
        for r in siparisler:
            sid = int(r["id"])
            kalem = kalem_totals.get(sid, {"miktar": 0})
            teslim = teslim_totals.get(sid, {"miktar": 0, "last_tarih": ""})
            teslim_tarih = str(r["teslim_tarihi"] or "")
            last_teslim = str(teslim.get("last_tarih") or "")
            gecikme = ""
            if teslim_tarih:
                try:
                    plan = parse_date_smart(teslim_tarih)
                    plan_date = date.fromisoformat(plan)
                    if last_teslim:
                        last_date = date.fromisoformat(str(last_teslim))
                    else:
                        last_date = today
                    gecikme_val = (last_date - plan_date).days
                    gecikme = str(gecikme_val)
                except Exception:
                    gecikme = ""

            siparis_miktar = float(kalem.get("miktar", 0))
            teslim_miktar = float(teslim.get("miktar", 0))
            kismi = "Evet" if 0 < teslim_miktar < siparis_miktar else "Hayır"

            rows.append(
                {
                    "siparis_no": str(r["siparis_no"] or ""),
                    "tarih": fmt_tr_date(r["tarih"]),
                    "teslim_tarihi": fmt_tr_date(teslim_tarih),
                    "son_teslim": fmt_tr_date(last_teslim),
                    "tedarikci": str(r["tedarikci_ad"] or ""),
                    "durum": str(r["durum"] or ""),
                    "miktar": fmt_amount(siparis_miktar),
                    "teslim": fmt_amount(teslim_miktar),
                    "gecikme": gecikme,
                    "kismi": kismi,
                }
            )

        summary = f"Teslim performansı: {len(rows)} sipariş"
        return rows, summary

    def _report_expected_cost(self, db: DB, filters: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], str]:
        siparisler = db.satin_alma_siparis_list(**filters)
        ids = [int(r["id"]) for r in siparisler]
        kalem_totals = db.satin_alma_siparis_kalem_totals(ids)
        teslim_totals = db.satin_alma_teslim_summary_by_siparis(ids)

        rows = []
        total_tl = 0.0
        for r in siparisler:
            if str(r["durum"] or "") in ("teslim alındı", "iptal"):
                continue
            sid = int(r["id"])
            kalem = kalem_totals.get(sid, {"toplam": 0, "iskonto": 0})
            teslim = teslim_totals.get(sid, {"toplam": 0})
            net = float(kalem.get("toplam", 0)) - float(kalem.get("iskonto", 0))
            kalan = net - float(teslim.get("toplam", 0))
            if kalan <= 0:
                continue
            kur = float(r["kur"] or 1)
            tl_val = kalan * kur
            total_tl += tl_val
            rows.append(
                {
                    "siparis_no": str(r["siparis_no"] or ""),
                    "tarih": fmt_tr_date(r["tarih"]),
                    "tedarikci": str(r["tedarikci_ad"] or ""),
                    "para": str(r["para"] or "TL"),
                    "acik_tutar": fmt_amount(kalan),
                    "kur": fmt_amount(kur),
                    "tl_tutar": fmt_amount(tl_val),
                    "iskonto": fmt_amount(float(kalem.get("iskonto", 0))),
                    "durum": str(r["durum"] or ""),
                }
            )

        summary = f"Beklenen maliyet (TL): {fmt_amount(total_tl)}"
        return rows, summary

    def _report_matching(self, db: DB, filters: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], str]:
        teslimler = db.satin_alma_teslim_list(
            tedarikci_id=filters.get("tedarikci_id"),
            date_from=filters.get("date_from"),
            date_to=filters.get("date_to"),
            depo_id=filters.get("depo_id"),
        )

        rows = []
        for r in teslimler:
            fatura_no = str(r["fatura_no"] or "")
            fatura_durum = str(r["fatura_durum"] or "")
            if fatura_no:
                eslesme = "Eşleşti"
            else:
                eslesme = "Eksik"
            rows.append(
                {
                    "teslim_id": int(r["id"]),
                    "tarih": fmt_tr_date(r["tarih"]),
                    "siparis_no": str(r["siparis_no"] or ""),
                    "tedarikci": str(r["tedarikci_ad"] or ""),
                    "depo": str(r["depo_ad"] or ""),
                    "fatura_no": fatura_no,
                    "fatura_durum": fatura_durum,
                    "eslesme": eslesme,
                }
            )

        summary = f"Mal kabul eşleşme: {len(rows)} kayıt"
        return rows, summary

    def _current_report_key(self) -> str:
        try:
            tab_id = self.nb.select()
        except Exception:
            return "acik"
        for key, tab in zip(REPORTS.keys(), self.nb.tabs()):
            if tab == tab_id:
                return key
        return "acik"

    def export_csv(self, key: Optional[str] = None):
        report_key = key or self._current_report_key()
        rows = self.report_rows.get(report_key, [])
        if not rows:
            messagebox.showinfo(APP_TITLE, "Export için rapor yok.")
            return
        p = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            title="CSV Kaydet",
        )
        if not p:
            return
        headers = self.report_headers.get(report_key, [])
        keys = self.report_keys.get(report_key, [])
        try:
            with open(p, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f, delimiter=";")
                writer.writerow(headers)
                for r in rows:
                    writer.writerow([r.get(k, "") for k in keys])
            messagebox.showinfo(APP_TITLE, "CSV export tamamlandı.")
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"CSV export hatası: {e}")

    def export_excel(self, key: Optional[str] = None):
        report_key = key or self._current_report_key()
        rows = self.report_rows.get(report_key, [])
        if not rows:
            messagebox.showinfo(APP_TITLE, "Export için rapor yok.")
            return
        if not HAS_OPENPYXL:
            messagebox.showerror(APP_TITLE, "openpyxl yok. Excel export için kur: pip install openpyxl")
            return
        p = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            title="Excel Kaydet",
        )
        if not p:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter

            headers = self.report_headers.get(report_key, [])
            keys = self.report_keys.get(report_key, [])

            wb = Workbook()
            ws = wb.active
            ws.append(headers)
            for r in rows:
                ws.append([r.get(k, "") for k in keys])
            for idx, _ in enumerate(headers, start=1):
                ws.column_dimensions[get_column_letter(idx)].width = 18
            wb.save(p)
            messagebox.showinfo(APP_TITLE, "Excel export tamamlandı.")
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"Excel export hatası: {e}")

    def export_pdf(self, key: Optional[str] = None):
        report_key = key or self._current_report_key()
        rows = self.report_rows.get(report_key, [])
        if not rows:
            messagebox.showinfo(APP_TITLE, "Export için rapor yok.")
            return
        if not HAS_REPORTLAB:
            messagebox.showerror(APP_TITLE, "reportlab yok. PDF için kur: pip install reportlab")
            return
        p = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            title="PDF Kaydet",
        )
        if not p:
            return
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors

            headers = self.report_headers.get(report_key, [])
            keys = self.report_keys.get(report_key, [])

            data = [headers]
            for r in rows:
                data.append([str(r.get(k, "")) for k in keys])

            ensure_pdf_fonts()
            doc = SimpleDocTemplate(p, pagesize=A4)
            styles = getSampleStyleSheet()
            story = [Paragraph(REPORTS[report_key].title, styles["Title"]), Spacer(1, 12)]
            table = Table(data, repeatRows=1)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]
                )
            )
            story.append(table)
            doc.build(story)
            messagebox.showinfo(APP_TITLE, "PDF hazırlandı.")
        except Exception as e:
            messagebox.showerror(APP_TITLE, f"PDF oluşturulamadı: {e}")
