# -*- coding: utf-8 -*-
# mypy: ignore-errors
"""KasaPro v3 - Pencereler (Toplevel) ve import/ekstre ekranları"""

from __future__ import annotations

import os
import re
import sqlite3
from datetime import datetime, timedelta, date
from typing import Any, Optional, List, Dict, Tuple, TYPE_CHECKING

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

from ..config import APP_TITLE, HAS_OPENPYXL, HAS_REPORTLAB
from ..utils import (
    center_window,
    ensure_pdf_fonts,
    fmt_tr_date,
    safe_float,
    fmt_amount,
    norm_header,
)
from ..db.users_db import UsersDB
from .widgets import LabeledEntry, LabeledCombo
from .dialogs import simple_input, simple_choice

if TYPE_CHECKING:
    from ..app import App
class LoginWindow(tk.Toplevel):
    def __init__(self, root: tk.Tk, usersdb: UsersDB):
        super().__init__(root)
        self.usersdb = usersdb
        self.user: Optional[sqlite3.Row] = None
        self.title("Giriş")
        self.geometry("390x255")
        self.resizable(False, False)
        # root withdraw durumunda transient sorun çıkarabiliyor
        try:
            if bool(root.winfo_viewable()):
                self.transient(root)
        except Exception:
            pass
        self.protocol("WM_DELETE_WINDOW", self.do_exit)

        ttk.Label(self, text="KasaPro Giriş", font=("Calibri", 14, "bold")).pack(pady=(14, 6))
        frm = ttk.Frame(self)
        frm.pack(fill=tk.X, padx=18, pady=6)
        # Kullanıcı seçimi (liste)
        self.pick_user = LabeledCombo(frm, "Kullanıcı Seç:", self.usersdb.list_usernames(), 18)
        self.pick_user.pack(fill=tk.X, pady=6)
        try:
            self.pick_user.cmb.bind("<<ComboboxSelected>>", lambda _e: self._on_pick_user())
        except Exception:
            pass

        # Varsayılan kullanıcı: listedeki ilk değer
        try:
            vals = list(self.pick_user.cmb["values"])
            if vals:
                self.pick_user.set(vals[0])
        except Exception:
            pass


        self.e_user = LabeledEntry(frm, "Kullanıcı:", 18)
        self.e_user.pack(fill=tk.X, pady=6)
        self.e_pass = LabeledEntry(frm, "Şifre:", 18)
        self.e_pass.pack(fill=tk.X, pady=6)
        self.e_pass.ent.config(show="*")

        # Başlangıçta seçilen kullanıcıyı kullanıcı alanına yansıt
        try:
            self._on_pick_user()
        except Exception:
            pass

        btn = ttk.Frame(self)
        btn.pack(fill=tk.X, padx=18, pady=10)
        ttk.Button(btn, text="Giriş", command=self.do_login).pack(side=tk.LEFT)
        ttk.Button(btn, text="Çıkış", command=self.do_exit).pack(side=tk.RIGHT)

        ttk.Label(self, text="İlk kurulum: admin / admin", foreground="#666").pack(pady=(0, 6))
        self.bind("<Return>", lambda _e: self.do_login())
        # ✅ Daha sağlam modal gösterimi (root withdraw olsa bile)
        self.protocol("WM_DELETE_WINDOW", self.do_exit)
        center_window(self)

        # Bazı sistemlerde pencere görünür olmadan wait_visibility/grab_set kilitleyebiliyor.
        # Bu yüzden her şeyi biraz gecikmeli yapıyoruz.
        try:
            self.deiconify()
        except Exception:
            pass

        def _post_show():
            try:
                self.lift()
                self.focus_force()
            except Exception:
                pass
            try:
                self.attributes("-topmost", True)
                self.after(300, lambda: self.attributes("-topmost", False))
            except Exception:
                pass
            try:
                self.grab_set()
            except Exception:
                pass
            try:
                self.e_user.ent.focus_set()
            except Exception:
                pass

        self.after(50, _post_show)

    def _on_pick_user(self):
        u = self.pick_user.get().strip() if hasattr(self, 'pick_user') else ''
        if u:
            try:
                self.e_user.set(u)
            except Exception:
                try:
                    self.e_user.ent.delete(0, tk.END)
                    self.e_user.ent.insert(0, u)
                except Exception:
                    pass
            try:
                self.e_pass.ent.focus_set()
            except Exception:
                pass

    def do_login(self):
        u = self.e_user.get().strip()
        p = self.e_pass.get()
        user = self.usersdb.auth(u, p)
        if not user:
            messagebox.showerror(APP_TITLE, "Hatalı kullanıcı/şifre.")
            return
        self.user = user
        # Login başarılı (log, kullanıcı DB'si açıldıktan sonra App içinde yazılır)
        self.destroy()

    def do_exit(self):
        self.user = None
        self.destroy()

class SettingsWindow(tk.Toplevel):
    def __init__(self, app:"App"):
        super().__init__(app.root)
        self.app = app
        self.db = app.db
        self.title("Ayarlar")
        self.geometry("760x520")
        self._build()
        center_window(self, app.root)

    def _build(self):
        nb = ttk.Notebook(self)
        nb.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        tab1 = ttk.Frame(nb)
        nb.add(tab1, text="Listeler")
        tab2 = ttk.Frame(nb)
        nb.add(tab2, text="Kullanıcılar")

        self._list_editor(tab1, "Para Birimleri", "currencies", self.db.list_currencies(), y=0)
        self._list_editor(tab1, "Ödeme Tipleri", "payments", self.db.list_payments(), y=170)
        self._list_editor(tab1, "Kategoriler", "categories", self.db.list_categories(), y=340)

        if not self.app.is_admin:
            ttk.Label(tab2, text="Kullanıcı yönetimi sadece admin içindir.").pack(padx=10, pady=10, anchor="w")
            return

        top = ttk.Frame(tab2)
        top.pack(fill=tk.X, padx=10, pady=10)
        ttk.Button(top, text="Yenile", command=self._users_refresh).pack(side=tk.LEFT)
        ttk.Button(top, text="Yeni", command=self._user_add).pack(side=tk.LEFT, padx=6)
        ttk.Button(top, text="Şifre", command=self._user_pass).pack(side=tk.LEFT, padx=6)
        ttk.Button(top, text="Rol", command=self._user_role).pack(side=tk.LEFT, padx=6)
        ttk.Button(top, text="Sil", command=self._user_del).pack(side=tk.LEFT, padx=6)

        cols = ("id","username","role","created_at")
        self.tree = ttk.Treeview(tab2, columns=cols, show="headings", height=16)
        for c in cols:
            self.tree.heading(c, text=c.upper())
        self.tree.column("id", width=50, anchor="center")
        self.tree.column("username", width=160)
        self.tree.column("role", width=90, anchor="center")
        self.tree.column("created_at", width=160)
        self.tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0,10))
        self._users_refresh()

    def _list_editor(self, master, title, key, items, y=0):
        import json
        box = ttk.LabelFrame(master, text=title)
        box.place(x=10, y=10+y, width=720, height=150)

        txt = tk.Text(box, height=6)
        txt.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        txt.insert("1.0", "\n".join(items))

        def save():
            lines = [line.strip() for line in txt.get("1.0", tk.END).splitlines() if line.strip()]
            self.db.set_setting(key, json.dumps(lines, ensure_ascii=False))
            self.db.log("Settings", f"{key} updated ({len(lines)})")
            self.app.reload_settings()
            messagebox.showinfo(APP_TITLE, f"{title} kaydedildi.")

        ttk.Button(box, text="Kaydet", command=save).pack(anchor="e", padx=8, pady=(0,8))

    def _selected_user_id(self) -> Optional[int]:
        s = self.tree.selection()
        if not s:
            return None
        return int(self.tree.item(s[0], "values")[0])

    def _users_refresh(self):
        for i in self.tree.get_children():
            self.tree.delete(i)
        for r in self.db.users_list():
            self.tree.insert("", tk.END, values=(r["id"], r["username"], r["role"], r["created_at"]))

    def _user_add(self):
        u = simple_input(self, "Yeni Kullanıcı", "Kullanıcı adı:")
        if not u:
            return
        # Şifre doğrulama (iki kez sor)
        p1 = simple_input(self, "Yeni Kullanıcı", "Şifre:", password=True)
        if p1 is None:
            return
        p2 = simple_input(self, "Yeni Kullanıcı", "Şifre (tekrar):", password=True)
        if p2 is None:
            return
        if p1 != p2:
            messagebox.showerror(APP_TITLE, "Şifreler eşleşmiyor.")
            return
        role = simple_choice(self, "Rol", "Rol seç:", ["admin","user"], default="user")
        if not role:
            return
        try:
            self.db.user_add(u, p1, role)
            self._users_refresh()
        except Exception as e:
            messagebox.showerror(APP_TITLE, str(e))

    def _user_pass(self):
        uid = self._selected_user_id()
        if not uid:
            return
        # Şifre doğrulama (iki kez sor)
        p1 = simple_input(self, "Şifre Değiştir", "Yeni şifre:", password=True)
        if p1 is None:
            return
        p2 = simple_input(self, "Şifre Değiştir", "Yeni şifre (tekrar):", password=True)
        if p2 is None:
            return
        if p1 != p2:
            messagebox.showerror(APP_TITLE, "Şifreler eşleşmiyor.")
            return
        self.db.user_set_password(uid, p1)
        messagebox.showinfo(APP_TITLE, "Şifre güncellendi.")

    def _user_role(self):
        uid = self._selected_user_id()
        if not uid:
            return
        role = simple_choice(self, "Rol Değiştir", "Rol seç:", ["admin","user"], default="user")
        if not role:
            return
        self.db.user_set_role(uid, role)
        self._users_refresh()

    def _user_del(self):
        uid = self._selected_user_id()
        if not uid:
            return
        if messagebox.askyesno(APP_TITLE, "Kullanıcı silinsin mi?"):
            self.db.user_delete(uid)
            self._users_refresh()


# =========================
# YARDIM / DOKUMANTASYON
# =========================

def _tr_norm(s: str) -> str:
    """Arama için TR-dostu normalizasyon (I/İ davranışı dahil)."""
    s = "" if s is None else str(s)
    s = s.replace("I", "ı").replace("İ", "i")
    return s.lower()

HELP_TOPICS: List[Tuple[str, str]] = [
    ("Genel Bakış",
"""KasaPro
Kasa hareketleri, Cariler ve Cari Hareketleri tek bir veritabanında tutar.
Sol menüden ekranlar arasında geçiş yapabilirsin.

Roller:
- admin: silme/düzenleme, DB geri yükleme, kullanıcı yönetimi gibi yetkiler.
- user: kayıt ekleme/görüntüleme (bazı kritik işlemler kapalı olabilir).

Kısayollar:
- F1: Yardım penceresini açar.
"""),

    ("Navigasyon ve Genel Mantık",
"""Sol menüdeki ekranlar:
- 🏦 Kasa: gelir/gider hareketleri
- 👥 Cariler: cari kartları (unvan, telefon, notlar, açılış bakiyesi)
- 📒 Cari Hareket: borç/alacak hareketleri
- 📊 Raporlar: özetler ve dışa aktarımlar
- 🔎 Global Arama: her yerden hızlı arama
- 🧾 Log: yapılan işlemlerin kaydı

Her ekranda:
- Üst bölüm: kayıt ekleme / düzenleme formu
- Alt bölüm: kayıt listesi (filtre/arama + tablo)
"""),

    ("Tutar Girişi (TR Para Formatı)",
"""Programdaki tüm tutar girişleri TR formatını kullanır.

Örnek gösterim:
- 1111111  -> 1.111.111,00
- 100,32   -> 100,32
- 1.234,50 -> 1.234,50

Nasıl yazılır?
- Tam kısmı yaz: 1250 => 1.250,00
- Kuruş yazmak için ',' (veya '.') tuşla: 1250,75
- Yazarken otomatik nokta/virgül maskelemesi yapılır.
- Enter veya alandan çıkınca (focus out) format kesinleşir.

İpucu:
- Binlik ayıracı olarak '.' kullanılır.
- Ondalık ayıracı olarak ',' kullanılır.
"""),

    ("Kasa Ekranı",
"""Kasa Hareket Ekle:
- Tarih / Tip (Gelir-Gider) / Tutar / Para / Ödeme / Kategori / Cari (opsiyonel)
- Belge No ve Etiket alanlarıyla kayıtlarını sınıflandırabilirsin.
- Açıklama alanı bir butondur: 'Açıklama yaz…' → sekmeli pencerede yaz.

Kayıt listesi:
- Çift tık: Seçili kaydı düzenle (admin)
- Seçili Kaydı Düzenle / Sil (admin)

Kaydet sonrası:
- Yeni kayıt için form temizlenir (açıklama dahil).
"""),

    ("Cari Hareket Ekranı",
"""Cari Hareket Ekle:
- Tarih / Cari / Tip (Borç-Alacak) / Tutar / Para / Ödeme / Belge / Etiket
- Açıklama: butona basarak sekmeli pencereden girilir.

Çoklu seçim:
- 'Çoklu Seçim: Açık' iken satırlara tek tek tıklayarak çoklu seçebilirsin.
- 'Seçili Kaydı Sil' çoklu seçimde hepsini siler (admin).

Düzenleme:
- Düzenleme için tek kayıt seçili olmalıdır (admin).
- Çift tık: seçili kaydı düzenle (admin).
"""),

    ("Cariler Ekranı",
"""Cari kartı:
- Cari Adı (zorunlu), Tür, Telefon, Notlar
- Açılış Bakiyesi: TR para formatındadır.

İşlemler:
- Kaydet: yeni cari ekler / seçili cari üzerinde günceller
- Sil: admin yetkisiyle çalışır

Not:
- Bir cariyi silmek için ilişkili hareketler varsa önce hareketleri temizlemek gerekebilir.
"""),

    ("Açıklama Sekmesi (Buton)",
"""Kasa ve Cari Hareket ekranlarında Açıklama alanı butondur.

Kullanım:
- 'Açıklama yaz…' butonuna bas → sekmeli pencere açılır.
- Metni yaz → 'Uygula' veya 'Uygula & Kapat' ile forma aktarılır.
- Kaydet dedikten sonra yeni kayıt için açıklama otomatik temizlenir.
- Pencere ekranın ortasında açılır.
"""),

    ("Excel İçe Aktar / Export",
"""Excel İçe Aktar:
- openpyxl kurulu olmalı: pip install openpyxl
- Dosya seçilir → Eşleştirme Sihirbazı açılır
- Her tablo için sheet ve kolon eşleştirmesi yapılır
- 'Cari yoksa otomatik oluştur' seçeneği ile eksik cariler otomatik eklenebilir.

Excel Export:
- Veriler yeni bir Excel dosyasına aktarılır.
"""),

    ("PDF Dışa Aktarım (Türkçe Karakter)",
"""PDF'te Türkçe karakterler bozuluyorsa:
- PDF çıktısı için Unicode font gömme kullanılır (DejaVuSans/Arial vb.).
- Eğer sistemde uygun font bulunamazsa metinde kare/bozuk karakter görülebilir.

Çözüm:
- Windows'ta genelde Arial bulunduğundan otomatik düzelir.
- Gerekirse DejaVuSans.ttf dosyasını programın yanına koyabilirsin.
"""),

    ("DB Yedek / Geri Yükle",
"""DB Yedek:
- '💾 DB Yedek' butonu mevcut giriş yapılan kullanıcının veritabanını kopyalar.

DB Geri Yükle:
- Mevcut giriş yapılan kullanıcının DB'sinin üstüne yazar (geri dönüşü zordur). Önce yedek önerilir.
"""),

    ("Kullanıcılar (Çoklu Kullanıcı)",
"""- Giriş ekranında kullanıcı seçip şifreyle giriş yaparsın.
- Her kullanıcının verileri ayrı tutulur: kasa_data/ klasöründe kullanıcıya özel .db dosyası.
- Sadece admin: Sol menüde "👤 Kullanıcılar" bölümünden kullanıcı ekle/sil ve şifre sıfırla.
- "💾 DB Yedek" / "♻️ DB Geri Yükle" işlemleri mevcut giriş yapılan kullanıcının verisini etkiler.
"""),

    ("Sık Sorulanlar",
"""S: Tutar yazarken neden otomatik değişiyor?
C: Para girişleri TR formatında maskelenir
bu yanlış girişi azaltır.

S: Silme/Düzenleme butonları pasif.
C: Admin hesabıyla giriş yapmalısın.

S: Excel import görünmüyor.
C: openpyxl kurulu değilse import/export devre dışı kalır.

S: PDF'te Türkçe karakter bozuk.
C: Unicode font gömme ayarı gerekir (programda otomatik denenir).
"""),
]

class HelpWindow(tk.Toplevel):
    def __init__(self, app: "App"):
        super().__init__(app.root)
        self.app = app
        self.title("Yardım")
        self.geometry("980x640")
        self.minsize(820, 520)
        try:
            self.transient(app.root)
        except Exception:
            pass

        self._all_topics = HELP_TOPICS[:]
        self._topics_view: List[Tuple[str, str]] = self._all_topics[:]
        self._current_topic_index: int = 0
        self._last_find_index: str = "1.0"

        self._build()
        center_window(self, app.root)

    def _build(self):
        top = ttk.Frame(self)
        top.pack(fill=tk.X, padx=12, pady=10)

        ttk.Label(top, text="Yardım", font=("Calibri", 14, "bold")).pack(side=tk.LEFT)
        ttk.Label(top, text=f"  (Kullanıcı: {self.app.user['username']} / {self.app.user['role']})", foreground="#666").pack(side=tk.LEFT)

        self.var_q = tk.StringVar()
        ent = ttk.Entry(top, textvariable=self.var_q, width=42)
        ent.pack(side=tk.RIGHT, padx=(6, 0))
        ent.bind("<Return>", lambda _e: self.search())
        ttk.Button(top, text="Ara", command=self.search).pack(side=tk.RIGHT, padx=(6, 0))
        ttk.Button(top, text="Temizle", command=self.clear_search).pack(side=tk.RIGHT, padx=(6, 12))

        self.lbl_status = ttk.Label(self, text="", foreground="#666")
        self.lbl_status.pack(fill=tk.X, padx=12, pady=(0, 6))

        pw = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        pw.pack(fill=tk.BOTH, expand=True, padx=12, pady=(0, 12))

        left = ttk.Frame(pw)
        right = ttk.Frame(pw)
        pw.add(left, weight=1)
        pw.add(right, weight=3)

        # Sol: konu listesi
        ttk.Label(left, text="Konular").pack(anchor="w", padx=6, pady=(6, 2))
        self.lb = tk.Listbox(left, height=18)
        sb = ttk.Scrollbar(left, orient="vertical", command=self.lb.yview)
        self.lb.configure(yscrollcommand=sb.set)
        self.lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(6,0), pady=6)
        sb.pack(side=tk.RIGHT, fill=tk.Y, pady=6)

        self.lb.bind("<<ListboxSelect>>", self._on_select_topic)

        # Sağ: içerik + arama içinde gezinme
        nav = ttk.Frame(right)
        nav.pack(fill=tk.X, padx=6, pady=(6, 0))
        ttk.Button(nav, text="Önceki", command=lambda: self.find_next(backwards=True)).pack(side=tk.LEFT)
        ttk.Button(nav, text="Sonraki", command=lambda: self.find_next(backwards=False)).pack(side=tk.LEFT, padx=6)
        ttk.Button(nav, text="Kopyala", command=self.copy_current).pack(side=tk.RIGHT)

        self.txt = tk.Text(right, wrap="word")
        ysb = ttk.Scrollbar(right, orient="vertical", command=self.txt.yview)
        self.txt.configure(yscrollcommand=ysb.set)
        self.txt.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(6,0), pady=6)
        ysb.pack(side=tk.RIGHT, fill=tk.Y, pady=6)

        self.txt.tag_configure("h", font=("Calibri", 13, "bold"))
        self.txt.tag_configure("hl", background="#ffe08a")

        self._reload_list()
        self._show_topic(0)

        # kısayol
        try:
            self.bind("<Escape>", lambda _e: self.destroy())
        except Exception:
            pass

    def _reload_list(self):
        self.lb.delete(0, tk.END)
        for title, _body in self._topics_view:
            self.lb.insert(tk.END, title)
        if self._topics_view:
            self.lb.selection_clear(0, tk.END)
            self.lb.selection_set(0)
            self.lb.activate(0)

    def _on_select_topic(self, _e=None):
        sel = self.lb.curselection()
        if not sel:
            return
        self._show_topic(int(sel[0]))

    def _show_topic(self, idx: int):
        if not self._topics_view:
            return
        idx = max(0, min(idx, len(self._topics_view)-1))
        self._current_topic_index = idx
        title, body = self._topics_view[idx]

        self.txt.configure(state="normal")
        self.txt.delete("1.0", tk.END)
        self.txt.insert("1.0", title + "\n", ("h",))
        self.txt.insert("end", "\n" + body.strip() + "\n")
        self.txt.configure(state="disabled")

        self._last_find_index = "1.0"
        self._apply_highlight()

        # status
        q = self.var_q.get().strip()
        if q:
            self.lbl_status.config(text=f"Arama: '{q}'  |  Konu: {title}")
        else:
            self.lbl_status.config(text=f"Konu: {title}")

    def _apply_highlight(self):
        q = self.var_q.get().strip()
        self.txt.configure(state="normal")
        self.txt.tag_remove("hl", "1.0", tk.END)
        if q:
            start = "1.0"
            while True:
                pos = self.txt.search(q, start, stopindex=tk.END, nocase=True)
                if not pos:
                    break
                end = f"{pos}+{len(q)}c"
                self.txt.tag_add("hl", pos, end)
                start = end
        self.txt.configure(state="disabled")

    def search(self):
        q = self.var_q.get().strip()
        if not q:
            self.clear_search()
            return

        nq = _tr_norm(q)
        filtered = []
        for title, body in self._all_topics:
            if nq in _tr_norm(title) or nq in _tr_norm(body):
                filtered.append((title, body))

        self._topics_view = filtered if filtered else []
        self._reload_list()

        if not self._topics_view:
            self.lbl_status.config(text=f"'{q}' için sonuç bulunamadı.")
            self.txt.configure(state="normal")
            self.txt.delete("1.0", tk.END)
            self.txt.insert("1.0", "Sonuç bulunamadı. Aramayı değiştir veya 'Temizle'ye bas.")
            self.txt.configure(state="disabled")
            return

        self.lbl_status.config(text=f"'{q}' için {len(self._topics_view)} konu bulundu.")
        self._show_topic(0)

    def clear_search(self):
        self.var_q.set("")
        self._topics_view = self._all_topics[:]
        self._reload_list()
        self.lbl_status.config(text="")
        self._show_topic(0)

    def find_next(self, backwards: bool = False):
        q = self.var_q.get().strip()
        if not q:
            return

        self.txt.configure(state="normal")
        try:
            if backwards:
                pos = self.txt.search(q, self._last_find_index, stopindex="1.0", nocase=True, backwards=True)
            else:
                pos = self.txt.search(q, self._last_find_index, stopindex=tk.END, nocase=True)
            if not pos:
                # sar
                pos = self.txt.search(q, tk.END if backwards else "1.0", stopindex="1.0" if backwards else tk.END, nocase=True, backwards=backwards)
            if pos:
                end = f"{pos}+{len(q)}c"
                self.txt.tag_remove("sel", "1.0", tk.END)
                self.txt.tag_add("sel", pos, end)
                self.txt.mark_set(tk.INSERT, end)
                self.txt.see(pos)
                self._last_find_index = pos
        finally:
            self.txt.configure(state="disabled")

    def copy_current(self):
        try:
            sel = self.lb.curselection()
            idx = int(sel[0]) if sel else 0
            title, body = self._topics_view[idx]
            txt = f"{title}\n\n{body}"
            self.clipboard_clear()
            self.clipboard_append(txt)
            self.update_idletasks()
            self.lbl_status.config(text="Kopyalandı.")
        except Exception:
            pass

# =========================
# IMPORT WIZARD (Excel mapping)
# =========================

def _auto_header_row(ws, search_rows: int = 30) -> int:
    best_row = 1
    best_score = -1
    max_col = min(ws.max_column or 1, 60)
    for r in range(1, min(ws.max_row or 1, search_rows) + 1):
        texts = 0
        for c in range(1, max_col + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip():
                texts += 1
        if texts > best_score:
            best_score = texts
            best_row = r
    return best_row

def _headers(ws, header_row: int) -> List[str]:
    max_col = min(ws.max_column or 1, 60)
    out = []
    for c in range(1, max_col + 1):
        v = ws.cell(header_row, c).value
        out.append("" if v is None else str(v).strip())
    return out

def _suggest_col(headers: List[str], keys: List[str]) -> int:
    norm = [norm_header(h) for h in headers]
    for k in keys:
        kk = norm_header(k)
        for i, h in enumerate(norm):
            if h == kk:
                return i
    for k in keys:
        kk = norm_header(k)
        for i, h in enumerate(norm):
            if kk and kk in h:
                return i
    return -1

class ImportWizard(tk.Toplevel):
    def __init__(self, app:"App", xlsx_path: str):
        super().__init__(app.root)
        self.app = app
        self.db = app.db
        self.xlsx_path = xlsx_path
        self.title("Excel İçe Aktar - Eşleştirme Sihirbazı")
        self.geometry("1080x720")
        self.resizable(True, True)
        self.grab_set()

        try:
            import openpyxl  # type: ignore
        except Exception:
            messagebox.showerror(APP_TITLE, "openpyxl kurulu değil. Kur: pip install openpyxl")
            self.destroy()
            return

        self.openpyxl = openpyxl
        self.wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        self.sheetnames = self.wb.sheetnames[:]
        self.result_counts: Optional[Dict[str, int]] = None
        self.mappings: Dict[str, Dict[str, Any]] = {}

        self._build()
        center_window(self, app.root)

    def _build(self):
        ttk.Label(self, text=f"Dosya: {os.path.basename(self.xlsx_path)}", font=("Calibri", 12, "bold")).pack(anchor="w", padx=12, pady=(10, 4))
        ttk.Label(self, text="Sayfa seç → Kolon eşleştir → İçe Aktar. (Başlıklar farklı olabilir, sorun değil.)", foreground="#666").pack(anchor="w", padx=12, pady=(0, 10))

        self.nb = ttk.Notebook(self)
        self.nb.pack(fill=tk.BOTH, expand=True, padx=12, pady=8)

        self._add_tab("Cariler", fields=[
            ("ad", "Cari Adı (Zorunlu)"),
            ("tur", "Tür"),
            ("telefon", "Telefon"),
            ("notlar", "Notlar"),
            ("acilis", "Açılış Bakiyesi"),
        ], suggest={
            "ad": ["cari", "cari adı", "ad", "unvan", "isim"],
            "tur": ["tur", "tip"],
            "telefon": ["telefon", "tel", "gsm"],
            "notlar": ["not", "aciklama", "açıklama"],
            "acilis": ["acilis", "açılış", "acilis bakiyesi", "bakiye"],
        })

        self._add_tab("CariHareket", fields=[
            ("tarih", "Tarih"),
            ("cari", "Cari (Zorunlu)"),
            ("tip", "Tip (Borç/Alacak)"),
            ("tutar", "Tutar (Zorunlu)"),
            ("para", "Para"),
            ("odeme", "Ödeme"),
            ("belge", "Belge"),
            ("etiket", "Etiket"),
            ("aciklama", "Açıklama"),
        ], suggest={
            "tarih": ["tarih", "date"],
            "cari": ["cari", "cari adı", "unvan", "isim"],
            "tip": ["tip", "borç", "borc", "alacak"],
            "tutar": ["tutar", "miktar", "amount", "tahsilat", "odeme", "ödeme", "gelir", "gider", "gelen"],
            "para": ["para", "pb", "doviz", "döviz"],
            "odeme": ["odeme", "ödeme", "odeme tipi", "ödeme tipi"],
            "belge": ["belge", "fis", "fiş", "dekont"],
            "etiket": ["etiket", "tag"],
            "aciklama": ["aciklama", "açıklama", "not"],
        })

        self._add_tab("KasaHareket", fields=[
            ("tarih", "Tarih"),
            ("tip", "Tip (Gelir/Gider)"),
            ("tutar", "Tutar (Zorunlu)"),
            ("para", "Para"),
            ("odeme", "Ödeme"),
            ("kategori", "Kategori/Grup"),
            ("cari", "Cari (opsiyonel)"),
            ("belge", "Belge"),
            ("etiket", "Etiket"),
            ("aciklama", "Açıklama"),
        ], suggest={
            "tarih": ["tarih", "date"],
            "tip": ["tip", "gelir", "gider"],
            "tutar": ["tutar", "miktar", "amount", "tahsilat", "odeme", "ödeme", "gelir", "gider", "gelen"],
            "para": ["para", "pb", "doviz", "döviz"],
            "odeme": ["odeme", "ödeme", "odeme tipi", "ödeme tipi"],
            "kategori": ["kategori", "grup", "group"],
            "cari": ["cari", "unvan", "isim"],
            "belge": ["belge", "fis", "fiş", "dekont"],
            "etiket": ["etiket", "tag"],
            "aciklama": ["aciklama", "açıklama", "not"],
        })

        bottom = ttk.Frame(self)
        bottom.pack(fill=tk.X, padx=12, pady=10)

        self.var_create_missing_cari = tk.BooleanVar(value=True)
        ttk.Checkbutton(bottom, text="İçe aktarma sırasında cari yoksa otomatik oluştur", variable=self.var_create_missing_cari).pack(side=tk.LEFT)

        ttk.Button(bottom, text="İptal", command=self._cancel).pack(side=tk.RIGHT, padx=6)
        ttk.Button(bottom, text="İçe Aktar", command=self._do_import).pack(side=tk.RIGHT, padx=6)

    def _guess_sheet(self, tab_name: str) -> str:
        t = tab_name.lower()
        for s in self.sheetnames:
            sl = s.lower()
            if t == "cariler" and ("cari" in sl and "hareket" not in sl):
                return s
            if t == "carihareket" and ("cari" in sl and "hareket" in sl):
                return s
            if t == "kasahareket" and ("kasa" in sl):
                return s
        return self.sheetnames[0] if self.sheetnames else ""

    def _add_tab(self, name: str, fields: List[Tuple[str, str]], suggest: Dict[str, List[str]]):
        tab = ttk.Frame(self.nb)
        self.nb.add(tab, text=name)

        top = ttk.Frame(tab)
        top.pack(fill=tk.X, padx=10, pady=10)
        ttk.Label(top, text="Sayfa:", width=8).pack(side=tk.LEFT)

        guess = self._guess_sheet(name)
        sheet_var = tk.StringVar(value=(guess if guess else "(Atla)"))
        sheet_cmb = ttk.Combobox(top, textvariable=sheet_var, values=["(Atla)"] + self.sheetnames, state="readonly", width=28)
        sheet_cmb.pack(side=tk.LEFT, padx=(0,10))
        ttk.Button(top, text="Önizleme/Yenile", command=lambda: self._refresh_preview(name)).pack(side=tk.LEFT)

        mapbox = ttk.LabelFrame(tab, text="Kolon Eşleştirme")
        mapbox.pack(fill=tk.X, padx=10, pady=(0,10))

        mapping_vars: Dict[str, tk.StringVar] = {}
        for row, (key, label) in enumerate(fields):
            var = tk.StringVar(value="(Yok)")
            mapping_vars[key] = var
            ttk.Label(mapbox, text=label, width=22).grid(row=row, column=0, sticky="w", padx=10, pady=4)
            cmb = ttk.Combobox(mapbox, textvariable=var, values=["(Yok)"], state="readonly", width=40)
            cmb.grid(row=row, column=1, sticky="w", padx=6, pady=4)

        prev = ttk.LabelFrame(tab, text="Önizleme (ilk 20 veri satırı)")
        prev.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0,10))

        tree = ttk.Treeview(prev, show="headings", height=14)
        tree.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)
        scr = ttk.Scrollbar(prev, orient="vertical", command=tree.yview)
        scr.pack(fill=tk.Y, side=tk.RIGHT)
        tree.configure(yscrollcommand=scr.set)

        self.mappings[name] = {
            "tab": tab,
            "sheet_var": sheet_var,
            "sheet_cmb": sheet_cmb,
            "mapping_vars": mapping_vars,
            "tree": tree,
            "fields": fields,
            "suggest": suggest,
            "header_row": 1,
            "headers": [],
        }

        sheet_cmb.bind("<<ComboboxSelected>>", lambda _e: self._refresh_preview(name))
        self._refresh_preview(name)

    def _refresh_preview(self, tab_name: str):
        info = self.mappings[tab_name]
        sheet_name = info["sheet_var"].get()
        tree: ttk.Treeview = info["tree"]

        if not sheet_name or sheet_name == "(Atla)":
            for i in tree.get_children():
                tree.delete(i)
            tree["columns"] = ()
            return

        ws = self.wb[sheet_name]
        header_row = _auto_header_row(ws, search_rows=30)
        headers = _headers(ws, header_row)
        info["header_row"] = header_row
        info["headers"] = headers

        col_options = ["(Yok)"]
        for idx, h in enumerate(headers, start=1):
            col_letter = self.openpyxl.utils.get_column_letter(idx)
            col_options.append(f"{col_letter}: {h}" if h else f"{col_letter}: (boş)")

        tab = info["tab"]
        if len(tab.winfo_children()) >= 2:
            mapbox = tab.winfo_children()[1]
            for w in mapbox.winfo_children():
                if isinstance(w, ttk.Combobox):
                    w["values"] = col_options

        suggest: Dict[str, List[str]] = info["suggest"]
        for key, var in info["mapping_vars"].items():
            if var.get() not in ("", "(Yok)"):
                continue
            idx0 = _suggest_col(headers, suggest.get(key, []))
            if idx0 >= 0:
                col_letter = self.openpyxl.utils.get_column_letter(idx0 + 1)
                h = headers[idx0] if idx0 < len(headers) else ""
                var.set(f"{col_letter}: {h}" if h else f"{col_letter}: (boş)")

        for i in tree.get_children():
            tree.delete(i)

        max_col = min(ws.max_column or 1, max(1, len(headers)))
        cols = [f"C{c}" for c in range(1, max_col + 1)]
        tree["columns"] = cols
        for i, c in enumerate(cols, start=1):
            h = headers[i-1] if i-1 < len(headers) else ""
            col_letter = self.openpyxl.utils.get_column_letter(i)
            tree.heading(c, text=f"{col_letter} {h}".strip())
            tree.column(c, width=140, anchor="w")

        start = header_row + 1
        end = min(ws.max_row or start, start + 19)
        for r in range(start, end + 1):
            vals = []
            for c in range(1, max_col + 1):
                v = ws.cell(r, c).value
                if isinstance(v, (datetime, date)):
                    vals.append(v.strftime("%Y-%m-%d"))
                else:
                    s = "" if v is None else str(v)
                    vals.append(s[:60] + ("…" if len(s) > 60 else ""))
            tree.insert("", tk.END, values=vals)

    def _parse_col_choice(self, choice: str) -> Optional[int]:
        if not choice or choice == "(Yok)":
            return None
        m = re.match(r"^([A-Z]{1,3}):", choice.strip())
        if not m:
            return None
        return self.openpyxl.utils.column_index_from_string(m.group(1))

    def _cancel(self):
        self.result_counts = None
        self.destroy()

    def _do_import(self):
        plan: Dict[str, Dict[str, Any]] = {}
        for tab_name, info in self.mappings.items():
            cols = {}
            for key, _label in info["fields"]:
                cols[key] = self._parse_col_choice(info["mapping_vars"][key].get())
            plan[tab_name] = {"sheet": info["sheet_var"].get(), "header_row": info["header_row"], "cols": cols}

        if plan["CariHareket"]["sheet"] and plan["CariHareket"]["sheet"] != "(Atla)":
            if plan["CariHareket"]["cols"].get("cari") is None or plan["CariHareket"]["cols"].get("tutar") is None:
                messagebox.showerror(APP_TITLE, "CariHareket için en az 'Cari' ve 'Tutar' kolonlarını seçmelisin.")
                return

        if plan["KasaHareket"]["sheet"] and plan["KasaHareket"]["sheet"] != "(Atla)":
            if plan["KasaHareket"]["cols"].get("tutar") is None:
                messagebox.showerror(APP_TITLE, "KasaHareket için en az 'Tutar' kolonunu seçmelisin.")
                return

        try:
            counts = self._run_import(plan, create_missing_cari=self.var_create_missing_cari.get())
            self.result_counts = counts
            messagebox.showinfo(APP_TITLE, f"İçe aktarıldı:\nCariler: {counts['cariler']}\nCari Hareket: {counts['cari_hareket']}\nKasa: {counts['kasa']}")
            self.destroy()
        except Exception as e:
            messagebox.showerror(APP_TITLE, str(e))

    def _run_import(self, plan: Dict[str, Dict[str, Any]], create_missing_cari: bool) -> Dict[str, int]:
        counts = {"cariler": 0, "cari_hareket": 0, "kasa": 0}

        def cell(ws, r, col):
            if col is None:
                return None
            return ws.cell(r, col).value

        car = plan.get("Cariler")
        if car and car.get("sheet") and car.get("sheet") != "(Atla)":
            ws = self.wb[car["sheet"]]
            hr = int(car["header_row"])
            c = car["cols"]
            if c.get("ad"):
                for r in range(hr + 1, (ws.max_row or hr) + 1):
                    ad = cell(ws, r, c.get("ad"))
                    if not ad or not str(ad).strip():
                        continue
                    tur = str(cell(ws, r, c.get("tur")) or "")
                    tel = str(cell(ws, r, c.get("telefon")) or "")
                    notlar = str(cell(ws, r, c.get("notlar")) or "")
                    acilis = safe_float(cell(ws, r, c.get("acilis")))
                    self.db.cari_upsert(str(ad), tur, tel, notlar, acilis)
                    counts["cariler"] += 1

        ch = plan.get("CariHareket")
        if ch and ch.get("sheet") and ch.get("sheet") != "(Atla)":
            ws = self.wb[ch["sheet"]]
            hr = int(ch["header_row"])
            c = ch["cols"]
            for r in range(hr + 1, (ws.max_row or hr) + 1):
                cari = cell(ws, r, c.get("cari"))
                tutar = cell(ws, r, c.get("tutar"))
                if not cari or not str(cari).strip():
                    continue
                if safe_float(tutar) == 0:
                    continue

                cari_name = str(cari).strip()
                c_row = self.db.cari_get_by_name(cari_name)
                if not c_row:
                    if not create_missing_cari:
                        continue
                    cari_id = self.db.cari_upsert(cari_name)
                else:
                    cari_id = int(c_row["id"])

                tarih = cell(ws, r, c.get("tarih"))
                tip_raw = str(cell(ws, r, c.get("tip")) or "Borç")
                tipn = "Alacak" if "alacak" in norm_header(tip_raw) else "Borç"
                para = str(cell(ws, r, c.get("para")) or "TL")
                odeme = str(cell(ws, r, c.get("odeme")) or "")
                belge = str(cell(ws, r, c.get("belge")) or "")
                if not belge.strip():
                    belge = self.db.next_belge_no("C")
                etiket = str(cell(ws, r, c.get("etiket")) or "")
                aciklama = str(cell(ws, r, c.get("aciklama")) or "")

                self.db.cari_hareket_add(tarih, cari_id, tipn, safe_float(tutar), para, aciklama, odeme, belge, etiket)
                counts["cari_hareket"] += 1

        kh = plan.get("KasaHareket")
        if kh and kh.get("sheet") and kh.get("sheet") != "(Atla)":
            ws = self.wb[kh["sheet"]]
            hr = int(kh["header_row"])
            c = kh["cols"]
            for r in range(hr + 1, (ws.max_row or hr) + 1):
                tutar = cell(ws, r, c.get("tutar"))
                if safe_float(tutar) == 0:
                    continue

                tarih = cell(ws, r, c.get("tarih"))
                tip_raw = str(cell(ws, r, c.get("tip")) or "Gider")
                tipn = "Gelir" if "gelir" in norm_header(tip_raw) else "Gider"
                para = str(cell(ws, r, c.get("para")) or "TL")
                odeme = str(cell(ws, r, c.get("odeme")) or "")
                kategori = str(cell(ws, r, c.get("kategori")) or "")
                belge = str(cell(ws, r, c.get("belge")) or "")
                if not belge.strip():
                    belge = self.db.next_belge_no("K")
                etiket = str(cell(ws, r, c.get("etiket")) or "")
                aciklama = str(cell(ws, r, c.get("aciklama")) or "")

                cari_id = None
                cari = cell(ws, r, c.get("cari"))
                if cari and str(cari).strip():
                    cari_name = str(cari).strip()
                    c_row = self.db.cari_get_by_name(cari_name)
                    if not c_row:
                        if create_missing_cari:
                            cari_id = self.db.cari_upsert(cari_name)
                    else:
                        cari_id = int(c_row["id"])

                self.db.kasa_add(tarih, tipn, safe_float(tutar), para, odeme, kategori, cari_id, aciklama, belge, etiket)
                counts["kasa"] += 1

        self.db.log("Excel Import Wizard", f"{os.path.basename(self.xlsx_path)} | {counts}")
        return counts


# =========================
# CARİ EKSTRE WINDOW
# =========================

class CariEkstreWindow(tk.Toplevel):
    def __init__(self, app: "App", cari_id: int):
        super().__init__(app.root)
        self.app = app
        self.db = app.db
        self.cari_id = int(cari_id)

        cari = self.db.cari_get(self.cari_id)
        self.title(f"Cari Ekstre: {cari['ad'] if cari else self.cari_id}")
        self.geometry("1180x720")
        self.resizable(True, True)
        self.transient(app.root)
        self.grab_set()

        self._build()
        self.refresh()
        center_window(self, app.root)

    def _build(self):
        top = ttk.LabelFrame(self, text="Filtre")
        top.pack(fill=tk.X, padx=10, pady=10)

        r1 = ttk.Frame(top)
        r1.pack(fill=tk.X, pady=6)
        self.f_from = LabeledEntry(r1, "Başlangıç:", 12)
        self.f_from.pack(side=tk.LEFT, padx=6)
        self.f_to = LabeledEntry(r1, "Bitiş:", 12)
        self.f_to.pack(side=tk.LEFT, padx=6)
        self.f_q = LabeledEntry(r1, "Ara:", 22)
        self.f_q.pack(side=tk.LEFT, padx=6)

        ttk.Button(r1, text="Son 30 gün", command=self.last30).pack(side=tk.LEFT, padx=6)
        ttk.Button(r1, text="Yenile", command=self.refresh).pack(side=tk.LEFT, padx=6)

        ttk.Separator(top, orient="horizontal").pack(fill=tk.X, padx=6, pady=8)
        r2 = ttk.Frame(top)
        r2.pack(fill=tk.X, pady=6)
        self.btn_pdf = ttk.Button(r2, text="📄 PDF Ekstre", command=self.export_pdf)
        self.btn_pdf.pack(side=tk.LEFT, padx=6)
        if not HAS_REPORTLAB:
            self.btn_pdf.config(state="disabled")

        self.btn_xlsx = ttk.Button(r2, text="📊 Excel Ekstre", command=self.export_excel)
        self.btn_xlsx.pack(side=tk.LEFT, padx=6)
        if not HAS_OPENPYXL:
            self.btn_xlsx.config(state="disabled")

        ttk.Button(r2, text="Kapat", command=self.destroy).pack(side=tk.RIGHT, padx=6)

        self.lbl_sum = ttk.Label(top, text="")
        self.lbl_sum.pack(anchor="w", padx=10, pady=(0,8))

        mid = ttk.LabelFrame(self, text="Hareketler")
        mid.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0,10))

        cols = ("tarih","tip","borc","alacak","para","odeme","belge","etiket","aciklama","bakiye")
        self.tree = ttk.Treeview(mid, columns=cols, show="headings", height=18)
        for c in cols:
            self.tree.heading(c, text=c.upper())

        self.tree.column("tarih", width=90)
        self.tree.column("tip", width=70)
        self.tree.column("borc", width=90, anchor="e")
        self.tree.column("alacak", width=90, anchor="e")
        self.tree.column("para", width=55, anchor="center")
        self.tree.column("odeme", width=110)
        self.tree.column("belge", width=90)
        self.tree.column("etiket", width=90)
        self.tree.column("aciklama", width=360)
        self.tree.column("bakiye", width=110, anchor="e")

        self.tree.pack(fill=tk.BOTH, expand=True, side=tk.LEFT, padx=(6,0), pady=6)

        scr = ttk.Scrollbar(mid, orient="vertical", command=self.tree.yview)
        scr.pack(fill=tk.Y, side=tk.RIGHT, padx=(0,6), pady=6)
        self.tree.configure(yscrollcommand=scr.set)

    def last30(self):
        d_to = date.today()
        d_from = d_to - timedelta(days=30)
        self.f_from.set(d_from.strftime("%d.%m.%Y"))
        self.f_to.set(d_to.strftime("%d.%m.%Y"))
        self.refresh()

    def refresh(self):
        for i in self.tree.get_children():
            self.tree.delete(i)

        data = self.db.cari_ekstre(
            self.cari_id,
            date_from=self.f_from.get(),
            date_to=self.f_to.get(),
            q=self.f_q.get()
        )

        self.lbl_sum.config(
            text=(
                f"Açılış: {fmt_amount(data['opening'])} | "
                f"Borç: {fmt_amount(data['total_borc'])} | "
                f"Alacak: {fmt_amount(data['total_alacak'])} | "
                f"Net: {fmt_amount(data['net_degisim'])} | "
                f"Kapanış: {fmt_amount(data['closing'])}"
            )
        )

        for r in data["rows"]:
            self.tree.insert("", tk.END, values=(
                fmt_tr_date(r["tarih"]),
                r["tip"],
                f"{fmt_amount(r['borc'])}" if r["borc"] else "",
                f"{fmt_amount(r['alacak'])}" if r["alacak"] else "",
                r["para"],
                r["odeme"],
                r["belge"],
                r["etiket"],
                r["aciklama"],
                f"{fmt_amount(r['bakiye'])}",
            ))

    def export_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except Exception:
            messagebox.showerror(APP_TITLE, "openpyxl kurulu değil. Kur: pip install openpyxl")
            return

        data = self.db.cari_ekstre(self.cari_id, self.f_from.get(), self.f_to.get(), self.f_q.get())
        p = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            title="Cari Ekstre Kaydet"
        )
        if not p:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Ekstre"

        ws.append(["Cari", data["cari_ad"]])
        ws.append(["Tarih Aralığı", f"{data['df'] or '-'} → {data['dt'] or '-'}"])
        ws.append(["Arama", data["q"] or "-"])
        ws.append([])
        ws.append(["Açılış", data["opening"]])
        ws.append(["Borç", data["total_borc"]])
        ws.append(["Alacak", data["total_alacak"]])
        ws.append(["Net", data["net_degisim"]])
        ws.append(["Kapanış", data["closing"]])
        ws.append([])

        headers = ["Tarih","Tip","Borç","Alacak","Para","Ödeme","Belge","Etiket","Açıklama","Bakiye"]
        ws.append(headers)

        for r in data["rows"]:
            ws.append([
                r["tarih"],
                r["tip"],
                r["borc"],
                r["alacak"],
                r["para"],
                r["odeme"],
                r["belge"],
                r["etiket"],
                r["aciklama"],
                r["bakiye"],
            ])

        for i, h in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(45, max(12, len(str(h))+2))

        ws.freeze_panes = "A13"
        wb.save(p)
        messagebox.showinfo(APP_TITLE, f"Excel ekstre kaydedildi:\n{p}")

    def export_pdf(self):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
        except Exception:
            messagebox.showerror(APP_TITLE, "PDF için reportlab yok. Kur: pip install reportlab")
            return

        data = self.db.cari_ekstre(self.cari_id, self.f_from.get(), self.f_to.get(), self.f_q.get())
        p = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            title="Cari Ekstre PDF Kaydet"
        )
        if not p:
            return

        doc = SimpleDocTemplate(p, pagesize=A4, leftMargin=28, rightMargin=28, topMargin=28, bottomMargin=28)
        font_reg, font_bold = ensure_pdf_fonts()

        styles = getSampleStyleSheet()

        # Türkçe karakterler için PDF fontlarını uygula

        for _k in ("Normal", "Title", "Heading1", "Heading2"):

            if _k in styles:

                styles[_k].fontName = (font_bold if _k != "Normal" else font_reg)

        story = []

        story.append(Paragraph(f"<b>Cari Ekstre</b> - {data['cari_ad']}", styles["Title"]))
        story.append(Spacer(1, 8))
        story.append(Paragraph(f"Tarih Aralığı: {data['df'] or '-'} → {data['dt'] or '-'}", styles["Normal"]))
        story.append(Paragraph(f"Arama: {data['q'] or '-'}", styles["Normal"]))
        story.append(Spacer(1, 10))
        story.append(Paragraph(
            f"Açılış: {fmt_amount(data['opening'])} | Borç: {fmt_amount(data['total_borc'])} | "
            f"Alacak: {fmt_amount(data['total_alacak'])} | Net: {fmt_amount(data['net_degisim'])} | "
            f"Kapanış: {fmt_amount(data['closing'])}",
            styles["Normal"]
        ))
        story.append(Spacer(1, 12))

        table_data = [["Tarih","Tip","Borç","Alacak","Para","Ödeme","Belge","Etiket","Açıklama","Bakiye"]]
        for r in data["rows"]:
            table_data.append([
                r["tarih"],
                r["tip"],
                f"{fmt_amount(r['borc'])}" if r["borc"] else "",
                f"{fmt_amount(r['alacak'])}" if r["alacak"] else "",
                r["para"],
                (r["odeme"] or "")[:20],
                (r["belge"] or "")[:16],
                (r["etiket"] or "")[:16],
                (r["aciklama"] or "")[:40],
                f"{fmt_amount(r['bakiye'])}",
            ])

        tbl = Table(table_data, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("FONTNAME", (0,0), (-1,0), font_bold),
            ("FONTNAME", (0,1), (-1,-1), font_reg),
            ("FONTSIZE", (0,0), (-1,0), 9),
            ("FONTSIZE", (0,1), (-1,-1), 8),
            ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
            ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
            ("ALIGN", (2,1), (3,-1), "RIGHT"),
            ("ALIGN", (-1,1), (-1,-1), "RIGHT"),
        ]))

        story.append(tbl)
        doc.build(story)

        messagebox.showinfo(APP_TITLE, f"PDF ekstre kaydedildi:\n{p}")


# =========================
# FRAMES
# =========================
