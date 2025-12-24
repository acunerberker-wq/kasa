# -*- coding: utf-8 -*-
"""Dışa aktarımlar (Excel/PDF).

UI tarafındaki openpyxl/reportlab bağımlılıklarını tek noktada toplamak için.
"""

from __future__ import annotations

from typing import Any, Dict, Iterable, List

from ..utils import fmt_amount, ensure_pdf_fonts


class ExportService:
    def export_table_csv(self, headers: List[str], rows: Iterable[Iterable[Any]], filepath: str) -> None:
        import csv

        with open(filepath, "w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(headers)
            for row in rows:
                writer.writerow(list(row))

    def export_table_excel(self, headers: List[str], rows: Iterable[Iterable[Any]], filepath: str) -> None:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append(list(row))
        for i, h in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(42, max(12, len(str(h)) + 2))
        ws.freeze_panes = "A2"
        wb.save(filepath)

    def export_cari_ekstre_excel(self, data: Dict[str, Any], filepath: str):
        """Cari ekstreyi Excel'e yazar.

        data: DB.cari_ekstre çıktı sözlüğü
        """
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Ekstre"

        ws.append(["Cari", data["cari_ad"]])
        ws.append(["Tarih Aralığı", f"{data['df'] or '-'} → {data['dt'] or '-'}"])
        ws.append(["Arama", data["q"] or "-"])
        ws.append([])
        ws.append(["Açılış", data["opening"]])
        ws.append(["Borç", data["total_borc"]])
        ws.append(["Alacak", data["total_alacak"]])
        ws.append(["Net", data["net_degisim"]])
        ws.append(["Kapanış", data["closing"]])
        ws.append([])

        headers = ["Tarih","Tip","Borç","Alacak","Para","Ödeme","Belge","Etiket","Açıklama","Bakiye"]
        ws.append(headers)

        for r in data["rows"]:
            ws.append([
                r["tarih"],
                r["tip"],
                r["borc"],
                r["alacak"],
                r["para"],
                r["odeme"],
                r["belge"],
                r["etiket"],
                r["aciklama"],
                r["bakiye"],
            ])

        for i, h in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(45, max(12, len(str(h))+2))

        ws.freeze_panes = "A13"
        wb.save(filepath)

    def export_cari_ekstre_pdf(self, data: Dict[str, Any], filepath: str):
        """Cari ekstreyi PDF'e yazar."""
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors

        doc = SimpleDocTemplate(filepath, pagesize=A4, leftMargin=28, rightMargin=28, topMargin=28, bottomMargin=28)
        font_reg, font_bold = ensure_pdf_fonts()

        styles = getSampleStyleSheet()
        for k in ("Normal", "Title", "Heading1", "Heading2"):
            if k in styles:
                styles[k].fontName = (font_bold if k != "Normal" else font_reg)

        story = []
        story.append(Paragraph(f"<b>Cari Ekstre</b> - {data['cari_ad']}", styles["Title"]))
        story.append(Spacer(1, 8))
        story.append(Paragraph(f"Tarih Aralığı: {data['df'] or '-'} → {data['dt'] or '-'}", styles["Normal"]))
        story.append(Paragraph(f"Arama: {data['q'] or '-'}", styles["Normal"]))
        story.append(Spacer(1, 10))
        story.append(Paragraph(
            f"Açılış: {fmt_amount(data['opening'])} | Borç: {fmt_amount(data['total_borc'])} | "
            f"Alacak: {fmt_amount(data['total_alacak'])} | Net: {fmt_amount(data['net_degisim'])} | "
            f"Kapanış: {fmt_amount(data['closing'])}",
            styles["Normal"],
        ))
        story.append(Spacer(1, 12))

        table_data = [["Tarih","Tip","Borç","Alacak","Para","Ödeme","Belge","Etiket","Açıklama","Bakiye"]]
        for r in data["rows"]:
            table_data.append([
                r["tarih"],
                r["tip"],
                f"{fmt_amount(r['borc'])}" if r["borc"] else "",
                f"{fmt_amount(r['alacak'])}" if r["alacak"] else "",
                r["para"],
                (r["odeme"] or "")[:20],
                (r["belge"] or "")[:16],
                (r["etiket"] or "")[:16],
                (r["aciklama"] or "")[:40],
                f"{fmt_amount(r['bakiye'])}",
            ])

        tbl = Table(table_data, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("FONTNAME", (0,0), (-1,0), font_bold),
            ("FONTNAME", (0,1), (-1,-1), font_reg),
            ("FONTSIZE", (0,0), (-1,0), 9),
            ("FONTSIZE", (0,1), (-1,-1), 8),
            ("GRID", (0,0), (-1,-1), 0.25, colors.grey),
            ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
            ("ALIGN", (2,1), (3,-1), "RIGHT"),
            ("ALIGN", (-1,1), (-1,-1), "RIGHT"),
        ]))

        story.append(tbl)
        doc.build(story)

    def export_purchase_report_excel(self, data: Dict[str, Any], filepath: str):
        """Satın alma raporunu Excel'e yazar."""
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Özet"

        filters = data.get("filters") or {}
        ws_summary.append(["Başlangıç", filters.get("date_from") or "-"])
        ws_summary.append(["Bitiş", filters.get("date_to") or "-"])
        ws_summary.append(["Tedarikçi", filters.get("supplier_id") or "-"])
        ws_summary.append(["Ürün", filters.get("product_name") or "-"])
        ws_summary.append(["Kategori", filters.get("category") or "-"])
        ws_summary.append(["Ödeme Tipi", filters.get("payment_type") or "-"])
        ws_summary.append([])

        ws_summary.append(["Periyot", "Alış", "İade", "Gider", "Ödeme", "Net"])
        for r in data.get("summary") or []:
            ws_summary.append([
                r.get("key"),
                r.get("alis"),
                r.get("iade"),
                r.get("gider"),
                r.get("odeme"),
                r.get("net"),
            ])

        for i in range(1, 7):
            ws_summary.column_dimensions[get_column_letter(i)].width = 18

        ws_mov = wb.create_sheet("Hareketler")
        headers = [
            "tarih",
            "hareket",
            "kaynak",
            "belge",
            "tedarikci",
            "urun",
            "kategori",
            "depo",
            "odeme",
            "miktar",
            "birim",
            "tutar",
            "kdv",
            "iskonto",
            "para",
        ]
        ws_mov.append(headers)
        for r in data.get("movements") or []:
            ws_mov.append([r.get(h) for h in headers])

        for i, h in enumerate(headers, start=1):
            ws_mov.column_dimensions[get_column_letter(i)].width = min(28, max(12, len(str(h)) + 2))

        wb.save(filepath)

    def export_purchase_report_pdf(self, data: Dict[str, Any], filepath: str, *, printable: bool = False):
        """Satın alma raporunu PDF'e yazar."""
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors

        pagesize = landscape(A4) if printable else A4
        doc = SimpleDocTemplate(filepath, pagesize=pagesize, leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
        font_reg, font_bold = ensure_pdf_fonts()

        styles = getSampleStyleSheet()
        for k in ("Normal", "Title", "Heading1", "Heading2"):
            if k in styles:
                styles[k].fontName = (font_bold if k != "Normal" else font_reg)

        filters = data.get("filters") or {}
        story = []
        story.append(Paragraph("<b>Satın Alma İşlemleri Raporu</b>", styles["Title"]))
        story.append(Spacer(1, 8))
        story.append(Paragraph(f"Tarih Aralığı: {filters.get('date_from') or '-'} → {filters.get('date_to') or '-'}", styles["Normal"]))
        story.append(Spacer(1, 10))

        summary_data = [["Periyot", "Alış", "İade", "Gider", "Ödeme", "Net"]]
        for r in data.get("summary") or []:
            summary_data.append([
                r.get("key"),
                fmt_amount(r.get("alis")),
                fmt_amount(r.get("iade")),
                fmt_amount(r.get("gider")),
                fmt_amount(r.get("odeme")),
                fmt_amount(r.get("net")),
            ])
        tbl = Table(summary_data, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), font_bold),
            ("FONTNAME", (0, 1), (-1, -1), font_reg),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ]))
        story.append(tbl)
        story.append(PageBreak())

        mov_headers = [
            "Tarih",
            "Hareket",
            "Kaynak",
            "Belge",
            "Tedarikçi",
            "Ürün",
            "Kategori",
            "Depo",
            "Ödeme",
            "Miktar",
            "Birim",
            "Tutar",
            "KDV",
            "İskonto",
            "Para",
        ]
        mov_data = [mov_headers]
        for r in data.get("movements") or []:
            mov_data.append([
                r.get("tarih"),
                r.get("hareket"),
                r.get("kaynak"),
                r.get("belge"),
                (r.get("tedarikci") or "")[:18],
                (r.get("urun") or "")[:18],
                (r.get("kategori") or "")[:14],
                (r.get("depo") or "")[:12],
                (r.get("odeme") or "")[:12],
                fmt_amount(r.get("miktar")),
                r.get("birim") or "",
                fmt_amount(r.get("tutar")),
                fmt_amount(r.get("kdv")),
                fmt_amount(r.get("iskonto")),
                r.get("para") or "",
            ])
        mov_tbl = Table(mov_data, repeatRows=1)
        mov_tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), font_bold),
            ("FONTNAME", (0, 1), (-1, -1), font_reg),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("ALIGN", (9, 1), (-1, -1), "RIGHT"),
        ]))
        story.append(mov_tbl)
        doc.build(story)
